
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics
from django.db.models import Sum
from .models import ExcelUpload, BudgetRecord
from .serializers import ExcelUploadSerializer, BudgetRecordSerializer, ExcelFileSerializer
from .utils import auto_correct_records, parse_excel
# from .mappings import REGION_MAPPING, ACTIVITE_MAPPING, FAMILLE_ORDER, get_famille_nom
from .mappings import  ACTIVITE_MAPPING
from .discovery import discover_service
from django.utils import timezone
# External service URLs
SERVICE1_APP = 'AUTHENTICATION-SERVICE'
import requests
import xml.etree.ElementTree as ET
from django.core.exceptions import ValidationError

# ─────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────
def get_service1_url():
    try:
        res = requests.get("http://registry:8761/eureka/apps/AUTHENTICATION-SERVICE", headers={'Accept': 'application/json'})
        instances = res.json()['application']['instance']
        instance = instances[0] if isinstance(instances, list) else instances
        host = instance['hostName']
        port = instance['port']['$']
        return f"http://{host}:{port}"
    except Exception as e:
        print("Error resolving service1 from Eureka:", e)
        return "http://localhost:8001"


def get_service_param_url():
    try:
        print("[DEBUG] Trying to resolve SERVICE-NODE-PARAM from Eureka...")
        res = requests.get(
            "http://localhost:8761/eureka/apps/SERVICE-NODE-PARAM",
            headers={'Accept': 'application/json'},
            timeout=5
        )
        
        if res.status_code == 200:
            data = res.json()
            instances = data['application']['instance']
            instance = instances[0] if isinstance(instances, list) else instances
            host = instance['hostName']
            port = instance['port']['$']
            url = f"http://{host}:{port}"
            print(f"[DEBUG] Service resolved from Eureka: {url}")
            
            # Vérifier si le service est accessible
            try:
                test_resp = requests.get(f"{url}/params/regions", timeout=2)
                if test_resp.status_code == 200:
                    print(f"[DEBUG] Service is accessible on {url}")
                    return url
                else:
                    print(f"[DEBUG] Service on {url} returned {test_resp.status_code}, using fallback")
            except Exception as test_e:
                print(f"[DEBUG] Service on {url} not accessible: {test_e}")
        else:
            print(f"[DEBUG] Eureka returned status {res.status_code}")
            
    except Exception as e:
        print(f"[DEBUG] Error resolving SERVICE-NODE-PARAM from Eureka: {e}")
    
    # Fallback vers le bon port
    print("[DEBUG] Using fallback URL: http://localhost:8083")
    return "http://localhost:8083"

NUMERIC_FIELDS = [
    'cout_initial_total', 'cout_initial_dont_dex',
    'realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex',
    'real_s1_n_total', 'real_s1_n_dont_dex',
    'prev_s2_n_total', 'prev_s2_n_dont_dex',
    'prev_cloture_n_total', 'prev_cloture_n_dont_dex',
    'prev_n_plus1_total', 'prev_n_plus1_dont_dex',
    'reste_a_realiser_total', 'reste_a_realiser_dont_dex',
    'prev_n_plus2_total', 'prev_n_plus2_dont_dex',
    'prev_n_plus3_total', 'prev_n_plus3_dont_dex',
    'prev_n_plus4_total', 'prev_n_plus4_dont_dex',
    'prev_n_plus5_total', 'prev_n_plus5_dont_dex',
    'janvier_total', 'fevrier_total', 'mars_total',
    'avril_total', 'mai_total', 'juin_total',
    'juillet_total', 'aout_total', 'septembre_total',
    'octobre_total', 'novembre_total', 'decembre_total',
]

PREVISION_FIELDS = [
    'prev_s2_n_total', 'prev_s2_n_dont_dex',
    'prev_cloture_n_total', 'prev_cloture_n_dont_dex',
    'prev_n_plus1_total', 'prev_n_plus1_dont_dex',
    'reste_a_realiser_total', 'reste_a_realiser_dont_dex',
    'prev_n_plus2_total', 'prev_n_plus2_dont_dex',
    'prev_n_plus3_total', 'prev_n_plus3_dont_dex',
    'prev_n_plus4_total', 'prev_n_plus4_dont_dex',
    'prev_n_plus5_total', 'prev_n_plus5_dont_dex',
]

SAISIE_FIELDS = [
    'cout_initial_total', 'cout_initial_dont_dex',
    'realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex',
    'real_s1_n_total', 'real_s1_n_dont_dex',
    'janvier_total', 'janvier_dont_dex',
    'fevrier_total', 'fevrier_dont_dex',
    'mars_total', 'mars_dont_dex',
    'avril_total', 'avril_dont_dex',
    'mai_total', 'mai_dont_dex',
    'juin_total', 'juin_dont_dex',
    'juillet_total', 'juillet_dont_dex',
    'aout_total', 'aout_dont_dex',
    'septembre_total', 'septembre_dont_dex',
    'octobre_total', 'octobre_dont_dex',
    'novembre_total', 'novembre_dont_dex',
    'decembre_total', 'decembre_dont_dex',
]

TOTAL_DONT_DEX_PAIRS = [
    ('cout_initial_total',              'cout_initial_dont_dex'),
    ('realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex'),
    ('real_s1_n_total',                 'real_s1_n_dont_dex'),
    ('janvier_total',                   'janvier_dont_dex'),
    ('fevrier_total',                   'fevrier_dont_dex'),
    ('mars_total',                      'mars_dont_dex'),
    ('avril_total',                     'avril_dont_dex'),
    ('mai_total',                       'mai_dont_dex'),
    ('juin_total',                      'juin_dont_dex'),
    ('juillet_total',                   'juillet_dont_dex'),
    ('aout_total',                      'aout_dont_dex'),
    ('septembre_total',                 'septembre_dont_dex'),
    ('octobre_total',                   'octobre_dont_dex'),
    ('novembre_total',                  'novembre_dont_dex'),
    ('decembre_total',                  'decembre_dont_dex'),
]

REGION_EXCLUSIONS   = [None, '', 'Région', 'REGION', 'region', 'Total', 'TOTAL']
FAMILLE_EXCLUSIONS  = [None, '', 'Famille', 'FAMILLE', 'famille', 'Total', 'TOTAL']
ACTIVITE_EXCLUSIONS = [None, '', 'Activité', 'ACTIVITE', 'activite', 'Total', 'TOTAL']


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def build_aggregation():
    return {f: Sum(f) for f in NUMERIC_FIELDS}


def apply_mapping(data, key_field, mapping):
    result = []
    for row in data:
        code = str(row.get(key_field, '') or '').strip()
        row[key_field + '_code'] = code
        row[key_field + '_nom'] = mapping.get(code, code)
        result.append(row)
    return result


def clean_queryset(qs):
    return qs.exclude(activite__isnull=True)\
             .exclude(activite__in=ACTIVITE_EXCLUSIONS)\
             .exclude(region__isnull=True)\
             .exclude(region__in=REGION_EXCLUSIONS)\
             .exclude(famille__isnull=True)\
             .exclude(famille__in=FAMILLE_EXCLUSIONS)


# def group_by_famille(data):
#     grouped = {}

#     for row in data:
#         code = str(row.get('famille', '') or '').strip()
#         if not code:
#             continue

#         nom = get_famille_nom(code)

#         if nom not in grouped:
#             grouped[nom] = {field: 0 for field in NUMERIC_FIELDS}
#             grouped[nom]['famille_nom'] = nom

#         for field in NUMERIC_FIELDS:
#             val = row.get(field) or 0
#             grouped[nom][field] += float(val)

#     return sorted(
#         grouped.values(),
#         key=lambda x: FAMILLE_ORDER.index(x['famille_nom'])
#         if x['famille_nom'] in FAMILLE_ORDER else 99
#     )


# ─────────────────────────────────────────
# UPLOAD
# ─────────────────────────────────────────
from rest_framework.permissions import IsAuthenticated
from .permissions import *
from .remote_auth import RemoteJWTAuthentication


class ExcelUploadView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAgent]

    def post(self, request):
        serializer = ExcelFileSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        file = serializer.validated_data['file']

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Only Excel files allowed'}, status=400)

        upload = ExcelUpload.objects.create(file_name=file.name, status='pending')

        try:
            count = parse_excel(file, upload)
            qs = clean_queryset(BudgetRecord.objects.filter(upload=upload))
            corrected_count = auto_correct_records(qs)

            upload.status = 'processed'
            upload.save()

            return Response({
                'message': f'{count} records importés',
                'upload_id': upload.id,
                'corrections': {
                    'records_corrigés': corrected_count,
                    'message': (
                        f'{corrected_count} record(s) corrigé(s) automatiquement'
                        if corrected_count
                        else 'Aucune correction nécessaire'
                    )
                }
            }, status=201)

        except Exception as e:
            upload.status = 'failed'
            upload.save()
            return Response({'error': str(e)}, status=500)


class UploadListView(generics.ListAPIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAgent]

    queryset = ExcelUpload.objects.all().order_by('-uploaded_at')
    serializer_class = ExcelUploadSerializer


class BudgetRecordListView(generics.ListAPIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetRecordSerializer

    def get_queryset(self):
        user = self.request.user
        role = getattr(user, 'role', None)
        structure_id = getattr(user, 'structure_id', None)
        region_id = getattr(user, 'region_id', None)
        
        qs = BudgetRecord.objects.all()
        
        if role == 'responsable_structure':
            # Responsable structure → voit sa structure
            if structure_id:
                qs = qs.filter(structure_id=structure_id)
            else:
                return BudgetRecord.objects.none()
        elif role == 'directeur_region':
            # Directeur région → voit sa région
            if region_id:
                qs = qs.filter(region_id=region_id)
            else:
                return BudgetRecord.objects.none()
        elif role == 'agent':
            # Agent → voit ses projets
            qs = qs.filter(created_by=user.id)
        # Chef, directeur, divisionnaire → voient tout
        
        uid = self.request.query_params.get('upload_id')
        if uid:
            qs = qs.filter(upload_id=uid)
        
        return qs.order_by('-id')

#voir plus
class GetProjetByIdView(APIView):
    """
    GET /recap/budget/projet/<id>/
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        try:
            projet = BudgetRecord.objects.get(id=id)
            serializer = BudgetRecordSerializer(projet, context={'request': request})
            return Response({
                'success': True,
                'data': serializer.data
            })
        except BudgetRecord.DoesNotExist:
            return Response(
                {'error': f'Projet avec ID {id} non trouvé'},
                status=404
            )
# ─────────────────────────────────────────
# RECAPS
# ─────────────────────────────────────────
# class RecapParRegionView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('region')
#             .annotate(**build_aggregation())
#             .order_by('region')
#         )

#         total = qs.aggregate(**build_aggregation())

#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')
        
#         result = []
#         for row in data:
#             code = str(row.get('region', '') or '').strip()
            
#             # Récupérer le nom de la région
#             region_name = code  # fallback
#             if code and code not in ['', '-', 'None']:
#                 try:
#                     url = f"{service_url}/params/regions/{code}"
#                     response = requests.get(url, headers={'Authorization': token}, timeout=5)
#                     if response.status_code == 200:
#                         region_data = response.json().get('data', {})
#                         region_name = region_data.get('nom_region', code)
#                         print(f"[DEBUG] Region {code} -> {region_name}")
#                 except Exception as e:
#                     print(f"[DEBUG] Error: {e}")
            
#             # Remplacer la clé 'region' par le nom
#             row['region'] = region_name
            
#             result.append(row)

#         return Response({
#             "regions": result,
#             "total_division": total
#         })

        

# class RecapParFamilleView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         print("=" * 80)
#         print("[DEBUG] === Début RecapParFamilleView ===")
#         print("=" * 80)
        
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('famille')
#             .annotate(**build_aggregation())
#             .order_by('famille')
#         )

#         total = qs.aggregate(**build_aggregation())

#         service_url = get_service_param_url()
#         print(f"[DEBUG] Service URL: {service_url}")
        
#         # Récupérer le token de l'utilisateur actuel
#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
#         print(f"[DEBUG] Token present: {bool(token)}")
        
#         # Récupérer TOUTES les familles en UNE SEULE requête avec le token
#         famille_mapping = {}
#         try:
#             print("[DEBUG] Fetching all families from service with token...")
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             print(f"[DEBUG] Response status: {response.status_code}")
            
#             if response.status_code == 200:
#                 response_data = response.json()
#                 print(f"[DEBUG] Response keys: {response_data.keys() if isinstance(response_data, dict) else 'list'}")
                
#                 # Gérer différentes structures de réponse
#                 if isinstance(response_data, dict):
#                     if 'data' in response_data:
#                         familles_list = response_data['data']
#                     else:
#                         familles_list = [response_data] if response_data else []
#                 elif isinstance(response_data, list):
#                     familles_list = response_data
#                 else:
#                     familles_list = []
                
#                 print(f"[DEBUG] Number of families found: {len(familles_list)}")
                
#                 for famille in familles_list:
#                     if isinstance(famille, dict):
#                         code = famille.get('code_famille') or famille.get('code') or famille.get('id')
#                         nom = famille.get('nom_famille') or famille.get('name') or famille.get('libelle') or famille.get('nom')
#                         if code:
#                             famille_mapping[str(code)] = nom
#                             print(f"[DEBUG] Mapped {code} -> {nom}")
                
#                 print(f"[DEBUG] Total mapped families: {len(famille_mapping)}")
#             else:
#                 print(f"[DEBUG] Failed to fetch families: {response.status_code}")
#                 if response.text:
#                     print(f"[DEBUG] Response: {response.text[:200]}")
                    
#         except Exception as e:
#             print(f"[DEBUG] Error fetching families: {e}")
#             import traceback
#             traceback.print_exc()
        
#         # Appliquer le mapping
#         result = []
#         for row in data:
#             code = str(row.get('famille', '') or '').strip()
#             if code and code not in ['', '-', 'None']:
#                 row['famille'] = famille_mapping.get(code, code)
#             else:
#                 row['famille'] = '-'
#             print(f"[DEBUG] Mapped: {code} -> {row['famille']}")
#             result.append(row)

#         print("=" * 80)
#         print("[DEBUG] === Fin RecapParFamilleView ===")
#         print("=" * 80)

#         return Response({
#             "familles": result,
#             "total_division_production": total
#         })
# class RecapParActiviteView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]

#     def get(self, request):
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('activite')
#             .annotate(**build_aggregation())
#             .order_by('activite')
#         )

#         total_qs = qs.aggregate(**build_aggregation())

#         return Response({
#             "activites": apply_mapping(data, 'activite', ACTIVITE_MAPPING),
#             "total_division": total_qs,
#         })


# class RecapGlobalView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]

#     def get(self, request):
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)
#         agg = build_aggregation()
        
#         # Récupérer le token
#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
#         service_url = get_service_param_url()
        
#         # Récupérer toutes les régions
#         region_mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
            
#             if response.status_code == 200:
#                 response_data = response.json()
#                 regions_list = response_data.get('data', [])
#                 for region in regions_list:
#                     code = region.get('code_region')
#                     nom = region.get('nom_region')
#                     if code:
#                         region_mapping[str(code)] = nom
#         except Exception:
#             pass
        
#         # Récupérer toutes les familles
#         famille_mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
            
#             if response.status_code == 200:
#                 response_data = response.json()
#                 familles_list = response_data.get('data', [])
#                 for famille in familles_list:
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     if code:
#                         famille_mapping[str(code)] = nom
#         except Exception:
#             pass
        
#         # Traiter les régions
#         regions_data = list(qs.values('region').annotate(**agg))
#         regions_result = []
#         for row in regions_data:
#             code = str(row.get('region', '') or '').strip()
#             region_nom = region_mapping.get(code, code) if code and code not in ['', '-', 'None'] else '-'
#             regions_result.append({
#                 'region_code': code,
#                 'region_nom': region_nom,
#                 **{k: v for k, v in row.items() if k != 'region'}
#             })
        
#         # Traiter les familles
#         familles_data = list(qs.values('famille').annotate(**agg))
#         familles_result = []
#         for row in familles_data:
#             code = str(row.get('famille', '') or '').strip()
#             famille_nom = famille_mapping.get(code, code) if code and code not in ['', '-', 'None'] else '-'
#             familles_result.append({
#                 'famille_code': code,
#                 'famille_nom': famille_nom,
#                 **{k: v for k, v in row.items() if k != 'famille'}
#             })
        
#         # Traiter les activités
#         activites_data = list(qs.values('activite').annotate(**agg))
#         activites_result = []
#         for row in activites_data:
#             code = str(row.get('activite', '') or '').strip()
#             activite_nom = ACTIVITE_MAPPING.get(code, code)
#             activites_result.append({
#                 'activite_code': code,
#                 'activite_nom': activite_nom,
#                 **{k: v for k, v in row.items() if k != 'activite'}
#             })

#         return Response({
#             'par_region': regions_result,
#             'par_famille': familles_result,
#             'par_activite': activites_result,
#         })
    



# class RecapFamilleParActiviteView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     @staticmethod
#     def _get_familles_ordered(token):
#         """Récupère les familles dans l'ordre défini par le backend"""
#         service_url = get_service_param_url()
#         famille_mapping = {}
#         famille_order = []
        
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
            
#             if response.status_code == 200:
#                 response_data = response.json()
#                 familles_list = response_data.get('data', [])
                
#                 # Créer le mapping code -> nom
#                 # et la liste ordonnée des noms
#                 for famille in familles_list:
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     order = famille.get('ordre', 999)  # Récupérer l'ordre depuis l'API
                    
#                     if code:
#                         famille_mapping[str(code)] = nom
                    
#                     # Ajouter à la liste ordonnée
#                     famille_order.append({
#                         'code': code,
#                         'nom': nom,
#                         'ordre': order
#                     })
                
#                 # Trier par l'ordre défini par le backend
#                 famille_order.sort(key=lambda x: x['ordre'])
                
#         except Exception as e:
#             print(f"Erreur récupération familles: {e}")
        
#         return famille_mapping, famille_order

#     def get(self, request):
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('activite', 'famille')
#             .annotate(**build_aggregation())
#             .order_by('activite', 'famille')
#         )

#         # Récupérer le token
#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
        
#         # Récupérer les familles avec leur ordre
#         famille_mapping, famille_order_list = self._get_familles_ordered(token)
        
#         # Créer un dictionnaire pour l'ordre
#         famille_order_dict = {}
#         for idx, fam in enumerate(famille_order_list):
#             famille_order_dict[fam['nom']] = idx
#             famille_order_dict[fam['code']] = idx

#         activites = {}

#         for row in data:
#             act_code = str(row.get('activite') or '').strip()
#             act_nom = ACTIVITE_MAPPING.get(act_code, act_code)
#             fam_code = str(row.get('famille') or '').strip()
#             fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code and fam_code not in ['', '-', 'None'] else '-'

#             if act_code not in activites:
#                 activites[act_code] = {
#                     'activite_code': act_code,
#                     'activite_nom': act_nom,
#                     'familles': {},
#                     'total': {f: 0 for f in NUMERIC_FIELDS},
#                 }

#             if fam_code not in activites[act_code]['familles']:
#                 activites[act_code]['familles'][fam_code] = {
#                     'famille_code': fam_code,
#                     'famille_nom': fam_nom,
#                     **{f: 0 for f in NUMERIC_FIELDS}
#                 }

#             for field in NUMERIC_FIELDS:
#                 val = float(row.get(field) or 0)
#                 activites[act_code]['familles'][fam_code][field] += val
#                 activites[act_code]['total'][field] += val

#         result = []
#         total_global = {f: 0 for f in NUMERIC_FIELDS}

#         for act in activites.values():
#             # ⭐ Trier selon l'ordre du backend
#             familles = sorted(
#                 act['familles'].values(),
#                 key=lambda x: famille_order_dict.get(x['famille_nom'], 999)
#             )

#             result.append({
#                 'activite_code': act['activite_code'],
#                 'activite_nom': act['activite_nom'],
#                 'familles': familles,
#                 'total_activite': act['total'],
#             })

#             for f in NUMERIC_FIELDS:
#                 total_global[f] += act['total'][f]

#         return Response({
#             'detail': result,
#             'total_global': total_global,
#         })

# class RecapRegionFamilleView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]
#     """
#     GET /api/recap/region-famille/?upload_id=1
#     Retourne chaque région avec ses familles + total par région
#     """

#     def get(self, request):
#         qs = BudgetRecord.objects.all()

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('region', 'famille')
#             .annotate(**build_aggregation())
#             .order_by('region', 'famille')
#         )

#         # Récupérer le token
#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
#         service_url = get_service_param_url()
        
#         # Récupérer toutes les régions et familles depuis l'API
#         headers = {'Authorization': f'Bearer {token}'} if token else {}
        
#         region_mapping = {}
#         famille_mapping = {}
#         famille_order = []  # Pour stocker l'ordre des familles
        
#         try:
#             # Récupérer les régions
#             region_response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
#             if region_response.status_code == 200:
#                 regions_data = region_response.json().get('data', [])
#                 for region in regions_data:
#                     code = region.get('code_region')
#                     nom = region.get('nom_region')
#                     if code:
#                         region_mapping[str(code)] = nom
            
#             # Récupérer les familles
#             famille_response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if famille_response.status_code == 200:
#                 familles_data = famille_response.json().get('data', [])
#                 for famille in familles_data:
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     ordre = famille.get('ordre', 999)  # Récupérer l'ordre
#                     if code:
#                         famille_mapping[str(code)] = nom
#                         famille_order.append({
#                             'code': code,
#                             'nom': nom,
#                             'ordre': ordre
#                         })
                
#                 # Trier les familles par ordre défini par l'API
#                 famille_order.sort(key=lambda x: x['ordre'])
                
#         except Exception as e:
#             print(f"Erreur lors de l'appel au service param: {e}")

#         # Créer un dictionnaire pour l'ordre des familles
#         famille_ordre_dict = {}
#         for idx, fam in enumerate(famille_order):
#             famille_ordre_dict[fam['nom']] = idx
#             famille_ordre_dict[fam['code']] = idx

#         regions = {}

#         for row in data:
#             reg_code = str(row.get('region') or '').strip()
#             # Récupérer le nom depuis l'API ou utiliser le code par défaut
#             reg_nom = region_mapping.get(reg_code, reg_code) if reg_code and reg_code not in ['', '-', 'None'] else '-'
            
#             fam_code = str(row.get('famille') or '').strip()
#             fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code and fam_code not in ['', '-', 'None'] else '-'

#             if reg_code not in regions:
#                 regions[reg_code] = {
#                     'region_code': reg_code,
#                     'region_nom': reg_nom,
#                     'familles': {},
#                     'total': {f: 0 for f in NUMERIC_FIELDS},
#                 }

#             if fam_code not in regions[reg_code]['familles']:
#                 regions[reg_code]['familles'][fam_code] = {
#                     'famille_code': fam_code,
#                     'famille_nom': fam_nom,
#                     **{f: 0 for f in NUMERIC_FIELDS}
#                 }

#             for field in NUMERIC_FIELDS:
#                 val = float(row.get(field) or 0)
#                 regions[reg_code]['familles'][fam_code][field] += val
#                 regions[reg_code]['total'][field] += val

#         result = []
#         total_global = {f: 0 for f in NUMERIC_FIELDS}

#         for reg_code, reg_data in sorted(regions.items()):
#             # Trier les familles selon l'ordre défini par l'API
#             familles_triees = sorted(
#                 reg_data['familles'].values(),
#                 key=lambda x: famille_ordre_dict.get(x['famille_nom'], 999)
#             )

#             result.append({
#                 'region_code': reg_data['region_code'],
#                 'region_nom': reg_data['region_nom'],
#                 'familles': familles_triees,
#                 'total_region': reg_data['total'],
#             })

#             for f in NUMERIC_FIELDS:
#                 total_global[f] += reg_data['total'][f]

#         return Response({
#             'detail': result,
#             'total_global': total_global,
#         })
# class RecapParFamilleView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         next_year = datetime.now().year + 1
        
#         qs = BudgetRecord.objects.filter(annee_debut_pmt=next_year)

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('famille')
#             .annotate(**build_aggregation())
#             .order_by('famille')
#         )

#         total = qs.aggregate(**build_aggregation())

#         service_url = get_service_param_url()
#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''

#         famille_mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
            
#             if response.status_code == 200:
#                 response_data = response.json()
                
#                 if isinstance(response_data, dict):
#                     familles_list = response_data.get('data', [])
#                 elif isinstance(response_data, list):
#                     familles_list = response_data
#                 else:
#                     familles_list = []
                
#                 for famille in familles_list:
#                     if isinstance(famille, dict):
#                         code = famille.get('code_famille') or famille.get('code') or famille.get('id')
#                         nom = famille.get('nom_famille') or famille.get('name') or famille.get('libelle') or famille.get('nom')
#                         if code:
#                             famille_mapping[str(code)] = nom
                            
#         except Exception as e:
#             print(f"Error fetching families: {e}")

#         result = []
#         for row in data:
#             code = str(row.get('famille', '') or '').strip()
#             row['famille'] = famille_mapping.get(code, code) if code and code not in ['', '-', 'None'] else '-'
#             result.append(row)

#         return Response({
#             "familles": result,
#             "total_division_production": total,
#             "annee_filtre": next_year,  # utile pour debug côté frontend
#         })
# class RecapParActiviteView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]

#     def get(self, request):
#         next_year = datetime.now().year + 1

#         qs = BudgetRecord.objects.filter(annee_debut_pmt=next_year)

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('activite')
#             .annotate(**build_aggregation())
#             .order_by('activite')
#         )

#         total_qs = qs.aggregate(**build_aggregation())

#         return Response({
#             "activites": apply_mapping(data, 'activite', ACTIVITE_MAPPING),
#             "total_division": total_qs,
#             "annee_filtre": next_year,
#         })
# class RecapGlobalView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]

#     def get(self, request):
#         next_year = datetime.now().year + 1

#         qs = BudgetRecord.objects.filter(annee_debut_pmt=next_year)

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)
#         agg = build_aggregation()

#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
#         service_url = get_service_param_url()

#         region_mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 for region in response.json().get('data', []):
#                     code = region.get('code_region')
#                     nom = region.get('nom_region')
#                     if code:
#                         region_mapping[str(code)] = nom
#         except Exception:
#             pass

#         famille_mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 for famille in response.json().get('data', []):
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     if code:
#                         famille_mapping[str(code)] = nom
#         except Exception:
#             pass

#         regions_result = []
#         for row in list(qs.values('region').annotate(**agg)):
#             code = str(row.get('region', '') or '').strip()
#             regions_result.append({
#                 'region_code': code,
#                 'region_nom': region_mapping.get(code, code) if code not in ['', '-', 'None'] else '-',
#                 **{k: v for k, v in row.items() if k != 'region'}
#             })

#         familles_result = []
#         for row in list(qs.values('famille').annotate(**agg)):
#             code = str(row.get('famille', '') or '').strip()
#             familles_result.append({
#                 'famille_code': code,
#                 'famille_nom': famille_mapping.get(code, code) if code not in ['', '-', 'None'] else '-',
#                 **{k: v for k, v in row.items() if k != 'famille'}
#             })

#         activites_result = []
#         for row in list(qs.values('activite').annotate(**agg)):
#             code = str(row.get('activite', '') or '').strip()
#             activites_result.append({
#                 'activite_code': code,
#                 'activite_nom': ACTIVITE_MAPPING.get(code, code),
#                 **{k: v for k, v in row.items() if k != 'activite'}
#             })

#         return Response({
#             'par_region': regions_result,
#             'par_famille': familles_result,
#             'par_activite': activites_result,
#             'annee_filtre': next_year,
#         })
    
# class RecapFamilleParActiviteView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     @staticmethod
#     def _get_familles_ordered(token):
#         service_url = get_service_param_url()
#         famille_mapping = {}
#         famille_order = []
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 for famille in response.json().get('data', []):
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     order = famille.get('ordre', 999)
#                     if code:
#                         famille_mapping[str(code)] = nom
#                         famille_order.append({'code': code, 'nom': nom, 'ordre': order})
#                 famille_order.sort(key=lambda x: x['ordre'])
#         except Exception as e:
#             print(f"Erreur récupération familles: {e}")
#         return famille_mapping, famille_order

#     def get(self, request):
#         next_year = datetime.now().year + 1

#         qs = BudgetRecord.objects.filter(annee_debut_pmt=next_year)

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('activite', 'famille')
#             .annotate(**build_aggregation())
#             .order_by('activite', 'famille')
#         )

#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''

#         famille_mapping, famille_order_list = self._get_familles_ordered(token)

#         famille_order_dict = {}
#         for idx, fam in enumerate(famille_order_list):
#             famille_order_dict[fam['nom']] = idx
#             famille_order_dict[fam['code']] = idx

#         activites = {}
#         for row in data:
#             act_code = str(row.get('activite') or '').strip()
#             act_nom = ACTIVITE_MAPPING.get(act_code, act_code)
#             fam_code = str(row.get('famille') or '').strip()
#             fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code not in ['', '-', 'None'] else '-'

#             if act_code not in activites:
#                 activites[act_code] = {
#                     'activite_code': act_code,
#                     'activite_nom': act_nom,
#                     'familles': {},
#                     'total': {f: 0 for f in NUMERIC_FIELDS},
#                 }

#             if fam_code not in activites[act_code]['familles']:
#                 activites[act_code]['familles'][fam_code] = {
#                     'famille_code': fam_code,
#                     'famille_nom': fam_nom,
#                     **{f: 0 for f in NUMERIC_FIELDS}
#                 }

#             for field in NUMERIC_FIELDS:
#                 val = float(row.get(field) or 0)
#                 activites[act_code]['familles'][fam_code][field] += val
#                 activites[act_code]['total'][field] += val

#         result = []
#         total_global = {f: 0 for f in NUMERIC_FIELDS}

#         for act in activites.values():
#             familles = sorted(
#                 act['familles'].values(),
#                 key=lambda x: famille_order_dict.get(x['famille_nom'], 999)
#             )
#             result.append({
#                 'activite_code': act['activite_code'],
#                 'activite_nom': act['activite_nom'],
#                 'familles': familles,
#                 'total_activite': act['total'],
#             })
#             for f in NUMERIC_FIELDS:
#                 total_global[f] += act['total'][f]

#         return Response({
#             'detail': result,
#             'total_global': total_global,
#             'annee_filtre': next_year,
#         })
# class RecapRegionFamilleView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAll]

#     def get(self, request):
#         next_year = datetime.now().year + 1

#         qs = BudgetRecord.objects.filter(annee_debut_pmt=next_year)

#         uid = request.query_params.get('upload_id')
#         if uid:
#             qs = qs.filter(upload_id=uid)

#         qs = clean_queryset(qs)

#         data = list(
#             qs.values('region', 'famille')
#             .annotate(**build_aggregation())
#             .order_by('region', 'famille')
#         )

#         auth_header = request.headers.get('Authorization', '')
#         token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''
#         service_url = get_service_param_url()
#         headers = {'Authorization': f'Bearer {token}'} if token else {}

#         region_mapping = {}
#         famille_mapping = {}
#         famille_order = []

#         try:
#             region_response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
#             if region_response.status_code == 200:
#                 for region in region_response.json().get('data', []):
#                     code = region.get('code_region')
#                     if code:
#                         region_mapping[str(code)] = region.get('nom_region')

#             famille_response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if famille_response.status_code == 200:
#                 for famille in famille_response.json().get('data', []):
#                     code = famille.get('code_famille')
#                     nom = famille.get('nom_famille')
#                     ordre = famille.get('ordre', 999)
#                     if code:
#                         famille_mapping[str(code)] = nom
#                         famille_order.append({'code': code, 'nom': nom, 'ordre': ordre})
#                 famille_order.sort(key=lambda x: x['ordre'])

#         except Exception as e:
#             print(f"Erreur service param: {e}")

#         famille_ordre_dict = {}
#         for idx, fam in enumerate(famille_order):
#             famille_ordre_dict[fam['nom']] = idx
#             famille_ordre_dict[fam['code']] = idx

#         regions = {}
#         for row in data:
#             reg_code = str(row.get('region') or '').strip()
#             reg_nom = region_mapping.get(reg_code, reg_code) if reg_code not in ['', '-', 'None'] else '-'
#             fam_code = str(row.get('famille') or '').strip()
#             fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code not in ['', '-', 'None'] else '-'

#             if reg_code not in regions:
#                 regions[reg_code] = {
#                     'region_code': reg_code,
#                     'region_nom': reg_nom,
#                     'familles': {},
#                     'total': {f: 0 for f in NUMERIC_FIELDS},
#                 }

#             if fam_code not in regions[reg_code]['familles']:
#                 regions[reg_code]['familles'][fam_code] = {
#                     'famille_code': fam_code,
#                     'famille_nom': fam_nom,
#                     **{f: 0 for f in NUMERIC_FIELDS}
#                 }

#             for field in NUMERIC_FIELDS:
#                 val = float(row.get(field) or 0)
#                 regions[reg_code]['familles'][fam_code][field] += val
#                 regions[reg_code]['total'][field] += val

#         result = []
#         total_global = {f: 0 for f in NUMERIC_FIELDS}

#         for reg_code, reg_data in sorted(regions.items()):
#             familles_triees = sorted(
#                 reg_data['familles'].values(),
#                 key=lambda x: famille_ordre_dict.get(x['famille_nom'], 999)
#             )
#             result.append({
#                 'region_code': reg_data['region_code'],
#                 'region_nom': reg_data['region_nom'],
#                 'familles': familles_triees,
#                 'total_region': reg_data['total'],
#             })
#             for f in NUMERIC_FIELDS:
#                 total_global[f] += reg_data['total'][f]

#         return Response({
#             'detail': result,
#             'total_global': total_global,
#             'annee_filtre': next_year,
#         })
from datetime import datetime
import requests
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated


def _get_token(request):
    auth_header = request.headers.get('Authorization', '')
    return auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''


def _get_region_mapping(token):
    service_url = get_service_param_url()
    region_mapping = {}
    try:
        headers = {'Authorization': f'Bearer {token}'} if token else {}
        response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
        if response.status_code == 200:
            for region in response.json().get('data', []):
                code = region.get('code_region')
                if code:
                    region_mapping[str(code)] = region.get('nom_region')
    except Exception as e:
        print(f"Error fetching regions: {e}")
    return region_mapping


def _get_famille_mapping(token):
    service_url = get_service_param_url()
    famille_mapping = {}
    famille_order = []
    try:
        headers = {'Authorization': f'Bearer {token}'} if token else {}
        response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
        if response.status_code == 200:
            for famille in response.json().get('data', []):
                code = famille.get('code_famille')
                nom = famille.get('nom_famille')
                ordre = famille.get('ordre', 999)
                if code:
                    famille_mapping[str(code)] = nom
                    famille_order.append({'code': code, 'nom': nom, 'ordre': ordre})
            famille_order.sort(key=lambda x: x['ordre'])
    except Exception as e:
        print(f"Error fetching familles: {e}")
    return famille_mapping, famille_order


def _base_qs():
    next_year = datetime.now().year + 1
    return BudgetRecord.objects.filter(
        annee_debut_pmt=next_year,
        # statut='valide_divisionnaire'
    )


# ──────────────────────────────────────────────────────────────────────────────

class RecapParRegionView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())

        data = list(
            qs.values('region')
            .annotate(**build_aggregation())
            .order_by('region')
        )
        total = qs.aggregate(**build_aggregation())

        token = _get_token(request)
        region_mapping, _ = _get_region_mapping(token), None
        region_mapping = _get_region_mapping(token)

        result = []
        for row in data:
            code = str(row.get('region', '') or '').strip()
            row['region'] = region_mapping.get(code, code) if code and code not in ['', '-', 'None'] else '-'
            result.append(row)

        return Response({
            "regions": result,
            "total_division": total,
        })


class RecapParFamilleView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())

        data = list(
            qs.values('famille')
            .annotate(**build_aggregation())
            .order_by('famille')
        )
        total = qs.aggregate(**build_aggregation())

        token = _get_token(request)
        famille_mapping, _ = _get_famille_mapping(token)

        result = []
        for row in data:
            code = str(row.get('famille', '') or '').strip()
            row['famille'] = famille_mapping.get(code, code) if code and code not in ['', '-', 'None'] else '-'
            result.append(row)

        return Response({
            "familles": result,
            "total_division_production": total,
        })


class RecapParActiviteView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())

        data = list(
            qs.values('activite')
            .annotate(**build_aggregation())
            .order_by('activite')
        )
        total_qs = qs.aggregate(**build_aggregation())

        return Response({
            "activites": apply_mapping(data, 'activite', ACTIVITE_MAPPING),
            "total_division": total_qs,
        })


class RecapGlobalView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())
        agg = build_aggregation()

        token = _get_token(request)
        region_mapping = _get_region_mapping(token)
        famille_mapping, _ = _get_famille_mapping(token)

        regions_result = []
        for row in list(qs.values('region').annotate(**agg)):
            code = str(row.get('region', '') or '').strip()
            regions_result.append({
                'region_code': code,
                'region_nom': region_mapping.get(code, code) if code not in ['', '-', 'None'] else '-',
                **{k: v for k, v in row.items() if k != 'region'}
            })

        familles_result = []
        for row in list(qs.values('famille').annotate(**agg)):
            code = str(row.get('famille', '') or '').strip()
            familles_result.append({
                'famille_code': code,
                'famille_nom': famille_mapping.get(code, code) if code not in ['', '-', 'None'] else '-',
                **{k: v for k, v in row.items() if k != 'famille'}
            })

        activites_result = []
        for row in list(qs.values('activite').annotate(**agg)):
            code = str(row.get('activite', '') or '').strip()
            activites_result.append({
                'activite_code': code,
                'activite_nom': ACTIVITE_MAPPING.get(code, code),
                **{k: v for k, v in row.items() if k != 'activite'}
            })

        return Response({
            'par_region': regions_result,
            'par_famille': familles_result,
            'par_activite': activites_result,
        })


class RecapFamilleParActiviteView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())

        data = list(
            qs.values('activite', 'famille')
            .annotate(**build_aggregation())
            .order_by('activite', 'famille')
        )

        token = _get_token(request)
        famille_mapping, famille_order_list = _get_famille_mapping(token)

        famille_order_dict = {}
        for idx, fam in enumerate(famille_order_list):
            famille_order_dict[fam['nom']] = idx
            famille_order_dict[fam['code']] = idx

        activites = {}
        for row in data:
            act_code = str(row.get('activite') or '').strip()
            act_nom = ACTIVITE_MAPPING.get(act_code, act_code)
            fam_code = str(row.get('famille') or '').strip()
            fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code not in ['', '-', 'None'] else '-'

            if act_code not in activites:
                activites[act_code] = {
                    'activite_code': act_code,
                    'activite_nom': act_nom,
                    'familles': {},
                    'total': {f: 0 for f in NUMERIC_FIELDS},
                }
            if fam_code not in activites[act_code]['familles']:
                activites[act_code]['familles'][fam_code] = {
                    'famille_code': fam_code,
                    'famille_nom': fam_nom,
                    **{f: 0 for f in NUMERIC_FIELDS}
                }
            for field in NUMERIC_FIELDS:
                val = float(row.get(field) or 0)
                activites[act_code]['familles'][fam_code][field] += val
                activites[act_code]['total'][field] += val

        result = []
        total_global = {f: 0 for f in NUMERIC_FIELDS}

        for act in activites.values():
            familles = sorted(
                act['familles'].values(),
                key=lambda x: famille_order_dict.get(x['famille_nom'], 999)
            )
            result.append({
                'activite_code': act['activite_code'],
                'activite_nom': act['activite_nom'],
                'familles': familles,
                'total_activite': act['total'],
            })
            for f in NUMERIC_FIELDS:
                total_global[f] += act['total'][f]

        return Response({
            'detail': result,
            'total_global': total_global,
        })


class RecapRegionFamilleView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = clean_queryset(_base_qs())

        data = list(
            qs.values('region', 'famille')
            .annotate(**build_aggregation())
            .order_by('region', 'famille')
        )

        token = _get_token(request)
        region_mapping = _get_region_mapping(token)
        famille_mapping, famille_order = _get_famille_mapping(token)

        famille_ordre_dict = {}
        for idx, fam in enumerate(famille_order):
            famille_ordre_dict[fam['nom']] = idx
            famille_ordre_dict[fam['code']] = idx

        regions = {}
        for row in data:
            reg_code = str(row.get('region') or '').strip()
            reg_nom = region_mapping.get(reg_code, reg_code) if reg_code not in ['', '-', 'None'] else '-'
            fam_code = str(row.get('famille') or '').strip()
            fam_nom = famille_mapping.get(fam_code, fam_code) if fam_code not in ['', '-', 'None'] else '-'

            if reg_code not in regions:
                regions[reg_code] = {
                    'region_code': reg_code,
                    'region_nom': reg_nom,
                    'familles': {},
                    'total': {f: 0 for f in NUMERIC_FIELDS},
                }
            if fam_code not in regions[reg_code]['familles']:
                regions[reg_code]['familles'][fam_code] = {
                    'famille_code': fam_code,
                    'famille_nom': fam_nom,
                    **{f: 0 for f in NUMERIC_FIELDS}
                }
            for field in NUMERIC_FIELDS:
                val = float(row.get(field) or 0)
                regions[reg_code]['familles'][fam_code][field] += val
                regions[reg_code]['total'][field] += val

        result = []
        total_global = {f: 0 for f in NUMERIC_FIELDS}

        for reg_code, reg_data in sorted(regions.items()):
            familles_triees = sorted(
                reg_data['familles'].values(),
                key=lambda x: famille_ordre_dict.get(x['famille_nom'], 999)
            )
            result.append({
                'region_code': reg_data['region_code'],
                'region_nom': reg_data['region_nom'],
                'familles': familles_triees,
                'total_region': reg_data['total'],
            })
            for f in NUMERIC_FIELDS:
                total_global[f] += reg_data['total'][f]

        return Response({
            'detail': result,
            'total_global': total_global,
        })
# ─────────────────────────────────────────
# PDF
# ─────────────────────────────────────────
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime

# class BudgetRecordPDFView(APIView):
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]

#     def get(self, request, pk):
#         record = get_object_or_404(BudgetRecord, pk=pk)

#         # Récupérer l'année de début PMT
#         annee_debut = record.annee_debut_pmt
        
#         # Si l'année n'est pas définie, utiliser l'année courante + 1 comme fallback
#         current_year = datetime.now().year
#         if not annee_debut:
#             annee_debut = current_year + 1
            
#         # N = année de début PMT - 1
#         N = annee_debut - 1
#         N_plus_1 = N + 1
#         N_plus_2 = N + 2
#         N_plus_3 = N + 3
#         N_plus_4 = N + 4
#         N_plus_5 = N + 5
        
#         # Les années de la période PMT
#         annee_fin = record.annee_fin_pmt if record.annee_fin_pmt else N_plus_5

#         activite = ACTIVITE_MAPPING.get(record.activite, record.activite or '-')
#         region   = REGION_MAPPING.get(record.region, record.region or '-')
#         famille  = get_famille_nom(record.famille or '-')

#         buffer = BytesIO()
#         doc    = SimpleDocTemplate(buffer, pagesize=A4)
#         styles = getSampleStyleSheet()
#         elements = []

#         elements.append(Paragraph("RAPPORT BUDGET", styles['Title']))
#         elements.append(Spacer(1, 10))

#         info_data = [
#             ["Activité",  activite],
#             ["Région",    region],
#             ["Famille",   famille],
#             ["Libellé",   record.libelle or '-'],
#             ["Période PMT", f"{annee_debut} - {annee_fin}"]
#         ]

#         info_table = Table(info_data, colWidths=[120, 350])
#         info_table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
#             ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
#             ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica'),
#         ]))
#         elements.append(info_table)
#         elements.append(Spacer(1, 15))

#         def v(val):
#             return val if val is not None else 0

#         main_data = [
#             ["Désignation", "Total", "Dont DEX"],
#             [f"Coût Global Initial PMT {annee_debut}/{annee_fin}", v(record.cout_initial_total), v(record.cout_initial_dont_dex)],
#             [f"Réalisations Cumulées à fin {N} au coût réel", v(record.realisation_cumul_n_mins1_total), v(record.realisation_cumul_n_mins1_dont_dex)],
#             [f"Prévisions de Clôture {N}", v(record.prev_cloture_n_total), v(record.prev_cloture_n_dont_dex)],
#             [f"Prévisions {N_plus_1}", v(record.prev_n_plus1_total), v(record.prev_n_plus1_dont_dex)],
#             [f"Reste à Réaliser {N_plus_2}/{annee_fin}", v(record.reste_a_realiser_total), v(record.reste_a_realiser_dont_dex)],
#             [f"Prévisions {N_plus_2}", v(record.prev_n_plus2_total), v(record.prev_n_plus2_dont_dex)],
#             [f"Prévisions {N_plus_3}", v(record.prev_n_plus3_total), v(record.prev_n_plus3_dont_dex)],
#             [f"Prévisions {N_plus_4}", v(record.prev_n_plus4_total), v(record.prev_n_plus4_dont_dex)],
#             [f"Prévisions {N_plus_5}", v(record.prev_n_plus5_total), v(record.prev_n_plus5_dont_dex)],
#         ]

#         main_table = Table(main_data, colWidths=[280, 100, 100])
#         main_table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
#             ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
#             ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('BACKGROUND', (0, 1), (-1, 1), colors.whitesmoke),
#             ('BACKGROUND', (0, 3), (-1, 3), colors.whitesmoke),
#             ('BACKGROUND', (0, 5), (-1, 5), colors.whitesmoke),
#             ('BACKGROUND', (0, 7), (-1, 7), colors.whitesmoke),
#             ('BACKGROUND', (0, 9), (-1, 9), colors.whitesmoke),
#             ('ALIGN',  (1, 0), (-1, -1), 'CENTER'),
#             ('GRID',   (0, 0), (-1, -1), 0.5, colors.grey),
#             ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
#         ]))

#         elements.append(Paragraph("<b>Données budgétaires</b>", styles['Heading2']))
#         elements.append(main_table)
#         elements.append(Spacer(1, 15))

#         mensuel_data = [
#             ["Mois", f"Prévisions {N_plus_2} - Total", f"Prévisions {N_plus_2} - Dont DEX"],
#             ["Janvier",   v(record.janvier_total),   v(record.janvier_dont_dex)],
#             ["Février",   v(record.fevrier_total),   v(record.fevrier_dont_dex)],
#             ["Mars",      v(record.mars_total),      v(record.mars_dont_dex)],
#             ["Avril",     v(record.avril_total),     v(record.avril_dont_dex)],
#             ["Mai",       v(record.mai_total),       v(record.mai_dont_dex)],
#             ["Juin",      v(record.juin_total),      v(record.juin_dont_dex)],
#             ["Juillet",   v(record.juillet_total),   v(record.juillet_dont_dex)],
#             ["Août",      v(record.aout_total),      v(record.aout_dont_dex)],
#             ["Septembre", v(record.septembre_total), v(record.septembre_dont_dex)],
#             ["Octobre",   v(record.octobre_total),   v(record.octobre_dont_dex)],
#             ["Novembre",  v(record.novembre_total),  v(record.novembre_dont_dex)],
#             ["Décembre",  v(record.decembre_total),  v(record.decembre_dont_dex)],
#         ]

#         mensuel_table = Table(mensuel_data, colWidths=[180, 150, 150])
#         mensuel_table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
#             ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
#             ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('ALIGN',  (1, 0), (-1, -1), 'CENTER'),
#             ('GRID',   (0, 0), (-1, -1), 0.5, colors.grey),
#             ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
#         ]))

#         elements.append(Paragraph(f"<b>Répartition mensuelle — Prévisions {N_plus_2}</b>", styles['Heading2']))
#         elements.append(mensuel_table)

#         doc.build(elements)
#         buffer.seek(0)

#         return HttpResponse(
#             buffer,
#             content_type='application/pdf',
#             headers={
#                 'Content-Disposition': f'attachment; filename="budget_{record.id}.pdf"'
#             },
#         )
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime
import requests

class BudgetRecordPDFView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def _get_region_name(self, code_region, token):
        """Récupère le nom de la région depuis le service param"""
        if not code_region or code_region in ['', '-', 'None']:
            return '-'
        
        service_url = get_service_param_url()
        
        try:
            headers = {'Authorization': f'Bearer {token}'} if token else {}
            response = requests.get(
                f"{service_url}/params/regions/{code_region}", 
                headers=headers, 
                timeout=5
            )
            
            if response.status_code == 200:
                data = response.json().get('data', {})
                return data.get('nom_region', code_region)
        except Exception as e:
            print(f"Erreur récupération région {code_region}: {e}")
        
        return code_region

    def _get_famille_name(self, code_famille, token):
        """Récupère le nom de la famille depuis le service param"""
        if not code_famille or code_famille in ['', '-', 'None']:
            return '-'
        
        service_url = get_service_param_url()
        
        try:
            headers = {'Authorization': f'Bearer {token}'} if token else {}
            response = requests.get(
                f"{service_url}/params/familles/by-code/{code_famille}", 
                headers=headers, 
                timeout=5
            )
            
            if response.status_code == 200:
                data = response.json().get('data', {})
                return data.get('nom_famille', code_famille)
        except Exception as e:
            print(f"Erreur récupération famille {code_famille}: {e}")
        
        return code_famille

    def get(self, request, pk):
        record = get_object_or_404(BudgetRecord, pk=pk)

        # Récupérer le token
        auth_header = request.headers.get('Authorization', '')
        token = auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else ''

        # Récupérer les noms depuis le backend
        activite = ACTIVITE_MAPPING.get(record.activite, record.activite or '-')
        region = self._get_region_name(record.region, token)
        famille = self._get_famille_name(record.famille, token)

        # Récupérer l'année de début PMT
        annee_debut = record.annee_debut_pmt
        
        # Si l'année n'est pas définie, utiliser l'année courante + 1 comme fallback
        current_year = datetime.now().year
        if not annee_debut:
            annee_debut = current_year + 1
            
        # N = année de début PMT - 1
        N = annee_debut - 1
        N_plus_1 = N + 1
        N_plus_2 = N + 2
        N_plus_3 = N + 3
        N_plus_4 = N + 4
        N_plus_5 = N + 5
        
        # Les années de la période PMT
        annee_fin = record.annee_fin_pmt if record.annee_fin_pmt else N_plus_5

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("RAPPORT BUDGET", styles['Title']))
        elements.append(Spacer(1, 10))

        info_data = [
            ["Activité", activite],
            ["Région", region],
            ["Famille", famille],
            ["Libellé", record.libelle or '-'],
            ["Période PMT", f"{annee_debut} - {annee_fin}"]
        ]

        info_table = Table(info_data, colWidths=[120, 350])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))

        def v(val):
            return val if val is not None else 0

        main_data = [
            ["Désignation", "Total", "Dont DEX"],
            [f"Coût Global Initial PMT {annee_debut}/{annee_fin}", v(record.cout_initial_total), v(record.cout_initial_dont_dex)],
            [f"Réalisations Cumulées à fin {N} au coût réel", v(record.realisation_cumul_n_mins1_total), v(record.realisation_cumul_n_mins1_dont_dex)],
            [f"Prévisions de Clôture {N}", v(record.prev_cloture_n_total), v(record.prev_cloture_n_dont_dex)],
            [f"Prévisions {N_plus_1}", v(record.prev_n_plus1_total), v(record.prev_n_plus1_dont_dex)],
            [f"Reste à Réaliser {N_plus_2}/{annee_fin}", v(record.reste_a_realiser_total), v(record.reste_a_realiser_dont_dex)],
            [f"Prévisions {N_plus_2}", v(record.prev_n_plus2_total), v(record.prev_n_plus2_dont_dex)],
            [f"Prévisions {N_plus_3}", v(record.prev_n_plus3_total), v(record.prev_n_plus3_dont_dex)],
            [f"Prévisions {N_plus_4}", v(record.prev_n_plus4_total), v(record.prev_n_plus4_dont_dex)],
            [f"Prévisions {N_plus_5}", v(record.prev_n_plus5_total), v(record.prev_n_plus5_dont_dex)],
        ]

        main_table = Table(main_data, colWidths=[280, 100, 100])
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.whitesmoke),
            ('BACKGROUND', (0, 3), (-1, 3), colors.whitesmoke),
            ('BACKGROUND', (0, 5), (-1, 5), colors.whitesmoke),
            ('BACKGROUND', (0, 7), (-1, 7), colors.whitesmoke),
            ('BACKGROUND', (0, 9), (-1, 9), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ]))

        elements.append(Paragraph("<b>Données budgétaires</b>", styles['Heading2']))
        elements.append(main_table)
        elements.append(Spacer(1, 15))

        mensuel_data = [
            ["Mois", f"Prévisions {N_plus_2} - Total", f"Prévisions {N_plus_2} - Dont DEX"],
            ["Janvier", v(record.janvier_total), v(record.janvier_dont_dex)],
            ["Février", v(record.fevrier_total), v(record.fevrier_dont_dex)],
            ["Mars", v(record.mars_total), v(record.mars_dont_dex)],
            ["Avril", v(record.avril_total), v(record.avril_dont_dex)],
            ["Mai", v(record.mai_total), v(record.mai_dont_dex)],
            ["Juin", v(record.juin_total), v(record.juin_dont_dex)],
            ["Juillet", v(record.juillet_total), v(record.juillet_dont_dex)],
            ["Août", v(record.aout_total), v(record.aout_dont_dex)],
            ["Septembre", v(record.septembre_total), v(record.septembre_dont_dex)],
            ["Octobre", v(record.octobre_total), v(record.octobre_dont_dex)],
            ["Novembre", v(record.novembre_total), v(record.novembre_dont_dex)],
            ["Décembre", v(record.decembre_total), v(record.decembre_dont_dex)],
        ]

        mensuel_table = Table(mensuel_data, colWidths=[180, 150, 150])
        mensuel_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ]))

        elements.append(Paragraph(f"<b>Répartition mensuelle — Prévisions {N_plus_2}</b>", styles['Heading2']))
        elements.append(mensuel_table)

        doc.build(elements)
        buffer.seek(0)

        return HttpResponse(
            buffer,
            content_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="budget_{record.id}.pdf"'
            },
        )
# ─────────────────────────────────────────
# VÉRIFICATION CALCULS
# ─────────────────────────────────────────


class VerificationCalculsView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAgent]
    """
    GET /verification/?upload_id=1
    Vérifie la cohérence des calculs pour chaque record
    """

    def get(self, request):
        print("[DEBUG] === Début VerificationCalculsView ===")
        qs = BudgetRecord.objects.all()

        uid = request.query_params.get('upload_id')
        print(f"[DEBUG] upload_id param: {uid}")
        if uid:
            qs = qs.filter(upload_id=uid)
            print(f"[DEBUG] Filtered by upload_id={uid}, count={qs.count()}")

        qs = clean_queryset(qs)

        errors   = []
        warnings = []
        total    = qs.count()
        ok_count = 0
        print(f"[DEBUG] Total records after clean_queryset: {total}")

        TOLERANCE = 1

        def val(x):
            return float(x or 0)

        def check(record, label, gauche, droite):
            diff = abs(gauche - droite)
            if diff > TOLERANCE:
                print(f"[DEBUG] Check failed - Record ID: {record.id}, Label: {label}, Diff: {diff}")
                return {
                    'record_id':  record.id,
                    'libelle':    record.libelle or '-',
                    'region':     REGION_MAPPING.get(record.region, record.region or '-'),
                    'famille':    get_famille_nom(record.famille or '-'),
                    'activite':   ACTIVITE_MAPPING.get(record.activite, record.activite or '-'),
                    'regle':      label,
                    'gauche':     round(gauche, 2),
                    'droite':     round(droite, 2),
                    'difference': round(diff, 2),
                }
            return None

        for idx, record in enumerate(qs):
            print(f"[DEBUG] --- Processing record {idx+1}/{total} (ID: {record.id}) ---")
            record_errors = []

            # Check 1
            gauche1 = val(record.real_s1_n_total) + val(record.prev_s2_n_total)
            droite1 = val(record.prev_cloture_n_total)
            print(f"[DEBUG] Record {record.id} - Check1: gauche={gauche1}, droite={droite1}")
            e = check(record, "Réal.S1 (total) + Prév.S2 (total) = Prév.Clôture N (total)", gauche1, droite1)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check1 FAILED")

            # Check 2
            gauche2 = val(record.real_s1_n_dont_dex) + val(record.prev_s2_n_dont_dex)
            droite2 = val(record.prev_cloture_n_dont_dex)
            print(f"[DEBUG] Record {record.id} - Check2: gauche={gauche2}, droite={droite2}")
            e = check(record, "Réal.S1 (DEX) + Prév.S2 (DEX) = Prév.Clôture N (DEX)", gauche2, droite2)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check2 FAILED")

            # Check 3
            gauche3 = (val(record.prev_n_plus2_total) + val(record.prev_n_plus3_total)
                      + val(record.prev_n_plus4_total) + val(record.prev_n_plus5_total))
            droite3 = val(record.reste_a_realiser_total)
            print(f"[DEBUG] Record {record.id} - Check3: gauche={gauche3}, droite={droite3}")
            e = check(record, "Reste à Réaliser (total) = Prév.N+2 + N+3 + N+4 + N+5 (total)", gauche3, droite3)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check3 FAILED")

            # Check 4
            gauche4 = (val(record.prev_n_plus2_dont_dex) + val(record.prev_n_plus3_dont_dex)
                      + val(record.prev_n_plus4_dont_dex) + val(record.prev_n_plus5_dont_dex))
            droite4 = val(record.reste_a_realiser_dont_dex)
            print(f"[DEBUG] Record {record.id} - Check4: gauche={gauche4}, droite={droite4}")
            e = check(record, "Reste à Réaliser (DEX) = Prév.N+2 + N+3 + N+4 + N+5 (DEX)", gauche4, droite4)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check4 FAILED")

            # Check 5
            somme_mois = (
                val(record.janvier_total)   + val(record.fevrier_total)  +
                val(record.mars_total)      + val(record.avril_total)    +
                val(record.mai_total)       + val(record.juin_total)     +
                val(record.juillet_total)   + val(record.aout_total)     +
                val(record.septembre_total) + val(record.octobre_total)  +
                val(record.novembre_total)  + val(record.decembre_total)
            )
            print(f"[DEBUG] Record {record.id} - Check5: somme_mois={somme_mois}, prev_n_plus1={val(record.prev_n_plus1_total)}")
            e = check(record, "Prév.N+1 (total) = Somme des 12 mois (total)", somme_mois, val(record.prev_n_plus1_total))
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check5 FAILED")

            # Check 6
            gauche6 = (val(record.realisation_cumul_n_mins1_total) + val(record.prev_cloture_n_total)
                      + val(record.prev_n_plus1_total) + val(record.reste_a_realiser_total))
            droite6 = val(record.cout_initial_total)
            print(f"[DEBUG] Record {record.id} - Check6: gauche={gauche6}, droite={droite6}")
            e = check(record, "Coût Global (total) = Réal.Cumul N-1 + Prév.Clôture N + Prév.N+1 + Reste à Réaliser (total)", gauche6, droite6)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check6 FAILED")

            # Check 7
            gauche7 = (val(record.realisation_cumul_n_mins1_dont_dex) + val(record.prev_cloture_n_dont_dex)
                      + val(record.prev_n_plus1_dont_dex) + val(record.reste_a_realiser_dont_dex))
            droite7 = val(record.cout_initial_dont_dex)
            print(f"[DEBUG] Record {record.id} - Check7: gauche={gauche7}, droite={droite7}")
            e = check(record, "Coût Global (DEX) = Réal.Cumul N-1 + Prév.Clôture N + Prév.N+1 + Reste à Réaliser (DEX)", gauche7, droite7)
            if e: 
                record_errors.append(e)
                print(f"[DEBUG] Record {record.id} - Check7 FAILED")

            if record_errors:
                errors.extend(record_errors)
                print(f"[DEBUG] Record {record.id} has {len(record_errors)} error(s)")
            else:
                ok_count += 1
                print(f"[DEBUG] Record {record.id} OK")

        print(f"[DEBUG] === Verification Summary: total={total}, ok={ok_count}, errors={len(errors)} ===")
        
        response_data = {
            'resume': {
                'total_records':  total,
                'records_ok':     ok_count,
                'records_errors': total - ok_count,
                'total_erreurs':  len(errors),
            },
            'erreurs': errors,
        }
        
        print(f"[DEBUG] Response prepared, sending {len(errors)} errors")
        return Response(response_data)


# ─────────────────────────────────────────
# SAISIE MANUELLE — NOUVEAU PROJET
# ─────────────────────────────────────────

from rest_framework import status as drf_status


# class NouveauProjetView(APIView):
#     """
#     POST /api/budget/nouveau-projet/
    
#     Crée la première version d'un projet (sans historique)
#     Tous les champs de réalisation sont NULL
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsResponsableStructure]

#     @staticmethod
#     def _to_float_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return float(val)
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         return round(sum(filtered), 2) if filtered else None

#     def post(self, request):
#         data = request.data
        
#         # 1. Informations depuis le token
#         region_id = getattr(request.user, 'region_id', None)
#         structure_id = getattr(request.user, 'structure_id', None)
#         created_by = request.user.id

#         if not region_id or not structure_id:
#             return Response({'error': 'region_id ou structure_id manquant'}, status=400)

#         # 2. Champs obligatoires
#         activite = data.get('activite')
#         perimetre_code = data.get('perimetre')
#         famille_code = data.get('famille')
#         code_division = data.get('code_division')
#         libelle = data.get('libelle')

#         missing = [f for f, v in {
#             'activite': activite, 'perimetre': perimetre_code,
#             'famille': famille_code, 'code_division': code_division, 'libelle': libelle
#         }.items() if not v]

#         if missing:
#             return Response({'error': f"Champs manquants: {', '.join(missing)}"}, status=400)

#         # 3. Vérifier que le code_division n'existe pas déjà
#         if BudgetRecord.objects.filter(code_division=code_division).exists():
#             return Response({
#                 'error': f"Le code_division '{code_division}' existe déjà. Utilisez l'API de modification."
#             }, status=400)

#         # 4. Intervalle PMT
#         intervalle_pmt = data.get('intervalle_pmt')
#         if intervalle_pmt and isinstance(intervalle_pmt, list) and len(intervalle_pmt) == 2:
#             annee_debut_pmt = int(intervalle_pmt[0])
#             annee_fin_pmt = int(intervalle_pmt[1])
#         else:
#             annee_debut_pmt = data.get('annee_debut_pmt')
#             annee_fin_pmt = data.get('annee_fin_pmt')

#         # 5. Résolution région via service param
#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')

#         try:
#             region_resp = requests.get(
#                 f"{service_url}/params/regions/id/{region_id}",
#                 headers={'Authorization': token},
#                 timeout=5
#             )
#             if region_resp.status_code != 200:
#                 return Response({'error': 'Erreur région'}, status=400)
#             region_data = region_resp.json().get('data', {})
#             code_region = region_data.get('code_region')
#         except Exception as e:
#             return Response({'error': f'Erreur service région: {e}'}, status=503)

#         # 6. Lecture des champs financiers
#         PREVISIONS_KEYS = ['prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5']
#         MOIS_KEYS = ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                      'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']

#         v = {}
#         for key in PREVISIONS_KEYS:
#             v[f'{key}_total'] = self._to_float_or_none(data.get(f'{key}_total'))
#             v[f'{key}_dont_dex'] = self._to_float_or_none(data.get(f'{key}_dont_dex'))
#         for mois in MOIS_KEYS:
#             v[f'{mois}_total'] = self._to_float_or_none(data.get(f'{mois}_total'))
#             v[f'{mois}_dont_dex'] = self._to_float_or_none(data.get(f'{mois}_dont_dex'))

#         # 7. Calculs pour nouveau projet
#         prev_n_plus1_total = self._safe_sum([v[f'{m}_total'] for m in MOIS_KEYS])
#         prev_n_plus1_dex = self._safe_sum([v[f'{m}_dont_dex'] for m in MOIS_KEYS])
#         rar_total = self._safe_sum([v[f'{k}_total'] for k in PREVISIONS_KEYS])
#         rar_dex = self._safe_sum([v[f'{k}_dont_dex'] for k in PREVISIONS_KEYS])
#         cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#         cout_dex = self._safe_sum([prev_n_plus1_dex, rar_dex])

#         # 8. Création
#         upload = ExcelUpload.objects.create(
#             file_name=f"nouveau_projet_{code_division}",
#             status='processed'
#         )

#         record = BudgetRecord.objects.create(
#             upload=upload,
#             activite=activite,
#             region=code_region,
#             perm=perimetre_code,
#             famille=famille_code,
#             code_division=code_division,
#             libelle=libelle,
#             annee_debut_pmt=annee_debut_pmt,
#             annee_fin_pmt=annee_fin_pmt,
#             region_id=region_id,
#             structure_id=structure_id,
#             created_by=created_by,
#             type_projet='nouveau',
#             description_technique=data.get('description_technique'),
#             opportunite_projet=data.get('opportunite_projet'),
            
#             # Versionnement
#             parent_id=None,
#             version=1,
#             is_active=True,
#             version_comment="Création initiale",
            
#             # Champs de réalisation (NULL pour nouveau projet)
#             realisation_cumul_n_mins1_total=None,
#             realisation_cumul_n_mins1_dont_dex=None,
#             real_s1_n_total=None,
#             real_s1_n_dont_dex=None,
#             prev_s2_n_total=None,
#             prev_s2_n_dont_dex=None,
#             prev_cloture_n_total=None,
#             prev_cloture_n_dont_dex=None,
            
#             # Champs calculés
#             prev_n_plus1_total=prev_n_plus1_total,
#             prev_n_plus1_dont_dex=prev_n_plus1_dex,
#             reste_a_realiser_total=rar_total,
#             reste_a_realiser_dont_dex=rar_dex,
#             cout_initial_total=cout_total,
#             cout_initial_dont_dex=cout_dex,
            
#             # Prévisions
#             **{k: v[k] for k in [f'{key}_total' for key in PREVISIONS_KEYS] + 
#                [f'{key}_dont_dex' for key in PREVISIONS_KEYS] +
#                [f'{mois}_total' for mois in MOIS_KEYS] +
#                [f'{mois}_dont_dex' for mois in MOIS_KEYS]}
#         )

#         serializer = BudgetRecordSerializer(record)
#         return Response({
#             'success': True,
#             'message': 'Projet créé avec succès (version 1)',
#             'data': serializer.data
#         }, status=201)
import logging
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.core.cache import cache
import requests
import traceback

# Configuration du logger
logger = logging.getLogger(__name__)

# class NouveauProjetView(APIView):
#     """
#     POST /api/budget/nouveau-projet/
    
#     Crée la première version d'un projet (sans historique)
#     Tous les champs de réalisation sont NULL
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsResponsableStructure]

#     @staticmethod
#     def _to_float_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return float(val)
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         return round(sum(filtered), 2) if filtered else None

#     def post(self, request):
#         logger.info("=" * 80)
#         logger.info("🔵 NOUVEAU PROJET - Début de la requête")
#         logger.info(f"📅 Timestamp: {datetime.now()}")
#         logger.info(f"🔑 Headers: {dict(request.headers)}")
#         logger.info(f"👤 User: {request.user}")
#         logger.info(f"🆔 User ID: {request.user.id if request.user else 'None'}")
        
#         data = request.data
#         logger.info(f"📦 Données reçues: {data}")
        
#         # 1. Informations depuis le token
#         region_id = getattr(request.user, 'region_id', None)
#         structure_id = getattr(request.user, 'structure_id', None)
#         created_by = request.user.id if request.user else None

#         logger.info(f"📍 region_id: {region_id}")
#         logger.info(f"🏢 structure_id: {structure_id}")
#         logger.info(f"👨‍💻 created_by: {created_by}")

#         if not region_id or not structure_id:
#             logger.error(f"❌ region_id ou structure_id manquant - region_id={region_id}, structure_id={structure_id}")
#             return Response({
#                 'error': 'region_id ou structure_id manquant',
#                 'debug': {
#                     'region_id': region_id,
#                     'structure_id': structure_id,
#                     'user_attrs': dir(request.user) if request.user else 'No user'
#                 }
#             }, status=400)

#         # 2. Champs obligatoires
#         activite = data.get('activite')
#         perimetre_code = data.get('perimetre')
#         famille_code = data.get('famille')
#         code_division = data.get('code_division')
#         libelle = data.get('libelle')

#         logger.info(f"📋 Champs obligatoires:")
#         logger.info(f"  - activite: {activite}")
#         logger.info(f"  - perimetre: {perimetre_code}")
#         logger.info(f"  - famille: {famille_code}")
#         logger.info(f"  - code_division: {code_division}")
#         logger.info(f"  - libelle: {libelle}")

#         missing = [f for f, v in {
#             'activite': activite, 
#             'perimetre': perimetre_code,
#             'famille': famille_code, 
#             'code_division': code_division, 
#             'libelle': libelle
#         }.items() if not v]

#         if missing:
#             logger.error(f"❌ Champs manquants: {missing}")
#             return Response({
#                 'error': f"Champs manquants: {', '.join(missing)}",
#                 'debug': {'missing_fields': missing, 'received_data': list(data.keys())}
#             }, status=400)

#         # 3. Vérifier que le code_division n'existe pas déjà
#         logger.info(f"🔍 Vérification existence code_division: {code_division}")
#         if BudgetRecord.objects.filter(code_division=code_division).exists():
#             logger.error(f"❌ code_division existe déjà: {code_division}")
#             return Response({
#                 'error': f"Le code_division '{code_division}' existe déjà. Utilisez l'API de modification.",
#                 'debug': {'code_division': code_division}
#             }, status=400)
#         logger.info(f"✅ code_division disponible")

#         # 4. Intervalle PMT
#         intervalle_pmt = data.get('intervalle_pmt')
#         if intervalle_pmt and isinstance(intervalle_pmt, list) and len(intervalle_pmt) == 2:
#             annee_debut_pmt = int(intervalle_pmt[0])
#             annee_fin_pmt = int(intervalle_pmt[1])
#         else:
#             annee_debut_pmt = data.get('annee_debut_pmt')
#             annee_fin_pmt = data.get('annee_fin_pmt')
        
#         logger.info(f"📅 PMT Intervalle: debut={annee_debut_pmt}, fin={annee_fin_pmt}")

#         # 5. Récupérer le code région via service param
#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')
        
#         logger.info(f"🌐 Service Param URL: {service_url}")
#         logger.info(f"🔑 Token (first 50 chars): {token[:50]}..." if token else "🔑 Token: None")
#         logger.info(f"📍 Region ID pour appel: {region_id}")

#         code_region = None
#         region_nom = None

#         try:
#             api_url = f"{service_url}/params/regions/id/{region_id}"
#             logger.info(f"📡 Appel API: {api_url}")
            
#             region_resp = requests.get(
#                 api_url,
#                 headers={'Authorization': token},
#                 timeout=5
#             )
            
#             logger.info(f"📊 Status code: {region_resp.status_code}")
#             logger.info(f"📄 Response text: {region_resp.text[:200]}" if region_resp.text else "📄 Response: empty")
            
#             if region_resp.status_code == 200:
#                 region_data = region_resp.json().get('data', {})
#                 code_region = region_data.get('code_region')
#                 region_nom = region_data.get('nom')
#                 logger.info(f"✅ Région trouvée - code: {code_region}, nom: {region_nom}")
#             else:
#                 logger.error(f"❌ Erreur région - Status: {region_resp.status_code}")
#                 return Response({
#                     'error': f'Erreur lors de la récupération de la région',
#                     'debug': {
#                         'status_code': region_resp.status_code,
#                         'response': region_resp.text,
#                         'region_id': region_id,
#                         'url': api_url
#                     }
#                 }, status=400)
                
#         except requests.exceptions.Timeout:
#             logger.error(f"⏰ Timeout sur l'appel au service param (5 secondes)")
#             return Response({
#                 'error': 'Timeout du service param',
#                 'debug': {'url': api_url, 'region_id': region_id}
#             }, status=503)
#         except Exception as e:
#             logger.error(f"💥 Exception lors de l'appel au service param: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur service région: {str(e)}',
#                 'debug': {'exception': str(e), 'region_id': region_id}
#             }, status=503)

#         if not code_region:
#             logger.error(f"❌ Code région non trouvé pour region_id={region_id}")
#             return Response({
#                 'error': 'Code région non trouvé',
#                 'debug': {'region_id': region_id, 'region_data': region_data if 'region_data' in locals() else None}
#             }, status=404)

#         # 6. Lecture des champs financiers
#         logger.info("💰 Lecture des champs financiers...")
#         PREVISIONS_KEYS = ['prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5']
#         MOIS_KEYS = ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                      'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']

#         v = {}
#         for key in PREVISIONS_KEYS:
#             v[f'{key}_total'] = self._to_float_or_none(data.get(f'{key}_total'))
#             v[f'{key}_dont_dex'] = self._to_float_or_none(data.get(f'{key}_dont_dex'))
#             logger.debug(f"  {key}_total: {v[f'{key}_total']}, {key}_dont_dex: {v[f'{key}_dont_dex']}")
            
#         for mois in MOIS_KEYS:
#             v[f'{mois}_total'] = self._to_float_or_none(data.get(f'{mois}_total'))
#             v[f'{mois}_dont_dex'] = self._to_float_or_none(data.get(f'{mois}_dont_dex'))
        
#         logger.info(f"✅ {len([x for x in v if v[x] is not None])} champs financiers chargés")

#         # 7. Calculs pour nouveau projet
#         logger.info("🧮 Calculs en cours...")
#         prev_n_plus1_total = self._safe_sum([v[f'{m}_total'] for m in MOIS_KEYS])
#         prev_n_plus1_dex = self._safe_sum([v[f'{m}_dont_dex'] for m in MOIS_KEYS])
#         rar_total = self._safe_sum([v[f'{k}_total'] for k in PREVISIONS_KEYS])
#         rar_dex = self._safe_sum([v[f'{k}_dont_dex'] for k in PREVISIONS_KEYS])
#         cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#         cout_dex = self._safe_sum([prev_n_plus1_dex, rar_dex])
        
#         logger.info(f"📊 Résultats calculs:")
#         logger.info(f"  - prev_n_plus1_total: {prev_n_plus1_total}")
#         logger.info(f"  - prev_n_plus1_dex: {prev_n_plus1_dex}")
#         logger.info(f"  - rar_total: {rar_total}")
#         logger.info(f"  - rar_dex: {rar_dex}")
#         logger.info(f"  - cout_total: {cout_total}")
#         logger.info(f"  - cout_dex: {cout_dex}")

#         # 8. Création en base
#         logger.info("💾 Création de l'upload...")
#         upload = ExcelUpload.objects.create(
#             file_name=f"nouveau_projet_{code_division}",
#             status='processed'
#         )
#         logger.info(f"✅ Upload créé: id={upload.id}")

#         logger.info("💾 Création du BudgetRecord...")
#         try:
#             record = BudgetRecord.objects.create(
#                 upload=upload,
#                 activite=activite,
#                 region=code_region,
#                 perm=perimetre_code,
#                 famille=famille_code,
#                 code_division=code_division,
#                 libelle=libelle,
#                 annee_debut_pmt=annee_debut_pmt,
#                 annee_fin_pmt=annee_fin_pmt,
#                 region_id=region_id,
#                 structure_id=structure_id,
#                 created_by=created_by,
#                 type_projet='nouveau',
#                 description_technique=data.get('description_technique'),
#                 opportunite_projet=data.get('opportunite_projet'),
                
#                 # Versionnement
#                 parent_id=None,
#                 version=1,
#                 is_active=True,
#                 version_comment="Création initiale",
#                 statut='soumis',
                
#                 # Champs de réalisation (NULL pour nouveau projet)
#                 realisation_cumul_n_mins1_total=None,
#                 realisation_cumul_n_mins1_dont_dex=None,
#                 real_s1_n_total=None,
#                 real_s1_n_dont_dex=None,
#                 prev_s2_n_total=None,
#                 prev_s2_n_dont_dex=None,
#                 prev_cloture_n_total=None,
#                 prev_cloture_n_dont_dex=None,
                
#                 # Champs calculés
#                 prev_n_plus1_total=prev_n_plus1_total,
#                 prev_n_plus1_dont_dex=prev_n_plus1_dex,
#                 reste_a_realiser_total=rar_total,
#                 reste_a_realiser_dont_dex=rar_dex,
#                 cout_initial_total=cout_total,
#                 cout_initial_dont_dex=cout_dex,
                
#                 # Prévisions
#                 **{k: v[k] for k in [f'{key}_total' for key in PREVISIONS_KEYS] + 
#                    [f'{key}_dont_dex' for key in PREVISIONS_KEYS] +
#                    [f'{mois}_total' for mois in MOIS_KEYS] +
#                    [f'{mois}_dont_dex' for mois in MOIS_KEYS]}
#             )
#             logger.info(f"✅ BudgetRecord créé: id={record.id}, code_division={record.code_division}")
#         except Exception as e:
#             logger.error(f"💥 Erreur création BudgetRecord: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur création en base: {str(e)}',
#                 'debug': {'exception': str(e)}
#             }, status=500)

#         # 9. Sérialisation avec contexte
#         logger.info("🔄 Sérialisation des données...")
#         try:
#             serializer = BudgetRecordSerializer(
#                 record, 
#                 context={'request': request}
#             )
#             serialized_data = serializer.data
#             logger.info(f"✅ Sérialisation réussie - champs: {list(serialized_data.keys())}")
#         except Exception as e:
#             logger.error(f"💥 Erreur sérialisation: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur sérialisation: {str(e)}',
#                 'debug': {'exception': str(e)}
#             }, status=500)
        
#         # 10. Réponse finale
#         logger.info("🎉 Succès - Projet créé!")
#         logger.info("=" * 80)
        
#         return Response({
#             'success': True,
#             'message': 'Projet créé avec succès (version 1)',
#             'data': serialized_data,
#             'debug_info': {  # Optionnel: à retirer en production
#                 'region_code': code_region,
#                 'region_nom': region_nom,
#                 'record_id': record.id
#             }
#         }, status=201)













# #########################################################################################################""
# class NouveauProjetView(APIView):
#     """
#     POST /api/budget/nouveau-projet/
    
#     Crée la première version d'un projet (sans historique)
#     Tous les champs de réalisation sont NULL
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsResponsableStructure]

#     @staticmethod
#     def _to_float_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return float(val)
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         return round(sum(filtered), 2) if filtered else None

#     # ⭐ NOUVELLE MÉTHODE : Validation total >= dex
#     def _validate_total_ge_dex(self, total, dex, field_name):
#         """Valide que total >= dex (si les deux sont non-None)"""
#         if total is not None and dex is not None and total < dex:
#             raise ValidationError(
#                 f"{field_name}: Le total ({total}) ne peut pas être inférieur au DEX ({dex})"
#             )
#         return True

#     # ⭐ NOUVELLE MÉTHODE : Valider tous les champs
#     def _validate_all_totals(self, data, v, prev_n_plus1_total, prev_n_plus1_dex, 
#                               rar_total, rar_dex, cout_total, cout_dex):
#         """Valide toutes les paires total/dont_dex"""
#         errors = []
        
#         # 1. Validation des mois
#         mois_list = ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                      'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']
        
#         for mois in mois_list:
#             total = v.get(f'{mois}_total')
#             dex = v.get(f'{mois}_dont_dex')
#             try:
#                 self._validate_total_ge_dex(total, dex, mois)
#             except ValidationError as e:
#                 errors.append(str(e))
        
#         # 2. Validation des prévisions annuelles (N+2 à N+5)
#         for annee in ['prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5']:
#             total = v.get(f'{annee}_total')
#             dex = v.get(f'{annee}_dont_dex')
#             try:
#                 self._validate_total_ge_dex(total, dex, annee)
#             except ValidationError as e:
#                 errors.append(str(e))
        
#         # 3. Validation des totaux calculés
#         try:
#             self._validate_total_ge_dex(prev_n_plus1_total, prev_n_plus1_dex, "Prévision N+1")
#             self._validate_total_ge_dex(rar_total, rar_dex, "Reste à réaliser (RAR)")
#             self._validate_total_ge_dex(cout_total, cout_dex, "Coût initial")
#         except ValidationError as e:
#             errors.append(str(e))
        
#         if errors:
#             raise ValidationError({"total_dex_mismatches": errors})
        
#         return True

#     def post(self, request):
#         logger.info("=" * 80)
#         logger.info("🔵 NOUVEAU PROJET - Début de la requête")
#         logger.info(f"📅 Timestamp: {datetime.now()}")
#         logger.info(f"🔑 Headers: {dict(request.headers)}")
#         logger.info(f"👤 User: {request.user}")
#         logger.info(f"🆔 User ID: {request.user.id if request.user else 'None'}")
        
#         data = request.data
#         logger.info(f"📦 Données reçues: {data}")
        
#         # 1. Informations depuis le token
#         region_id = getattr(request.user, 'region_id', None)
#         structure_id = getattr(request.user, 'structure_id', None)
#         created_by = request.user.id if request.user else None

#         logger.info(f"📍 region_id: {region_id}")
#         logger.info(f"🏢 structure_id: {structure_id}")
#         logger.info(f"👨‍💻 created_by: {created_by}")

#         if not region_id or not structure_id:
#             logger.error(f"❌ region_id ou structure_id manquant - region_id={region_id}, structure_id={structure_id}")
#             return Response({
#                 'error': 'region_id ou structure_id manquant',
#                 'debug': {
#                     'region_id': region_id,
#                     'structure_id': structure_id,
#                     'user_attrs': dir(request.user) if request.user else 'No user'
#                 }
#             }, status=400)

#         # 2. Champs obligatoires
#         activite = data.get('activite')
#         perimetre_code = data.get('perimetre')
#         famille_code = data.get('famille')
#         code_division = data.get('code_division')
#         libelle = data.get('libelle')

#         logger.info(f"📋 Champs obligatoires:")
#         logger.info(f"  - activite: {activite}")
#         logger.info(f"  - perimetre: {perimetre_code}")
#         logger.info(f"  - famille: {famille_code}")
#         logger.info(f"  - code_division: {code_division}")
#         logger.info(f"  - libelle: {libelle}")

#         missing = [f for f, v in {
#             'activite': activite, 
#             'perimetre': perimetre_code,
#             'famille': famille_code, 
#             'code_division': code_division, 
#             'libelle': libelle
#         }.items() if not v]

#         if missing:
#             logger.error(f"❌ Champs manquants: {missing}")
#             return Response({
#                 'error': f"Champs manquants: {', '.join(missing)}",
#                 'debug': {'missing_fields': missing, 'received_data': list(data.keys())}
#             }, status=400)

#         # 3. Vérifier que le code_division n'existe pas déjà
#         logger.info(f"🔍 Vérification existence code_division: {code_division}")
#         if BudgetRecord.objects.filter(code_division=code_division).exists():
#             logger.error(f"❌ code_division existe déjà: {code_division}")
#             return Response({
#                 'error': f"Le code_division '{code_division}' existe déjà. Utilisez l'API de modification.",
#                 'debug': {'code_division': code_division}
#             }, status=400)
#         logger.info(f"✅ code_division disponible")

#         # 4. Intervalle PMT
#         intervalle_pmt = data.get('intervalle_pmt')
#         if intervalle_pmt and isinstance(intervalle_pmt, list) and len(intervalle_pmt) == 2:
#             annee_debut_pmt = int(intervalle_pmt[0])
#             annee_fin_pmt = int(intervalle_pmt[1])
#         else:
#             annee_debut_pmt = data.get('annee_debut_pmt')
#             annee_fin_pmt = data.get('annee_fin_pmt')
        
#         # ⭐ Validation des années
#         if annee_debut_pmt and annee_fin_pmt and annee_debut_pmt > annee_fin_pmt:
#             logger.error(f"❌ Année début {annee_debut_pmt} > année fin {annee_fin_pmt}")
#             return Response({
#                 'error': "L'année début doit être inférieure à l'année fin",
#                 'debug': {'annee_debut_pmt': annee_debut_pmt, 'annee_fin_pmt': annee_fin_pmt}
#             }, status=400)
        
#         logger.info(f"📅 PMT Intervalle: debut={annee_debut_pmt}, fin={annee_fin_pmt}")

#         # 5. Récupérer le code région via service param
#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')
        
#         logger.info(f"🌐 Service Param URL: {service_url}")
#         logger.info(f"🔑 Token (first 50 chars): {token[:50]}..." if token else "🔑 Token: None")
#         logger.info(f"📍 Region ID pour appel: {region_id}")

#         code_region = None
#         region_nom = None

#         try:
#             api_url = f"{service_url}/params/regions/id/{region_id}"
#             logger.info(f"📡 Appel API: {api_url}")
            
#             region_resp = requests.get(
#                 api_url,
#                 headers={'Authorization': token},
#                 timeout=5
#             )
            
#             logger.info(f"📊 Status code: {region_resp.status_code}")
#             logger.info(f"📄 Response text: {region_resp.text[:200]}" if region_resp.text else "📄 Response: empty")
            
#             if region_resp.status_code == 200:
#                 region_data = region_resp.json().get('data', {})
#                 code_region = region_data.get('code_region')
#                 region_nom = region_data.get('nom')
#                 logger.info(f"✅ Région trouvée - code: {code_region}, nom: {region_nom}")
#             else:
#                 logger.error(f"❌ Erreur région - Status: {region_resp.status_code}")
#                 return Response({
#                     'error': f'Erreur lors de la récupération de la région',
#                     'debug': {
#                         'status_code': region_resp.status_code,
#                         'response': region_resp.text,
#                         'region_id': region_id,
#                         'url': api_url
#                     }
#                 }, status=400)
                
#         except requests.exceptions.Timeout:
#             logger.error(f"⏰ Timeout sur l'appel au service param (5 secondes)")
#             return Response({
#                 'error': 'Timeout du service param',
#                 'debug': {'url': api_url, 'region_id': region_id}
#             }, status=503)
#         except Exception as e:
#             logger.error(f"💥 Exception lors de l'appel au service param: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur service région: {str(e)}',
#                 'debug': {'exception': str(e), 'region_id': region_id}
#             }, status=503)

#         if not code_region:
#             logger.error(f"❌ Code région non trouvé pour region_id={region_id}")
#             return Response({
#                 'error': 'Code région non trouvé',
#                 'debug': {'region_id': region_id, 'region_data': region_data if 'region_data' in locals() else None}
#             }, status=404)

#         # 6. Lecture des champs financiers
#         logger.info("💰 Lecture des champs financiers...")
#         PREVISIONS_KEYS = ['prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5']
#         MOIS_KEYS = ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                      'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']

#         v = {}
#         for key in PREVISIONS_KEYS:
#             v[f'{key}_total'] = self._to_float_or_none(data.get(f'{key}_total'))
#             v[f'{key}_dont_dex'] = self._to_float_or_none(data.get(f'{key}_dont_dex'))
#             logger.debug(f"  {key}_total: {v[f'{key}_total']}, {key}_dont_dex: {v[f'{key}_dont_dex']}")
            
#         for mois in MOIS_KEYS:
#             v[f'{mois}_total'] = self._to_float_or_none(data.get(f'{mois}_total'))
#             v[f'{mois}_dont_dex'] = self._to_float_or_none(data.get(f'{mois}_dont_dex'))
        
#         logger.info(f"✅ {len([x for x in v if v[x] is not None])} champs financiers chargés")

#         # 7. Calculs pour nouveau projet
#         logger.info("🧮 Calculs en cours...")
#         prev_n_plus1_total = self._safe_sum([v[f'{m}_total'] for m in MOIS_KEYS])
#         prev_n_plus1_dex = self._safe_sum([v[f'{m}_dont_dex'] for m in MOIS_KEYS])
#         rar_total = self._safe_sum([v[f'{k}_total'] for k in PREVISIONS_KEYS])
#         rar_dex = self._safe_sum([v[f'{k}_dont_dex'] for k in PREVISIONS_KEYS])
#         cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#         cout_dex = self._safe_sum([prev_n_plus1_dex, rar_dex])
        
#         logger.info(f"📊 Résultats calculs:")
#         logger.info(f"  - prev_n_plus1_total: {prev_n_plus1_total}")
#         logger.info(f"  - prev_n_plus1_dex: {prev_n_plus1_dex}")
#         logger.info(f"  - rar_total: {rar_total}")
#         logger.info(f"  - rar_dex: {rar_dex}")
#         logger.info(f"  - cout_total: {cout_total}")
#         logger.info(f"  - cout_dex: {cout_dex}")

#         # ⭐⭐⭐ VALIDATION CRITIQUE : total >= dont_dex ⭐⭐⭐
#         logger.info("🔍 Validation des incohérences total/DEX...")
#         try:
#             self._validate_all_totals(data, v, prev_n_plus1_total, prev_n_plus1_dex,
#                                        rar_total, rar_dex, cout_total, cout_dex)
#             logger.info("✅ Validation total/DEX réussie")
#         except ValidationError as e:
#             logger.error(f"❌ Échec validation total/DEX: {e.message_dict}")
#             return Response({
#                 'error': 'Incohérence dans les montants financiers',
#                 'details': {
#                     'message': 'Le total doit être supérieur ou égal au DEX pour tous les champs',
#                     'violations': e.message_dict.get('total_dex_mismatches', [])
#                 }
#             }, status=400)

#         # 8. Création en base
#         logger.info("💾 Création de l'upload...")
#         upload = ExcelUpload.objects.create(
#             file_name=f"nouveau_projet_{code_division}",
#             status='processed'
#         )
#         logger.info(f"✅ Upload créé: id={upload.id}")

#         logger.info("💾 Création du BudgetRecord...")
#         try:
#             record = BudgetRecord.objects.create(
#                 upload=upload,
#                 activite=activite,
#                 region=code_region,
#                 perm=perimetre_code,
#                 famille=famille_code,
#                 code_division=code_division,
#                 libelle=libelle,
#                 annee_debut_pmt=annee_debut_pmt,
#                 annee_fin_pmt=annee_fin_pmt,
#                 region_id=region_id,
#                 structure_id=structure_id,
#                 created_by=created_by,
#                 type_projet='nouveau',
#                 description_technique=data.get('description_technique'),
#                 opportunite_projet=data.get('opportunite_projet'),
                
#                 # Versionnement
#                 parent_id=None,
#                 version=1,
#                 is_active=True,
#                 version_comment="Création initiale",
#                 # statut='soumis',
#                 statut_workflow='soumis',
#                 #############################################################
                
#                 # Champs de réalisation (NULL pour nouveau projet)
#                 realisation_cumul_n_mins1_total=None,
#                 realisation_cumul_n_mins1_dont_dex=None,
#                 real_s1_n_total=None,
#                 real_s1_n_dont_dex=None,
#                 prev_s2_n_total=None,
#                 prev_s2_n_dont_dex=None,
#                 prev_cloture_n_total=None,
#                 prev_cloture_n_dont_dex=None,
                
#                 # Champs calculés
#                 prev_n_plus1_total=prev_n_plus1_total,
#                 prev_n_plus1_dont_dex=prev_n_plus1_dex,
#                 reste_a_realiser_total=rar_total,
#                 reste_a_realiser_dont_dex=rar_dex,
#                 cout_initial_total=cout_total,
#                 cout_initial_dont_dex=cout_dex,
                
#                 # Prévisions
#                 **{k: v[k] for k in [f'{key}_total' for key in PREVISIONS_KEYS] + 
#                    [f'{key}_dont_dex' for key in PREVISIONS_KEYS] +
#                    [f'{mois}_total' for mois in MOIS_KEYS] +
#                    [f'{mois}_dont_dex' for mois in MOIS_KEYS]}
#             )
#             logger.info(f"✅ BudgetRecord créé: id={record.id}, code_division={record.code_division}")
#         except Exception as e:
#             logger.error(f"💥 Erreur création BudgetRecord: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur création en base: {str(e)}',
#                 'debug': {'exception': str(e)}
#             }, status=500)

#         # 9. Sérialisation avec contexte
#         logger.info("🔄 Sérialisation des données...")
#         try:
#             serializer = BudgetRecordSerializer(
#                 record, 
#                 context={'request': request}
#             )
#             serialized_data = serializer.data
#             logger.info(f"✅ Sérialisation réussie - champs: {list(serialized_data.keys())}")
#         except Exception as e:
#             logger.error(f"💥 Erreur sérialisation: {str(e)}")
#             logger.error(traceback.format_exc())
#             return Response({
#                 'error': f'Erreur sérialisation: {str(e)}',
#                 'debug': {'exception': str(e)}
#             }, status=500)
        
#         # 10. Réponse finale
#         logger.info("🎉 Succès - Projet créé!")
#         logger.info("=" * 80)
        
#         return Response({
#             'success': True,
#             'message': 'Projet créé avec succès (version 1)',
#             'data': serialized_data,
#             'debug_info': {
#                 'region_code': code_region,
#                 'region_nom': region_nom,
#                 'record_id': record.id
#             }
#         }, status=201)     
    







# ##############################################################################
# views.py

# 



"""
views_creation.py
=================
Deux endpoints de création de projet budget :

  POST /api/budget/nouveau-projet/structure/
      → responsable_structure  (region_id + structure_id depuis JWT)
      → body : activite, famille, code_division, libelle, perimetre
      → BudgetRecord.region    = code résolu via service param
      → BudgetRecord.direction = None

  POST /api/budget/nouveau-projet/departement/
      → responsable_departement  (direction_id + departement_id depuis JWT)
      → body : activite, famille, code_division, libelle
      → BudgetRecord.region    = None
      → BudgetRecord.direction = code résolu via service param
"""

import traceback
import requests
from datetime import datetime

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError

from .models import BudgetRecord, ExcelUpload
from .serializers import BudgetRecordSerializer, get_service_param_url
from .remote_auth import RemoteJWTAuthentication

import logging
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# CLASSE DE BASE — logique partagée
# ══════════════════════════════════════════════════════════════════════════════

class BaseNouveauProjetView(APIView):

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsAuthenticated]

    PREVISIONS_KEYS = [
        'prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5'
    ]
    MOIS_KEYS = [
        'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
        'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre'
    ]

    # ─────────────────────────────────────────────────────────────────
    # UTILITAIRES FINANCIERS
    # ─────────────────────────────────────────────────────────────────

    @staticmethod
    def _to_float_or_none(val):
        if val in (None, '', 'null', 'None'):
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _safe_sum(values):
        filtered = [v for v in values if v is not None]
        return round(sum(filtered), 2) if filtered else None

    def _validate_total_ge_dex(self, total, dex, field_name):
        if total is not None and dex is not None and total < dex:
            raise ValidationError(
                f"{field_name} : le total ({total}) ne peut pas être "
                f"inférieur au DEX ({dex})."
            )

    def _validate_all_totals(self, v,
                              prev_n1_t, prev_n1_d,
                              rar_t, rar_d,
                              cout_t, cout_d):
        errors = []
        for key in self.MOIS_KEYS + self.PREVISIONS_KEYS:
            try:
                self._validate_total_ge_dex(
                    v.get(f'{key}_total'),
                    v.get(f'{key}_dont_dex'),
                    key
                )
            except ValidationError as e:
                errors.append(str(e))

        for total, dex, label in [
            (prev_n1_t, prev_n1_d, "Prévision N+1"),
            (rar_t,     rar_d,     "Reste à réaliser (RAR)"),
            (cout_t,    cout_d,    "Coût initial"),
        ]:
            try:
                self._validate_total_ge_dex(total, dex, label)
            except ValidationError as e:
                errors.append(str(e))

        if errors:
            raise ValidationError({"total_dex_mismatches": errors})

    # ─────────────────────────────────────────────────────────────────
    # CHAMPS COMMUNS OBLIGATOIRES
    # ─────────────────────────────────────────────────────────────────

    def _validate_common_fields(self, data):
        """
        Valide activite, famille, code_division, libelle.
        Retourne (code_division, libelle, activite, famille).
        """
        fields = {
            'activite':      data.get('activite'),
            'famille':       data.get('famille'),
            'code_division': data.get('code_division'),
            'libelle':       data.get('libelle'),
        }
        missing = [k for k, v in fields.items() if not v]
        if missing:
            raise ValidationError(
                f"Champs obligatoires manquants : {', '.join(missing)}"
            )
        return (
            fields['code_division'],
            fields['libelle'],
            fields['activite'],
            fields['famille'],
        )

    # ─────────────────────────────────────────────────────────────────
    # INTERVALLE PMT
    # ─────────────────────────────────────────────────────────────────

    # def _parse_pmt(self, data):
    #     intervalle = data.get('intervalle_pmt')
    #     if isinstance(intervalle, list) and len(intervalle) == 2:
    #         annee_debut = int(intervalle[0])
    #         annee_fin   = int(intervalle[1])
    #     else:
    #         annee_debut = data.get('annee_debut_pmt')
    #         annee_fin   = data.get('annee_fin_pmt')

    #     if annee_debut and annee_fin:
    #         if int(annee_debut) > int(annee_fin):
    #             raise ValidationError(
    #                 "L'année début PMT ne peut pas être supérieure à l'année fin PMT."
    #             )
    #     return annee_debut, annee_fin
    def _parse_pmt(self, data):
        """Parse l'intervalle PMT et calcule automatiquement fin = debut + 4 si nécessaire"""
        
        intervalle = data.get('intervalle_pmt')
        
        # Cas 1: intervalle_pmt = [debut, fin]
        if isinstance(intervalle, list) and len(intervalle) == 2:
            annee_debut = int(intervalle[0])
            annee_fin = int(intervalle[1])
            
        # Cas 2: intervalle_pmt = [debut] seulement
        elif isinstance(intervalle, list) and len(intervalle) == 1:
            annee_debut = int(intervalle[0])
            annee_fin = annee_debut + 4  # ← +4 ans
            
        # Cas 3: annee_debut_pmt directement
        else:
            annee_debut = data.get('annee_debut_pmt')
            annee_fin = data.get('annee_fin_pmt')
            
            # Si fin est null mais debut existe, calculer fin = debut + 4
            if annee_debut and not annee_fin:
                annee_fin = annee_debut + 4
        
        if annee_debut and annee_fin:
            if int(annee_debut) > int(annee_fin):
                raise ValidationError(
                    "L'année début PMT ne peut pas être supérieure à l'année fin PMT."
                )
        
        return annee_debut, annee_fin

    # ─────────────────────────────────────────────────────────────────
    # CHAMPS FINANCIERS + CALCULS
    # ─────────────────────────────────────────────────────────────────

    def _load_financials(self, data):
        v = {}
        for key in self.PREVISIONS_KEYS + self.MOIS_KEYS:
            v[f'{key}_total']    = self._to_float_or_none(data.get(f'{key}_total'))
            v[f'{key}_dont_dex'] = self._to_float_or_none(data.get(f'{key}_dont_dex'))

        # prev_n_plus1 = somme des 12 mois
        prev_n1_t = self._safe_sum([v[f'{m}_total']    for m in self.MOIS_KEYS])
        prev_n1_d = self._safe_sum([v[f'{m}_dont_dex'] for m in self.MOIS_KEYS])

        # RAR = N+2 → N+5
        rar_t = self._safe_sum([v[f'{k}_total']    for k in self.PREVISIONS_KEYS])
        rar_d = self._safe_sum([v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS])

        # Coût initial = prev_n+1 + RAR
        cout_t = self._safe_sum([prev_n1_t, rar_t])
        cout_d = self._safe_sum([prev_n1_d, rar_d])

        logger.info(
            f"🧮 prev_n+1={prev_n1_t}/{prev_n1_d} | "
            f"RAR={rar_t}/{rar_d} | Coût={cout_t}/{cout_d}"
        )

        self._validate_all_totals(v, prev_n1_t, prev_n1_d, rar_t, rar_d, cout_t, cout_d)
        return v, prev_n1_t, prev_n1_d, rar_t, rar_d, cout_t, cout_d

    # ─────────────────────────────────────────────────────────────────
    # UNICITÉ code_division
    # ─────────────────────────────────────────────────────────────────

    def _check_code_division_unique(self, code_division):
        if BudgetRecord.objects.filter(code_division=code_division).exists():
            raise ValidationError(
                f"Le code_division '{code_division}' existe déjà. "
                f"Utilisez l'API de modification."
            )

    # ─────────────────────────────────────────────────────────────────
    # RÉSOLUTION CODE VIA SERVICE PARAM
    # ─────────────────────────────────────────────────────────────────

    def _resolve_code(self, lookup_id, source, token, service_url):
        """
        source='region'    → GET /params/regions/id/{id}    → retourne code_region
        source='direction' → GET /params/directions/id/{id} → retourne code_direction
        Retourne (code_value, label).
        """
        if source == 'region':
            endpoint = f"{service_url}/params/regions/id/{lookup_id}"
            code_key = 'code_region'
        else:
            endpoint = f"{service_url}/params/directions/{lookup_id}"
            code_key = 'code_direction'

        logger.info(f"📡 Résolution {source} → {endpoint}")

        try:
            resp = requests.get(
                endpoint,
                headers={'Authorization': token},
                timeout=5
            )
            if resp.status_code != 200:
                raise ValidationError(
                    f"Service '{source}' : HTTP {resp.status_code} — {resp.text[:200]}"
                )

            data_resp  = resp.json().get('data', {})
            code_value = data_resp.get(code_key)
            label      = data_resp.get('nom')

            if not code_value:
                raise ValidationError(
                    f"Champ '{code_key}' absent ou vide dans la réponse du service '{source}'."
                )

            logger.info(f"✅ {source} résolu : {code_value} (label={label})")
            return code_value, label

        except requests.exceptions.Timeout:
            raise ValidationError(f"Timeout (5s) — service '{source}'.")
        except ValidationError:
            raise
        except Exception as e:
            logger.error(traceback.format_exc())
            raise ValidationError(f"Exception inattendue service '{source}' : {e}")

    # ─────────────────────────────────────────────────────────────────
    # ÉCRITURE EN BASE (commune)
    # ─────────────────────────────────────────────────────────────────

    def _create_budget_record(
        self, request,
        code_division, libelle, activite, famille,
        region,        # code_region résolu   | None pour département
        direction,     # code_direction résolu | None pour structure
        perm,          # périmètre             | None pour département
        region_id, structure_id, direction_id, departement_id,
        annee_debut_pmt, annee_fin_pmt,
        v, prev_n1_t, prev_n1_d, rar_t, rar_d, cout_t, cout_d,
    ):
        data = request.data

        upload = ExcelUpload.objects.create(
            file_name=f"nouveau_projet_{code_division}",
            status='processed'
        )
        logger.info(f"📁 ExcelUpload id={upload.id}")

        financials_keys = (
            [f'{k}_total'    for k in self.PREVISIONS_KEYS] +
            [f'{k}_dont_dex' for k in self.PREVISIONS_KEYS] +
            [f'{m}_total'    for m in self.MOIS_KEYS]       +
            [f'{m}_dont_dex' for m in self.MOIS_KEYS]
        )

        # 🔥 Récupérer les nouveaux champs (utiliser des noms différents)
        duree_realisation_val = data.get('duree_realisation')
        point_situation_val = data.get('point_situation')
        commentaire_point_situation_val = data.get('commentaire_point_situation')

        record = BudgetRecord.objects.create(
            upload = upload,

            # Identité
            code_division         = code_division,
            libelle               = libelle,
            description_technique = data.get('description_technique'),
            opportunite_projet    = data.get('opportunite_projet'),
            type_projet           = 'nouveau',

            # ── Champs séparés région / direction ─────────────────────
            region    = region,       # rempli pour structure,   None pour département
            direction = direction,    # rempli pour département, None pour structure

            # Métier
            famille  = famille,
            activite = activite,
            perm     = perm,          # None pour département

            # IDs utilisateur
            region_id      = region_id,
            structure_id   = structure_id,    # None pour département
            direction_id   = direction_id,    # None pour structure
            departement_id = departement_id,  # None pour structure
            created_by     = request.user.id,

            # PMT
            annee_debut_pmt = annee_debut_pmt,
            annee_fin_pmt   = annee_fin_pmt,

            # Versionnement
            parent_id       = None,
            version         = 1,
            is_active       = True,
            version_comment = "Création initiale",
            statut_workflow = 'soumis',
            statut_final    = None,

            # 🔥 NOUVEAUX CHAMPS (avec les bonnes variables)
            duree_realisation = duree_realisation_val,
            point_situation = point_situation_val,
            commentaire_point_situation = commentaire_point_situation_val,

            # Réalisation NULL (nouveau projet)
            realisation_cumul_n_mins1_total    = None,
            realisation_cumul_n_mins1_dont_dex = None,
            real_s1_n_total                    = None,
            real_s1_n_dont_dex                 = None,
            prev_s2_n_total                    = None,
            prev_s2_n_dont_dex                 = None,
            prev_cloture_n_total               = None,
            prev_cloture_n_dont_dex            = None,

            # Champs calculés
            prev_n_plus1_total        = prev_n1_t,
            prev_n_plus1_dont_dex     = prev_n1_d,
            reste_a_realiser_total    = rar_t,
            reste_a_realiser_dont_dex = rar_d,
            cout_initial_total        = cout_t,
            cout_initial_dont_dex     = cout_d,

            # Prévisions mensuelles + annuelles
            **{k: v[k] for k in financials_keys}
        )

        logger.info(
            f"✅ BudgetRecord id={record.id} | "
            f"code_division={code_division} | "
            f"region={region} | direction={direction} | "
            f"duree={duree_realisation_val} | point={point_situation_val}"
        )
        return record


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT 1 — STRUCTURE
# POST /api/budget/nouveau-projet/structure/
# ══════════════════════════════════════════════════════════════════════════════

class NouveauProjetView(BaseNouveauProjetView):
    """
    Crée un nouveau projet pour un responsable_structure.

    JWT requis  : region_id, structure_id
    Body requis : activite, famille, code_division, libelle, perimetre
    En base     : region = code_region résolu | direction = None
    """

    ROLE_REQUIS = 'responsable_structure'

    def post(self, request):
        logger.info("=" * 60)
        logger.info("🔵 CRÉATION PROJET [STRUCTURE]")
        logger.info(f"👤 {request.user} | rôle={getattr(request.user, 'role', '?')}")
        data = request.data

        # ── 1. Vérification du rôle ───────────────────────────────────
        role = getattr(request.user, 'role', None)
        if role != self.ROLE_REQUIS:
            return Response(
                {'error': f"Rôle '{role}' non autorisé. Attendu : '{self.ROLE_REQUIS}'"},
                status=403
            )

        # ── 2. IDs depuis le token JWT ────────────────────────────────
        region_id    = getattr(request.user, 'region_id',    None)
        structure_id = getattr(request.user, 'structure_id', None)

        if not region_id:
            return Response({'error': "region_id manquant dans le token JWT."}, status=400)
        if not structure_id:
            return Response({'error': "structure_id manquant dans le token JWT."}, status=400)

        # ── 3. Périmètre (obligatoire pour structure) ─────────────────
        perimetre = data.get('perimetre')
        if not perimetre:
            return Response(
                {'error': "'perimetre' est obligatoire pour responsable_structure."},
                status=400
            )

        # ── 4. Champs communs ─────────────────────────────────────────
        try:
            code_division, libelle, activite, famille = \
                self._validate_common_fields(data)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 5. Unicité code_division ──────────────────────────────────
        try:
            self._check_code_division_unique(code_division)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 6. PMT ────────────────────────────────────────────────────
        try:
            annee_debut_pmt, annee_fin_pmt = self._parse_pmt(data)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 7. Résolution code_region via service param ───────────────
        token       = request.headers.get('Authorization', '')
        service_url = get_service_param_url()
        try:
            code_region, label = self._resolve_code(
                region_id, 'region', token, service_url
            )
        except ValidationError as e:
            return Response({'error': e.detail}, status=503)

        # ── 8. Champs financiers ──────────────────────────────────────
        try:
            v, prev_n1_t, prev_n1_d, rar_t, rar_d, cout_t, cout_d = \
                self._load_financials(data)
        except ValidationError as e:
            return Response(
                {'error': 'Incohérence montants financiers.', 'details': e.detail},
                status=400
            )

        # ── 9. Création en base ───────────────────────────────────────
        try:
            record = self._create_budget_record(
                request,
                code_division   = code_division,
                libelle         = libelle,
                activite        = activite,
                famille         = famille,
                region          = code_region,   # ← champ region rempli
                direction       = None,           # ← champ direction vide
                perm            = perimetre,
                region_id       = region_id,
                structure_id    = structure_id,
                direction_id    = None,
                departement_id  = None,
                annee_debut_pmt = annee_debut_pmt,
                annee_fin_pmt   = annee_fin_pmt,
                v=v,
                prev_n1_t=prev_n1_t, prev_n1_d=prev_n1_d,
                rar_t=rar_t,         rar_d=rar_d,
                cout_t=cout_t,       cout_d=cout_d,
            )
        except Exception as e:
            logger.error(traceback.format_exc())
            return Response({'error': f'Erreur création en base : {e}'}, status=500)

        # ── 10. Sérialisation ─────────────────────────────────────────
        try:
            serialized = BudgetRecordSerializer(
                record, context={'request': request}
            ).data
        except Exception as e:
            logger.error(traceback.format_exc())
            return Response({'error': f'Erreur sérialisation : {e}'}, status=500)

        logger.info(f"🎉 STRUCTURE créé — id={record.id} | region={code_region}")
        logger.info("=" * 60)

        return Response(
            {
                'success': True,
                'message': "Projet créé (version 1 | responsable_structure)",
                'data':    serialized,
                'meta': {
                    'role':         self.ROLE_REQUIS,
                    'region_id':    region_id,
                    'structure_id': structure_id,
                    'region':       code_region,
                    'label':        label,
                    'record_id':    record.id,
                },
            },
            status=201
        )


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT 2 — DÉPARTEMENT
# POST /api/budget/nouveau-projet/departement/
# ══════════════════════════════════════════════════════════════════════════════

class NouveauProjetDepartementView(BaseNouveauProjetView):
    """
    Crée un nouveau projet pour un responsable_departement.

    JWT requis  : direction_id, departement_id  (region_id optionnel)
    Body requis : activite, famille, code_division, libelle
                  (pas de perimetre → perm = None)
    En base     : region = None | direction = code_direction résolu
    """

    ROLE_REQUIS = 'responsable_departement'

    def post(self, request):
        logger.info("=" * 60)
        logger.info("🔵 CRÉATION PROJET [DÉPARTEMENT]")
        logger.info(f"👤 {request.user} | rôle={getattr(request.user, 'role', '?')}")
        data = request.data

        # ── 1. Vérification du rôle ───────────────────────────────────
        role = getattr(request.user, 'role', None)
        if role != self.ROLE_REQUIS:
            return Response(
                {'error': f"Rôle '{role}' non autorisé. Attendu : '{self.ROLE_REQUIS}'"},
                status=403
            )

        # ── 2. IDs depuis le token JWT ────────────────────────────────
        direction_id   = getattr(request.user, 'direction_id',   None)
        departement_id = getattr(request.user, 'departement_id', None)
        region_id      = getattr(request.user, 'region_id',      None)  # optionnel

        if not direction_id:
            return Response({'error': "direction_id manquant dans le token JWT."}, status=400)
        if not departement_id:
            return Response({'error': "departement_id manquant dans le token JWT."}, status=400)

        # ── 3. Champs communs ─────────────────────────────────────────
        try:
            code_division, libelle, activite, famille = \
                self._validate_common_fields(data)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 4. Unicité code_division ──────────────────────────────────
        try:
            self._check_code_division_unique(code_division)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 5. PMT ────────────────────────────────────────────────────
        try:
            annee_debut_pmt, annee_fin_pmt = self._parse_pmt(data)
        except ValidationError as e:
            return Response({'error': e.detail}, status=400)

        # ── 6. Résolution code_direction via service param ────────────
        token       = request.headers.get('Authorization', '')
        service_url = get_service_param_url()
        try:
            code_direction, label = self._resolve_code(
                direction_id, 'direction', token, service_url
            )
        except ValidationError as e:
            return Response({'error': e.detail}, status=503)

        # ── 7. Champs financiers ──────────────────────────────────────
        try:
            v, prev_n1_t, prev_n1_d, rar_t, rar_d, cout_t, cout_d = \
                self._load_financials(data)
        except ValidationError as e:
            return Response(
                {'error': 'Incohérence montants financiers.', 'details': e.detail},
                status=400
            )

        # ── 8. Création en base ───────────────────────────────────────
        try:
            record = self._create_budget_record(
                request,
                code_division   = code_division,
                libelle         = libelle,
                activite        = activite,
                famille         = famille,
                region          = None,            # ← champ region vide
                direction       = code_direction,  # ← champ direction rempli
                perm            = None,            # pas de périmètre
                region_id       = region_id,       # optionnel
                structure_id    = None,
                direction_id    = direction_id,
                departement_id  = departement_id,
                annee_debut_pmt = annee_debut_pmt,
                annee_fin_pmt   = annee_fin_pmt,
                v=v,
                prev_n1_t=prev_n1_t, prev_n1_d=prev_n1_d,
                rar_t=rar_t,         rar_d=rar_d,
                cout_t=cout_t,       cout_d=cout_d,
            )
        except Exception as e:
            logger.error(traceback.format_exc())
            return Response({'error': f'Erreur création en base : {e}'}, status=500)

        # ── 9. Sérialisation ──────────────────────────────────────────
        try:
            serialized = BudgetRecordSerializer(
                record, context={'request': request}
            ).data
        except Exception as e:
            logger.error(traceback.format_exc())
            return Response({'error': f'Erreur sérialisation : {e}'}, status=500)

        logger.info(f"🎉 DÉPARTEMENT créé — id={record.id} | direction={code_direction}")
        logger.info("=" * 60)

        return Response(
            {
                'success': True,
                'message': "Projet créé (version 1 | responsable_departement)",
                'data':    serialized,
                'meta': {
                    'role':           self.ROLE_REQUIS,
                    'direction_id':   direction_id,
                    'departement_id': departement_id,
                    'region_id':      region_id,
                    'direction':      code_direction,
                    'label':          label,
                    'record_id':      record.id,
                },
            },
            status=201
        )
from decimal import Decimal










from django.db import transaction
from django.db.models import Max
from django.core.cache import cache
from decimal import Decimal
import time
from rest_framework.views import APIView
from rest_framework.response import Response


# class ModifierProjetView(APIView):
#     """
#     POST /api/budget/modifier-projet/{code_division}/
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAgent]

#     PREVISIONS_KEYS = [
#         'prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5'
#     ]
#     MOIS_KEYS = [
#         'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre',
#     ]

#     LOCK_TIMEOUT  = 10   # secondes max que le verrou est tenu
#     LOCK_WAIT     = 8    # secondes max à attendre pour obtenir le verrou
#     LOCK_INTERVAL = 0.1  # intervalle de polling

#     # ------------------------------------------------------------------ #
#     # Helpers
#     # ------------------------------------------------------------------ #
#     @staticmethod
#     def _to_decimal_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return Decimal(str(val))
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         if not filtered:
#             return None
#         return sum(
#             v if isinstance(v, Decimal) else Decimal(str(v))
#             for v in filtered
#         )

#     def _is_admin(self, request):
#         return getattr(request.user, 'role', '') in ('admin', 'superadmin')

#     @staticmethod
#     def _all_financial_keys(previsions_keys, mois_keys):
#         keys = []
#         for k in previsions_keys:
#             keys += [f'{k}_total', f'{k}_dont_dex']
#         for m in mois_keys:
#             keys += [f'{m}_total', f'{m}_dont_dex']
#         return keys

#     def _parse_financial_fields(self, data, previsions_keys, mois_keys):
#         v = {}
#         for k in previsions_keys:
#             v[f'{k}_total']    = self._to_decimal_or_none(data.get(f'{k}_total'))
#             v[f'{k}_dont_dex'] = self._to_decimal_or_none(data.get(f'{k}_dont_dex'))
#         for m in mois_keys:
#             v[f'{m}_total']    = self._to_decimal_or_none(data.get(f'{m}_total'))
#             v[f'{m}_dont_dex'] = self._to_decimal_or_none(data.get(f'{m}_dont_dex'))
#         return v

#     def _acquire_lock(self, code_division):
#         """
#         Tente d'acquérir un verrou sur code_division.
#         Retourne True si obtenu, False si timeout.
#         """
#         lock_key     = f'budget_lock_{code_division}'
#         deadline     = time.time() + self.LOCK_WAIT

#         while time.time() < deadline:
#             # add() est atomique : échoue si la clé existe déjà
#             acquired = cache.add(lock_key, '1', timeout=self.LOCK_TIMEOUT)
#             if acquired:
#                 return True
#             time.sleep(self.LOCK_INTERVAL)

#         return False

#     def _release_lock(self, code_division):
#         cache.delete(f'budget_lock_{code_division}')

#     # ------------------------------------------------------------------ #
#     # Point d'entrée principal
#     # ------------------------------------------------------------------ #
#     def post(self, request, code_division):
#         data     = request.data
#         is_admin = self._is_admin(request)

#         projet_exists = BudgetRecord.objects.filter(
#             code_division=code_division
#         ).exists()

#         if not projet_exists:
#             return self._create_first_version(request, code_division)

#         # ── Acquérir le verrou applicatif ────────────────────────────────
#         if not self._acquire_lock(code_division):
#             return Response(
#                 {
#                     'error': (
#                         'Une modification est déjà en cours sur ce projet. '
#                         'Veuillez réessayer dans quelques secondes.'
#                     )
#                 },
#                 status=409
#             )

#         try:
#             return self._do_modification(
#                 request, data, code_division, is_admin
#             )
#         finally:
#             # Toujours libérer le verrou, même en cas d'exception
#             self._release_lock(code_division)

#     # ------------------------------------------------------------------ #
#     # Logique de modification
#     # ------------------------------------------------------------------ #
#     def _do_modification(self, request, data, code_division, is_admin):
#         with transaction.atomic():

#             old_version = (
#                 BudgetRecord.objects.filter(
#                     code_division=code_division, is_active=True
#                 ).first()
#                 or BudgetRecord.objects.filter(
#                     code_division=code_division
#                 ).order_by('-version').first()
#             )

#             if not old_version:
#                 return Response(
#                     {'error': f'Projet {code_division} non trouvé'},
#                     status=404
#                 )

#             # ── Champs interdits ────────────────────────────────────────
#             if not is_admin:
#                 forbidden = [
#                     field
#                     for field, attr, key in [
#                         ('region',    'region',   'region'),
#                         ('perimetre', 'perm',     'perimetre'),
#                         ('famille',   'famille',  'famille'),
#                         ('activite',  'activite', 'activite'),
#                     ]
#                     if key in data
#                     and data[key] != getattr(old_version, attr)
#                 ]
#                 if forbidden:
#                     return Response(
#                         {
#                             'error': (
#                                 f"Modification interdite : "
#                                 f"{', '.join(forbidden)}"
#                             )
#                         },
#                         status=403
#                     )

#             # ── Code final ──────────────────────────────────────────────
#             new_code_division = data.get('code_division', code_division)

#             # ── Numéro de version — calculé sous verrou applicatif ──────
#             # Le verrou cache garantit qu'une seule requête à la fois
#             # arrive ici pour ce code_division.
#             result = BudgetRecord.objects.filter(
#                 code_division=new_code_division
#             ).aggregate(max_version=Max('version'))

#             new_version_number = (result['max_version'] or 0) + 1

#             # ── Désactiver l'ancienne version ───────────────────────────
#             if old_version.is_active:
#                 old_version.is_active = False
#                 old_version.save(update_fields=['is_active'])

#             # ── Créer la nouvelle version ───────────────────────────────
#             new_version = self._create_new_version(
#                 old_version, data, request,
#                 new_code_division, new_version_number
#             )

#         serializer = BudgetRecordSerializer(new_version)
#         return Response(
#             {
#                 'success':          True,
#                 'message':          (
#                     f'Projet modifié — '
#                     f'Version {old_version.version} '
#                     f'→ {new_version_number}'
#                 ),
#                 'ancienne_version': old_version.version,
#                 'nouvelle_version': new_version_number,
#                 'data':             serializer.data,
#             },
#             status=201,
#         )

#     # ------------------------------------------------------------------ #
#     # Création version 1
#     # ------------------------------------------------------------------ #
#     def _create_first_version(self, request, code_division):
#         data       = request.data
#         final_code = data.get('code_division', code_division)

#         if BudgetRecord.objects.filter(code_division=final_code).exists():
#             return Response(
#                 {'error': f'Le projet {final_code} existe déjà.'},
#                 status=400,
#             )

#         v = self._parse_financial_fields(
#             data, self.PREVISIONS_KEYS, self.MOIS_KEYS
#         )

#         prev_n_plus1_total = self._safe_sum(
#             [v[f'{m}_total']    for m in self.MOIS_KEYS]
#         )
#         prev_n_plus1_dex   = self._safe_sum(
#             [v[f'{m}_dont_dex'] for m in self.MOIS_KEYS]
#         )
#         rar_total  = self._safe_sum(
#             [v[f'{k}_total']    for k in self.PREVISIONS_KEYS]
#         )
#         rar_dex    = self._safe_sum(
#             [v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS]
#         )
#         cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#         cout_dex   = self._safe_sum([prev_n_plus1_dex,   rar_dex])

#         upload = ExcelUpload.objects.create(
#             file_name=f'projet_{final_code}_v1',
#             status='processed'
#         )

#         record = BudgetRecord.objects.create(
#             upload=upload,
#             activite=data.get('activite'),
#             region=data.get('region'),
#             perm=data.get('perimetre'),
#             famille=data.get('famille'),
#             code_division=final_code,
#             libelle=data.get('libelle'),
#             annee_debut_pmt=data.get('annee_debut_pmt'),
#             annee_fin_pmt=data.get('annee_fin_pmt'),
#             region_id=getattr(request.user, 'region_id', None),
#             structure_id=getattr(request.user, 'structure_id', None),
#             created_by=request.user.id,
#             type_projet=data.get('type_projet', 'nouveau'),
#             description_technique=data.get('description_technique'),
#             opportunite_projet=data.get('opportunite_projet'),
#             parent_id=None,
#             version=1,
#             is_active=True,
#             version_comment=data.get('version_comment', 'Création initiale'),
#             realisation_cumul_n_mins1_total=None,
#             realisation_cumul_n_mins1_dont_dex=None,
#             real_s1_n_total=None,
#             real_s1_n_dont_dex=None,
#             prev_s2_n_total=None,
#             prev_s2_n_dont_dex=None,
#             prev_cloture_n_total=None,
#             prev_cloture_n_dont_dex=None,
#             prev_n_plus1_total=prev_n_plus1_total,
#             prev_n_plus1_dont_dex=prev_n_plus1_dex,
#             reste_a_realiser_total=rar_total,
#             reste_a_realiser_dont_dex=rar_dex,
#             cout_initial_total=cout_total,
#             cout_initial_dont_dex=cout_dex,
#             **{
#                 k: v[k]
#                 for k in self._all_financial_keys(
#                     self.PREVISIONS_KEYS, self.MOIS_KEYS
#                 )
#             },
#         )

#         return Response(
#             {
#                 'success': True,
#                 'message': 'Projet créé (version 1)',
#                 'data':    BudgetRecordSerializer(record).data,
#             },
#             status=201,
#         )

#     # ------------------------------------------------------------------ #
#     # Création d'une nouvelle version
#     # ------------------------------------------------------------------ #
#     def _create_new_version(
#         self, old_version, new_data, request,
#         new_code_division, new_version_number
#     ):
#         v = {}
#         for key in self.PREVISIONS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k    = f'{key}{suffix}'
#                 v[k] = (
#                     self._to_decimal_or_none(new_data.get(k))
#                     or getattr(old_version, k)
#                 )
#         for mois in self.MOIS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k    = f'{mois}{suffix}'
#                 v[k] = (
#                     self._to_decimal_or_none(new_data.get(k))
#                     or getattr(old_version, k)
#                 )

#         def nd(key):
#             raw = self._to_decimal_or_none(new_data.get(key))
#             return raw if raw is not None else getattr(old_version, key)

#         real_cumul_total = nd('realisation_cumul_n_mins1_total')
#         real_cumul_dex   = nd('realisation_cumul_n_mins1_dont_dex')
#         real_s1_total    = nd('real_s1_n_total')
#         real_s1_dex      = nd('real_s1_n_dont_dex')
#         prev_s2_total    = nd('prev_s2_n_total')
#         prev_s2_dex      = nd('prev_s2_n_dont_dex')

#         prev_n_plus1_total = self._safe_sum(
#             [v[f'{m}_total']    for m in self.MOIS_KEYS]
#         )
#         prev_n_plus1_dex   = self._safe_sum(
#             [v[f'{m}_dont_dex'] for m in self.MOIS_KEYS]
#         )
#         rar_total = self._safe_sum(
#             [v[f'{k}_total']    for k in self.PREVISIONS_KEYS]
#         )
#         rar_dex   = self._safe_sum(
#             [v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS]
#         )

#         projet_type = new_data.get('type_projet', old_version.type_projet)

#         if projet_type == 'en_cours' and (
#             real_s1_total is not None or prev_s2_total is not None
#         ):
#             prev_cloture_total = self._safe_sum([real_s1_total, prev_s2_total])
#             prev_cloture_dex   = self._safe_sum([real_s1_dex,   prev_s2_dex])
#             cout_total = self._safe_sum([
#                 real_cumul_total, prev_cloture_total,
#                 prev_n_plus1_total, rar_total
#             ])
#             cout_dex = self._safe_sum([
#                 real_cumul_dex, prev_cloture_dex,
#                 prev_n_plus1_dex, rar_dex
#             ])
#         else:
#             prev_cloture_total = None
#             prev_cloture_dex   = None
#             cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#             cout_dex   = self._safe_sum([prev_n_plus1_dex,   rar_dex])

#         parent_id = old_version.parent_id or old_version.id

#         upload = ExcelUpload.objects.create(
#             file_name=f'projet_{new_code_division}_v{new_version_number}',
#             status='processed',
#         )

#         return BudgetRecord.objects.create(
#             upload=upload,
#             activite=new_data.get('activite', old_version.activite),
#             region=new_data.get('region', old_version.region),
#             perm=new_data.get('perimetre', old_version.perm),
#             famille=new_data.get('famille', old_version.famille),
#             code_division=new_code_division,
#             libelle=new_data.get('libelle', old_version.libelle),
#             annee_debut_pmt=new_data.get(
#                 'annee_debut_pmt', old_version.annee_debut_pmt
#             ),
#             annee_fin_pmt=new_data.get(
#                 'annee_fin_pmt', old_version.annee_fin_pmt
#             ),
#             region_id=getattr(request.user, 'region_id', None),
#             structure_id=getattr(request.user, 'structure_id', None),
#             created_by=request.user.id,
#             type_projet=projet_type,
#             description_technique=new_data.get(
#                 'description_technique', old_version.description_technique
#             ),
#             opportunite_projet=new_data.get(
#                 'opportunite_projet', old_version.opportunite_projet
#             ),
#             parent_id=parent_id,
#             version=new_version_number,
#             is_active=True,
#             version_comment=new_data.get(
#                 'version_comment', f'Version {new_version_number}'
#             ),
#             realisation_cumul_n_mins1_total=real_cumul_total,
#             realisation_cumul_n_mins1_dont_dex=real_cumul_dex,
#             real_s1_n_total=real_s1_total,
#             real_s1_n_dont_dex=real_s1_dex,
#             prev_s2_n_total=prev_s2_total,
#             prev_s2_n_dont_dex=prev_s2_dex,
#             prev_cloture_n_total=prev_cloture_total,
#             prev_cloture_n_dont_dex=prev_cloture_dex,
#             prev_n_plus1_total=prev_n_plus1_total,
#             prev_n_plus1_dont_dex=prev_n_plus1_dex,
#             reste_a_realiser_total=rar_total,
#             reste_a_realiser_dont_dex=rar_dex,
#             cout_initial_total=cout_total,
#             cout_initial_dont_dex=cout_dex,
#             **{
#                 k: v[k]
#                 for k in self._all_financial_keys(
#                     self.PREVISIONS_KEYS, self.MOIS_KEYS
#                 )
#             },
#         )
# class HistoriqueProjetView(APIView):
#     """
#     GET /api/budget/historique/{code_division}/        → tout l'historique
#     GET /api/budget/historique/{code_division}/actif/  → version active uniquement
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsUser]

#     def get(self, request, code_division, mode=None):

#         qs = BudgetRecord.objects.filter(
#             code_division=code_division
#         ).order_by('-version')

#         if not qs.exists():
#             return Response(
#                 {'error': f'Projet {code_division} introuvable.'},
#                 status=404
#             )

#         # ── /actif/ → version active uniquement ─────────────────────────
#         if mode == 'actif':
#             actif = qs.filter(is_active=True).first() or qs.first()
#             return Response({
#                 'code_division': code_division,
#                 'version_active': BudgetRecordSerializer(actif).data,
#             })

#         # ── Historique complet ───────────────────────────────────────────
#         total     = qs.count()
#         actif     = qs.filter(is_active=True).first() or qs.first()
#         derniere  = qs.first()          # version la plus haute (ordering=-version)
#         premiere  = qs.last()           # version 1

#         return Response({
#             'code_division':    code_division,
#             'total_versions':   total,
#             'version_active':   actif.version,
#             'derniere_version': derniere.version,
#             'premiere_version': premiere.version,
#             'historique': BudgetRecordSerializer(qs, many=True).data,
#         })
##############################################################################################################""
# ================================================================== #
#  VUE 2 — Responsable Structure : champs identitaires auto-remplis
# ================================================================== #
# class ModifierProjetResponsableView(APIView):
#     """
#     GET  /api/budget/responsable/modifier-projet/{code_division}/
#          → retourne la version active avec region/perimetre/famille/activite
#            déjà remplis (lecture seule pour le front)

#     POST /api/budget/responsable/modifier-projet/{code_division}/
#          → modification ; region/perimetre/famille/activite sont
#            injectés automatiquement depuis la version active,
#            le responsable ne peut pas les changer
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsResponsableStructure]   # responsable structure

#     PREVISIONS_KEYS = [
#         'prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5'
#     ]
#     MOIS_KEYS = [
#         'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre',
#     ]
#     LOCK_TIMEOUT  = 10
#     LOCK_WAIT     = 8
#     LOCK_INTERVAL = 0.1

#     # Champs que le responsable ne peut PAS modifier — récupérés automatiquement
#     AUTO_FIELDS = [
#         ('region',    'region',   'region'),
#         ('perimetre', 'perm',     'perimetre'),
#         ('famille',   'famille',  'famille'),
#         ('activite',  'activite', 'activite'),
#     ]

#     # ------------------------------------------------------------------ #
#     # Helpers (identiques)
#     # ------------------------------------------------------------------ #
#     @staticmethod
#     def _to_decimal_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return Decimal(str(val))
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         if not filtered:
#             return None
#         return sum(
#             v if isinstance(v, Decimal) else Decimal(str(v))
#             for v in filtered
#         )

#     @staticmethod
#     def _all_financial_keys(previsions_keys, mois_keys):
#         keys = []
#         for k in previsions_keys:
#             keys += [f'{k}_total', f'{k}_dont_dex']
#         for m in mois_keys:
#             keys += [f'{m}_total', f'{m}_dont_dex']
#         return keys

#     def _parse_financial_fields(self, data, previsions_keys, mois_keys):
#         v = {}
#         for k in previsions_keys:
#             v[f'{k}_total']    = self._to_decimal_or_none(data.get(f'{k}_total'))
#             v[f'{k}_dont_dex'] = self._to_decimal_or_none(data.get(f'{k}_dont_dex'))
#         for m in mois_keys:
#             v[f'{m}_total']    = self._to_decimal_or_none(data.get(f'{m}_total'))
#             v[f'{m}_dont_dex'] = self._to_decimal_or_none(data.get(f'{m}_dont_dex'))
#         return v

#     def _acquire_lock(self, code_division):
#         lock_key = f'budget_lock_{code_division}'
#         deadline = time.time() + self.LOCK_WAIT
#         while time.time() < deadline:
#             if cache.add(lock_key, '1', timeout=self.LOCK_TIMEOUT):
#                 return True
#             time.sleep(self.LOCK_INTERVAL)
#         return False

#     def _release_lock(self, code_division):
#         cache.delete(f'budget_lock_{code_division}')

#     # ------------------------------------------------------------------ #
#     # GET — version active + champs auto mis en avant pour le front
#     # ------------------------------------------------------------------ #
#     def get(self, request, code_division):
#         actif = (
#             BudgetRecord.objects.filter(
#                 code_division=code_division, is_active=True
#             ).first()
#             or BudgetRecord.objects.filter(
#                 code_division=code_division
#             ).order_by('-version').first()
#         )
#         if not actif:
#             return Response(
#                 {'error': f'Projet {code_division} introuvable.'},
#                 status=404
#             )

#         serializer = BudgetRecordSerializer(actif)

#         # On expose explicitement les champs auto pour que le front
#         # sache quoi afficher en lecture seule
#         auto_values = {
#             data_key: getattr(actif, model_attr)
#             for _, model_attr, data_key in self.AUTO_FIELDS
#         }

#         return Response({
#             'code_division':  code_division,
#             'champs_auto':    auto_values,   # région, périmètre, famille, activité
#             'version_active': serializer.data,
#         })

#     # ------------------------------------------------------------------ #
#     # POST — modification avec injection automatique des champs auto
#     # ------------------------------------------------------------------ #
#     def post(self, request, code_division):
#         if not BudgetRecord.objects.filter(code_division=code_division).exists():
#             return Response(
#                 {'error': f'Projet {code_division} introuvable. Utilisez la création.'},
#                 status=404
#             )

#         if not self._acquire_lock(code_division):
#             return Response(
#                 {'error': 'Une modification est déjà en cours. Réessayez dans quelques secondes.'},
#                 status=409
#             )
#         try:
#             return self._do_modification(request, code_division)
#         finally:
#             self._release_lock(code_division)

#     # ------------------------------------------------------------------ #
#     # Logique de modification avec champs auto injectés
#     # ------------------------------------------------------------------ #
#     def _do_modification(self, request, code_division):
#         data = request.data.copy()   # copie mutable du QueryDict/dict

#         with transaction.atomic():
#             old_version = (
#                 BudgetRecord.objects.filter(
#                     code_division=code_division, is_active=True
#                 ).first()
#                 or BudgetRecord.objects.filter(
#                     code_division=code_division
#                 ).order_by('-version').first()
#             )
#             if not old_version:
#                 return Response(
#                     {'error': f'Projet {code_division} non trouvé'},
#                     status=404
#                 )

#             # ── Injection automatique des champs protégés ───────────────
#             # On écrase silencieusement ce que le front aurait pu envoyer
#             for _, model_attr, data_key in self.AUTO_FIELDS:
#                 data[data_key] = getattr(old_version, model_attr)

#             # ── Numéro de version ───────────────────────────────────────
#             # Le responsable ne peut pas changer le code_division
#             new_code_division  = code_division
#             result             = BudgetRecord.objects.filter(
#                 code_division=new_code_division
#             ).aggregate(max_version=Max('version'))
#             new_version_number = (result['max_version'] or 0) + 1

#             if old_version.is_active:
#                 old_version.is_active = False
#                 old_version.save(update_fields=['is_active'])

#             new_version = self._create_new_version(
#                 old_version, data, request,
#                 new_code_division, new_version_number
#             )

#         serializer = BudgetRecordSerializer(new_version)
#         return Response(
#             {
#                 'success':          True,
#                 'message':          f'Projet modifié — Version {old_version.version} → {new_version_number}',
#                 'ancienne_version': old_version.version,
#                 'nouvelle_version': new_version_number,
#                 'data':             serializer.data,
#             },
#             status=201,
#         )

#     # ------------------------------------------------------------------ #
#     # Création d'une nouvelle version (identique à l'original)
#     # ------------------------------------------------------------------ #
#     def _create_new_version(self, old_version, new_data, request, new_code_division, new_version_number):
#         v = {}
#         for key in self.PREVISIONS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k    = f'{key}{suffix}'
#                 v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)
#         for mois in self.MOIS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k    = f'{mois}{suffix}'
#                 v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)

#         def nd(key):
#             raw = self._to_decimal_or_none(new_data.get(key))
#             return raw if raw is not None else getattr(old_version, key)

#         real_cumul_total = nd('realisation_cumul_n_mins1_total')
#         real_cumul_dex   = nd('realisation_cumul_n_mins1_dont_dex')
#         real_s1_total    = nd('real_s1_n_total')
#         real_s1_dex      = nd('real_s1_n_dont_dex')
#         prev_s2_total    = nd('prev_s2_n_total')
#         prev_s2_dex      = nd('prev_s2_n_dont_dex')

#         prev_n_plus1_total = self._safe_sum([v[f'{m}_total']    for m in self.MOIS_KEYS])
#         prev_n_plus1_dex   = self._safe_sum([v[f'{m}_dont_dex'] for m in self.MOIS_KEYS])
#         rar_total          = self._safe_sum([v[f'{k}_total']    for k in self.PREVISIONS_KEYS])
#         rar_dex            = self._safe_sum([v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS])

#         projet_type = new_data.get('type_projet', old_version.type_projet)

#         if projet_type == 'en_cours' and (real_s1_total is not None or prev_s2_total is not None):
#             prev_cloture_total = self._safe_sum([real_s1_total, prev_s2_total])
#             prev_cloture_dex   = self._safe_sum([real_s1_dex,   prev_s2_dex])
#             cout_total = self._safe_sum([real_cumul_total, prev_cloture_total, prev_n_plus1_total, rar_total])
#             cout_dex   = self._safe_sum([real_cumul_dex,   prev_cloture_dex,   prev_n_plus1_dex,   rar_dex])
#         else:
#             prev_cloture_total = None
#             prev_cloture_dex   = None
#             cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#             cout_dex   = self._safe_sum([prev_n_plus1_dex,   rar_dex])

#         upload = ExcelUpload.objects.create(
#             file_name=f'projet_{new_code_division}_v{new_version_number}', status='processed'
#         )
#         return BudgetRecord.objects.create(
#             upload=upload,
#             # Champs injectés automatiquement (écrasés avant l'appel)
#             activite=new_data.get('activite',              old_version.activite),
#             region=new_data.get('region',                  old_version.region),
#             perm=new_data.get('perimetre',                 old_version.perm),
#             famille=new_data.get('famille',                old_version.famille),
#             code_division=new_code_division,
#             libelle=new_data.get('libelle',                old_version.libelle),
#             annee_debut_pmt=new_data.get('annee_debut_pmt', old_version.annee_debut_pmt),
#             annee_fin_pmt=new_data.get('annee_fin_pmt',    old_version.annee_fin_pmt),
#             region_id=getattr(request.user, 'region_id', None),
#             structure_id=getattr(request.user, 'structure_id', None),
#             created_by=request.user.id,
#             type_projet=projet_type,
#             description_technique=new_data.get('description_technique', old_version.description_technique),
#             opportunite_projet=new_data.get('opportunite_projet',       old_version.opportunite_projet),
#             parent_id=old_version.parent_id or old_version.id,
#             version=new_version_number,
#             is_active=True,
#             version_comment=new_data.get('version_comment', f'Version {new_version_number}'),
#             realisation_cumul_n_mins1_total=real_cumul_total,
#             realisation_cumul_n_mins1_dont_dex=real_cumul_dex,
#             real_s1_n_total=real_s1_total,
#             real_s1_n_dont_dex=real_s1_dex,
#             prev_s2_n_total=prev_s2_total,
#             prev_s2_n_dont_dex=prev_s2_dex,
#             prev_cloture_n_total=prev_cloture_total,
#             prev_cloture_n_dont_dex=prev_cloture_dex,
#             prev_n_plus1_total=prev_n_plus1_total,
#             prev_n_plus1_dont_dex=prev_n_plus1_dex,
#             reste_a_realiser_total=rar_total,
#             reste_a_realiser_dont_dex=rar_dex,
#             cout_initial_total=cout_total,
#             cout_initial_dont_dex=cout_dex,
#             **{k: v[k] for k in self._all_financial_keys(self.PREVISIONS_KEYS, self.MOIS_KEYS)},
#         )
# ══════════════════════════════════════════════════════════════════════════════
# MODIFICATION PROJET — RESPONSABLE STRUCTURE
# POST /api/budget/responsable/modifier-projet/structure/{code_division}/
# ══════════════════════════════════════════════════════════════════════════════

# class BaseModifierProjetView(APIView):
#     """Classe de base pour la modification de projet (responsable)"""

#     authentication_classes = [RemoteJWTAuthentication]

#     PREVISIONS_KEYS = [
#         'prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5'
#     ]
#     MOIS_KEYS = [
#         'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre',
#     ]
#     LOCK_TIMEOUT  = 10
#     LOCK_WAIT     = 8
#     LOCK_INTERVAL = 0.1

#     # Champs que le responsable ne peut PAS modifier — récupérés automatiquement
#     # À surcharger dans les sous-classes
#     AUTO_FIELDS = []  # Liste de tuples (data_key, model_attr, front_key)

#     @staticmethod
#     def _to_decimal_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return Decimal(str(val))
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _safe_sum(values):
#         filtered = [v for v in values if v is not None]
#         if not filtered:
#             return None
#         return sum(
#             v if isinstance(v, Decimal) else Decimal(str(v))
#             for v in filtered
#         )

#     @staticmethod
#     def _all_financial_keys(previsions_keys, mois_keys):
#         keys = []
#         for k in previsions_keys:
#             keys += [f'{k}_total', f'{k}_dont_dex']
#         for m in mois_keys:
#             keys += [f'{m}_total', f'{m}_dont_dex']
#         return keys

#     def _parse_financial_fields(self, data, previsions_keys, mois_keys):
#         v = {}
#         for k in previsions_keys:
#             v[f'{k}_total']    = self._to_decimal_or_none(data.get(f'{k}_total'))
#             v[f'{k}_dont_dex'] = self._to_decimal_or_none(data.get(f'{k}_dont_dex'))
#         for m in mois_keys:
#             v[f'{m}_total']    = self._to_decimal_or_none(data.get(f'{m}_total'))
#             v[f'{m}_dont_dex'] = self._to_decimal_or_none(data.get(f'{m}_dont_dex'))
#         return v

#     def _acquire_lock(self, code_division):
#         lock_key = f'budget_lock_{code_division}'
#         deadline = time.time() + self.LOCK_WAIT
#         while time.time() < deadline:
#             if cache.add(lock_key, '1', timeout=self.LOCK_TIMEOUT):
#                 return True
#             time.sleep(self.LOCK_INTERVAL)
#         return False

#     def _release_lock(self, code_division):
#         cache.delete(f'budget_lock_{code_division}')

#     def get(self, request, code_division):
#         actif = (
#             BudgetRecord.objects.filter(
#                 code_division=code_division, is_active=True
#             ).first()
#             or BudgetRecord.objects.filter(
#                 code_division=code_division
#             ).order_by('-version').first()
#         )
#         if not actif:
#             return Response(
#                 {'error': f'Projet {code_division} introuvable.'},
#                 status=404
#             )

#         serializer = BudgetRecordSerializer(actif)

#         auto_values = {}
#         for data_key, model_attr, front_key in self.AUTO_FIELDS:
#             auto_values[front_key] = getattr(actif, model_attr)

#         return Response({
#             'code_division':  code_division,
#             'champs_auto':    auto_values,
#             'version_active': serializer.data,
#         })

#     def post(self, request, code_division):
#         if not BudgetRecord.objects.filter(code_division=code_division).exists():
#             return Response(
#                 {'error': f'Projet {code_division} introuvable. Utilisez la création.'},
#                 status=404
#             )

#         if not self._acquire_lock(code_division):
#             return Response(
#                 {'error': 'Une modification est déjà en cours. Réessayez dans quelques secondes.'},
#                 status=409
#             )
#         try:
#             return self._do_modification(request, code_division)
#         finally:
#             self._release_lock(code_division)

#     def _do_modification(self, request, code_division):
#         data = request.data.copy()

#         with transaction.atomic():
#             old_version = (
#                 BudgetRecord.objects.filter(
#                     code_division=code_division, is_active=True
#                 ).first()
#                 or BudgetRecord.objects.filter(
#                     code_division=code_division
#                 ).order_by('-version').first()
#             )
#             if not old_version:
#                 return Response(
#                     {'error': f'Projet {code_division} non trouvé'},
#                     status=404
#                 )

#             # Injection automatique des champs protégés
#             for data_key, model_attr, front_key in self.AUTO_FIELDS:
#                 data[front_key] = getattr(old_version, model_attr)

#             new_code_division = code_division
#             result = BudgetRecord.objects.filter(
#                 code_division=new_code_division
#             ).aggregate(max_version=Max('version'))
#             new_version_number = (result['max_version'] or 0) + 1

#             if old_version.is_active:
#                 old_version.is_active = False
#                 old_version.save(update_fields=['is_active'])

#             new_version = self._create_new_version(
#                 old_version, data, request,
#                 new_code_division, new_version_number
#             )

#         serializer = BudgetRecordSerializer(new_version)
#         return Response(
#             {
#                 'success': True,
#                 'message': f'Projet modifié — Version {old_version.version} → {new_version_number}',
#                 'ancienne_version': old_version.version,
#                 'nouvelle_version': new_version_number,
#                 'data': serializer.data,
#             },
#             status=201,
#         )

#     def _create_new_version(self, old_version, new_data, request, new_code_division, new_version_number):
#         v = {}
#         for key in self.PREVISIONS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k = f'{key}{suffix}'
#                 v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)
#         for mois in self.MOIS_KEYS:
#             for suffix in ('_total', '_dont_dex'):
#                 k = f'{mois}{suffix}'
#                 v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)

#         def nd(key):
#             raw = self._to_decimal_or_none(new_data.get(key))
#             return raw if raw is not None else getattr(old_version, key)

#         real_cumul_total = nd('realisation_cumul_n_mins1_total')
#         real_cumul_dex   = nd('realisation_cumul_n_mins1_dont_dex')
#         real_s1_total    = nd('real_s1_n_total')
#         real_s1_dex      = nd('real_s1_n_dont_dex')
#         prev_s2_total    = nd('prev_s2_n_total')
#         prev_s2_dex      = nd('prev_s2_n_dont_dex')

#         prev_n_plus1_total = self._safe_sum([v[f'{m}_total'] for m in self.MOIS_KEYS])
#         prev_n_plus1_dex   = self._safe_sum([v[f'{m}_dont_dex'] for m in self.MOIS_KEYS])
#         rar_total          = self._safe_sum([v[f'{k}_total'] for k in self.PREVISIONS_KEYS])
#         rar_dex            = self._safe_sum([v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS])

#         projet_type = new_data.get('type_projet', old_version.type_projet)

#         if projet_type == 'en_cours' and (real_s1_total is not None or prev_s2_total is not None):
#             prev_cloture_total = self._safe_sum([real_s1_total, prev_s2_total])
#             prev_cloture_dex   = self._safe_sum([real_s1_dex, prev_s2_dex])
#             cout_total = self._safe_sum([real_cumul_total, prev_cloture_total, prev_n_plus1_total, rar_total])
#             cout_dex   = self._safe_sum([real_cumul_dex, prev_cloture_dex, prev_n_plus1_dex, rar_dex])
#         else:
#             prev_cloture_total = None
#             prev_cloture_dex   = None
#             cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
#             cout_dex   = self._safe_sum([prev_n_plus1_dex, rar_dex])

#         upload = ExcelUpload.objects.create(
#             file_name=f'projet_{new_code_division}_v{new_version_number}', status='processed'
#         )

#         # Récupérer les valeurs des champs AUTO_FIELDS depuis old_version
#         auto_values = {}
#         for _, model_attr, front_key in self.AUTO_FIELDS:
#             # Priorité à new_data (injecté), sinon old_version
#             auto_values[model_attr] = new_data.get(front_key, getattr(old_version, model_attr))

#         # Construction du dictionnaire des arguments
#         create_kwargs = {
#             'upload': upload,
#             'code_division': new_code_division,
#             'version': new_version_number,
#             'is_active': True,
#             'version_comment': new_data.get('version_comment', f'Version {new_version_number}'),
#             'created_by': request.user.id,
#             'parent_id': old_version.parent_id or old_version.id,

#             # Champs communs
#             'libelle': new_data.get('libelle', old_version.libelle),
#             'description_technique': new_data.get('description_technique', old_version.description_technique),
#             'opportunite_projet': new_data.get('opportunite_projet', old_version.opportunite_projet),
#             'type_projet': projet_type,
#             'annee_debut_pmt': new_data.get('annee_debut_pmt', old_version.annee_debut_pmt),
#             'annee_fin_pmt': new_data.get('annee_fin_pmt', old_version.annee_fin_pmt),

#             # Champs métier - avec prise en compte des AUTO_FIELDS
#             'activite': auto_values.get('activite', old_version.activite),
#             'famille': auto_values.get('famille', old_version.famille),
#             'region': auto_values.get('region', old_version.region),
#             'direction': auto_values.get('direction', old_version.direction),
#             'perm': auto_values.get('perm', old_version.perm),

#             # IDs utilisateur - CRUCIAL : garder les IDs de l'ancienne version
#             'region_id': old_version.region_id,
#             'structure_id': old_version.structure_id,
#             'direction_id': old_version.direction_id,
#             'departement_id': old_version.departement_id,

#             # Financiers
#             'realisation_cumul_n_mins1_total': real_cumul_total,
#             'realisation_cumul_n_mins1_dont_dex': real_cumul_dex,
#             'real_s1_n_total': real_s1_total,
#             'real_s1_n_dont_dex': real_s1_dex,
#             'prev_s2_n_total': prev_s2_total,
#             'prev_s2_n_dont_dex': prev_s2_dex,
#             'prev_cloture_n_total': prev_cloture_total,
#             'prev_cloture_n_dont_dex': prev_cloture_dex,
#             'prev_n_plus1_total': prev_n_plus1_total,
#             'prev_n_plus1_dont_dex': prev_n_plus1_dex,
#             'reste_a_realiser_total': rar_total,
#             'reste_a_realiser_dont_dex': rar_dex,
#             'cout_initial_total': cout_total,
#             'cout_initial_dont_dex': cout_dex,
#         }

#         # Ajout des prévisions mensuelles et annuelles
#         for k in self._all_financial_keys(self.PREVISIONS_KEYS, self.MOIS_KEYS):
#             create_kwargs[k] = v[k]

#         return BudgetRecord.objects.create(**create_kwargs)

# ══════════════════════════════════════════════════════════════════════════════
# BASE MODIFIER PROJET VIEW (version corrigée)
# ══════════════════════════════════════════════════════════════════════════════

class BaseModifierProjetView(APIView):
    """Classe de base pour la modification de projet (responsable)"""

    authentication_classes = [RemoteJWTAuthentication]

    PREVISIONS_KEYS = [
        'prev_n_plus2', 'prev_n_plus3', 'prev_n_plus4', 'prev_n_plus5'
    ]
    MOIS_KEYS = [
        'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
        'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre',
    ]
    LOCK_TIMEOUT  = 10
    LOCK_WAIT     = 8
    LOCK_INTERVAL = 0.1

    # Champs que le responsable ne peut PAS modifier — récupérés automatiquement
    # À surcharger dans les sous-classes
    AUTO_FIELDS = []  # Liste de tuples (data_key, model_attr, front_key)

    @staticmethod
    def _to_decimal_or_none(val):
        if val in (None, '', 'null', 'None'):
            return None
        try:
            return Decimal(str(val))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _safe_sum(values):
        filtered = [v for v in values if v is not None]
        if not filtered:
            return None
        return sum(
            v if isinstance(v, Decimal) else Decimal(str(v))
            for v in filtered
        )

    @staticmethod
    def _all_financial_keys(previsions_keys, mois_keys):
        keys = []
        for k in previsions_keys:
            keys += [f'{k}_total', f'{k}_dont_dex']
        for m in mois_keys:
            keys += [f'{m}_total', f'{m}_dont_dex']
        return keys

    def _parse_financial_fields(self, data, previsions_keys, mois_keys):
        v = {}
        for k in previsions_keys:
            v[f'{k}_total']    = self._to_decimal_or_none(data.get(f'{k}_total'))
            v[f'{k}_dont_dex'] = self._to_decimal_or_none(data.get(f'{k}_dont_dex'))
        for m in mois_keys:
            v[f'{m}_total']    = self._to_decimal_or_none(data.get(f'{m}_total'))
            v[f'{m}_dont_dex'] = self._to_decimal_or_none(data.get(f'{m}_dont_dex'))
        return v

    def _acquire_lock(self, code_division):
        lock_key = f'budget_lock_{code_division}'
        deadline = time.time() + self.LOCK_WAIT
        while time.time() < deadline:
            if cache.add(lock_key, '1', timeout=self.LOCK_TIMEOUT):
                return True
            time.sleep(self.LOCK_INTERVAL)
        return False

    def _release_lock(self, code_division):
        cache.delete(f'budget_lock_{code_division}')

    def get(self, request, code_division):
        actif = (
            BudgetRecord.objects.filter(
                code_division=code_division, is_active=True
            ).first()
            or BudgetRecord.objects.filter(
                code_division=code_division
            ).order_by('-version').first()
        )
        if not actif:
            return Response(
                {'error': f'Projet {code_division} introuvable.'},
                status=404
            )

        serializer = BudgetRecordSerializer(actif)

        auto_values = {}
        for data_key, model_attr, front_key in self.AUTO_FIELDS:
            auto_values[front_key] = getattr(actif, model_attr)

        return Response({
            'code_division':  code_division,
            'champs_auto':    auto_values,
            'version_active': serializer.data,
        })

    def post(self, request, code_division):
        if not BudgetRecord.objects.filter(code_division=code_division).exists():
            return Response(
                {'error': f'Projet {code_division} introuvable. Utilisez la création.'},
                status=404
            )

        if not self._acquire_lock(code_division):
            return Response(
                {'error': 'Une modification est déjà en cours. Réessayez dans quelques secondes.'},
                status=409
            )
        try:
            return self._do_modification(request, code_division)
        finally:
            self._release_lock(code_division)

    def _do_modification(self, request, code_division):
        data = request.data.copy()

        with transaction.atomic():
            old_version = (
                BudgetRecord.objects.filter(
                    code_division=code_division, is_active=True
                ).first()
                or BudgetRecord.objects.filter(
                    code_division=code_division
                ).order_by('-version').first()
            )
            if not old_version:
                return Response(
                    {'error': f'Projet {code_division} non trouvé'},
                    status=404
                )

            # Injection automatique des champs protégés
            for data_key, model_attr, front_key in self.AUTO_FIELDS:
                data[front_key] = getattr(old_version, model_attr)

            new_code_division = code_division
            result = BudgetRecord.objects.filter(
                code_division=new_code_division
            ).aggregate(max_version=Max('version'))
            new_version_number = (result['max_version'] or 0) + 1

            if old_version.is_active:
                old_version.is_active = False
                old_version.save(update_fields=['is_active'])

            new_version = self._create_new_version(
                old_version, data, request,
                new_code_division, new_version_number
            )

        serializer = BudgetRecordSerializer(new_version)
        return Response(
            {
                'success': True,
                'message': f'Projet modifié — Version {old_version.version} → {new_version_number}',
                'ancienne_version': old_version.version,
                'nouvelle_version': new_version_number,
                'data': serializer.data,
            },
            status=201,
        )

    def _create_new_version(self, old_version, new_data, request, new_code_division, new_version_number):
        v = {}
        for key in self.PREVISIONS_KEYS:
            for suffix in ('_total', '_dont_dex'):
                k = f'{key}{suffix}'
                v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)
        for mois in self.MOIS_KEYS:
            for suffix in ('_total', '_dont_dex'):
                k = f'{mois}{suffix}'
                v[k] = self._to_decimal_or_none(new_data.get(k)) or getattr(old_version, k)

        def nd(key):
            raw = self._to_decimal_or_none(new_data.get(key))
            return raw if raw is not None else getattr(old_version, key)

        real_cumul_total = nd('realisation_cumul_n_mins1_total')
        real_cumul_dex   = nd('realisation_cumul_n_mins1_dont_dex')
        real_s1_total    = nd('real_s1_n_total')
        real_s1_dex      = nd('real_s1_n_dont_dex')
        prev_s2_total    = nd('prev_s2_n_total')
        prev_s2_dex      = nd('prev_s2_n_dont_dex')

        prev_n_plus1_total = self._safe_sum([v[f'{m}_total'] for m in self.MOIS_KEYS])
        prev_n_plus1_dex   = self._safe_sum([v[f'{m}_dont_dex'] for m in self.MOIS_KEYS])
        rar_total          = self._safe_sum([v[f'{k}_total'] for k in self.PREVISIONS_KEYS])
        rar_dex            = self._safe_sum([v[f'{k}_dont_dex'] for k in self.PREVISIONS_KEYS])

        projet_type = new_data.get('type_projet', old_version.type_projet)

        if projet_type == 'en_cours' and (real_s1_total is not None or prev_s2_total is not None):
            prev_cloture_total = self._safe_sum([real_s1_total, prev_s2_total])
            prev_cloture_dex   = self._safe_sum([real_s1_dex, prev_s2_dex])
            cout_total = self._safe_sum([real_cumul_total, prev_cloture_total, prev_n_plus1_total, rar_total])
            cout_dex   = self._safe_sum([real_cumul_dex, prev_cloture_dex, prev_n_plus1_dex, rar_dex])
        else:
            prev_cloture_total = None
            prev_cloture_dex   = None
            cout_total = self._safe_sum([prev_n_plus1_total, rar_total])
            cout_dex   = self._safe_sum([prev_n_plus1_dex, rar_dex])

        upload = ExcelUpload.objects.create(
            file_name=f'projet_{new_code_division}_v{new_version_number}', status='processed'
        )

        # Récupérer les valeurs des champs AUTO_FIELDS depuis old_version
        auto_values = {}
        for _, model_attr, front_key in self.AUTO_FIELDS:
            # Priorité à new_data (injecté dans _do_modification), sinon old_version
            if front_key in new_data and new_data.get(front_key) is not None:
                auto_values[model_attr] = new_data.get(front_key)
            else:
                auto_values[model_attr] = getattr(old_version, model_attr)

        # Construction du dictionnaire des arguments
        create_kwargs = {
            'upload': upload,
            'code_division': new_code_division,
            'version': new_version_number,
            'is_active': True,
            'version_comment': new_data.get('version_comment', f'Version {new_version_number}'),
            'created_by': request.user.id,
            'parent_id': old_version.id,  # ← Pointer vers la version qu'on remplace

            # Champs communs
            'libelle': new_data.get('libelle', old_version.libelle),
            'description_technique': new_data.get('description_technique', old_version.description_technique),
            'opportunite_projet': new_data.get('opportunite_projet', old_version.opportunite_projet),
            'type_projet': projet_type,
            'annee_debut_pmt': new_data.get('annee_debut_pmt', old_version.annee_debut_pmt),
            'annee_fin_pmt': new_data.get('annee_fin_pmt', old_version.annee_fin_pmt),

            # Champs métier - avec prise en compte des AUTO_FIELDS
            'activite': auto_values.get('activite', old_version.activite),
            'famille': auto_values.get('famille', old_version.famille),
            'region': auto_values.get('region', old_version.region),
            'direction': auto_values.get('direction', old_version.direction),
            'perm': auto_values.get('perm', old_version.perm),

            # IDs utilisateur - TOUJOURS ceux de old_version (la version modifiée)
            'region_id': old_version.region_id,
            'structure_id': old_version.structure_id,
            'direction_id': old_version.direction_id,
            'departement_id': old_version.departement_id,

            # Financiers
            'realisation_cumul_n_mins1_total': real_cumul_total,
            'realisation_cumul_n_mins1_dont_dex': real_cumul_dex,
            'real_s1_n_total': real_s1_total,
            'real_s1_n_dont_dex': real_s1_dex,
            'prev_s2_n_total': prev_s2_total,
            'prev_s2_n_dont_dex': prev_s2_dex,
            'prev_cloture_n_total': prev_cloture_total,
            'prev_cloture_n_dont_dex': prev_cloture_dex,
            'prev_n_plus1_total': prev_n_plus1_total,
            'prev_n_plus1_dont_dex': prev_n_plus1_dex,
            'reste_a_realiser_total': rar_total,
            'reste_a_realiser_dont_dex': rar_dex,
            'cout_initial_total': cout_total,
            'cout_initial_dont_dex': cout_dex,
        }

        # Ajout des prévisions mensuelles et annuelles
        for k in self._all_financial_keys(self.PREVISIONS_KEYS, self.MOIS_KEYS):
            create_kwargs[k] = v[k]

        return BudgetRecord.objects.create(**create_kwargs)
class ModifierProjetStructureView(BaseModifierProjetView):
    """
    GET  /api/budget/responsable/modifier-projet/structure/{code_division}/
    POST /api/budget/responsable/modifier-projet/structure/{code_division}/
    
    Permet à un responsable_structure de modifier un projet.
    Les champs région, périmètre, famille, activité sont automatiquement
    injectés depuis la version active (non modifiables).
    """
    permission_classes = [IsResponsableStructure]

    AUTO_FIELDS = [
        ('region',    'region',   'region'),
        ('perimetre', 'perm',     'perimetre'),
        ('famille',   'famille',  'famille'),
        ('activite',  'activite', 'activite'),
    ]


class ModifierProjetDepartementView(BaseModifierProjetView):
    """
    GET  /api/budget/responsable/modifier-projet/departement/{code_division}/
    POST /api/budget/responsable/modifier-projet/departement/{code_division}/
    
    Permet à un responsable_departement de modifier un projet.
    Les champs direction, famille, activité sont automatiquement
    injectés depuis la version active (non modifiables).
    Note : le périmètre (perm) n'existe pas pour les départements.
    """
    permission_classes = [IsResponsableDepartement]

    AUTO_FIELDS = [
        ('direction', 'direction', 'direction'),
        ('famille',   'famille',   'famille'),
        ('activite',  'activite',  'activite'),
    ]

############################################################################################################
# ================================================================== #
#  PATCH — Admin : modification simple (1 ou plusieurs champs)
#  Pas de nouvelle version — mise à jour directe de la version active
# ================================================================== #
# class PatchProjetAdminView(APIView):
#     """
#     PATCH /api/budget/admin/patch-projet/{code_division}/

#     Permet à l'admin de modifier un ou plusieurs champs directement
#     sur la version active, sans créer de nouvelle version.

#     Exemple body :
#         { "libelle": "Nouveau libellé" }
#         { "region": "Nord", "famille": "F2" }
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAdmin]

#     PATCHABLE_FIELDS = {
#         # Identitaires
#         'region', 'perimetre', 'famille', 'activite',
#         'libelle', 'type_projet', 'code_division',
#         'annee_debut_pmt', 'annee_fin_pmt',
#         'description_technique', 'opportunite_projet',
#         'version_comment',
#         # Financiers — prévisions
#         'prev_n_plus2_total', 'prev_n_plus2_dont_dex',
#         'prev_n_plus3_total', 'prev_n_plus3_dont_dex',
#         'prev_n_plus4_total', 'prev_n_plus4_dont_dex',
#         'prev_n_plus5_total', 'prev_n_plus5_dont_dex',
#         # Financiers — mois
#         'janvier_total',   'janvier_dont_dex',
#         'fevrier_total',   'fevrier_dont_dex',
#         'mars_total',      'mars_dont_dex',
#         'avril_total',     'avril_dont_dex',
#         'mai_total',       'mai_dont_dex',
#         'juin_total',      'juin_dont_dex',
#         'juillet_total',   'juillet_dont_dex',
#         'aout_total',      'aout_dont_dex',
#         'septembre_total', 'septembre_dont_dex',
#         'octobre_total',   'octobre_dont_dex',
#         'novembre_total',  'novembre_dont_dex',
#         'decembre_total',  'decembre_dont_dex',
#         # Réalisations
#         'realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex',
#         'real_s1_n_total',  'real_s1_n_dont_dex',
#         'prev_s2_n_total',  'prev_s2_n_dont_dex',
#     }

#     READONLY_FIELDS = {
#         'id', 'version', 'is_active', 'parent_id',
#         'created_by', 'region_id', 'structure_id', 'upload',
#         'prev_n_plus1_total', 'prev_n_plus1_dont_dex',
#         'reste_a_realiser_total', 'reste_a_realiser_dont_dex',
#         'prev_cloture_n_total', 'prev_cloture_n_dont_dex',
#         'cout_initial_total', 'cout_initial_dont_dex',
#     }

#     DECIMAL_FIELDS = {f for f in PATCHABLE_FIELDS if '_total' in f or '_dont_dex' in f}

#     @staticmethod
#     def _to_decimal_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return Decimal(str(val))
#         except (ValueError, TypeError):
#             return None

#     def patch(self, request, code_division):
#         data = request.data

#         if not data:
#             return Response(
#                 {'error': 'Aucun champ fourni.'},
#                 status=400
#             )

#         # ── Récupérer la version active ─────────────────────────────────
#         actif = (
#             BudgetRecord.objects.filter(
#                 code_division=code_division, is_active=True
#             ).first()
#             or BudgetRecord.objects.filter(
#                 code_division=code_division
#             ).order_by('-version').first()
#         )
#         if not actif:
#             return Response(
#                 {'error': f'Projet {code_division} introuvable.'},
#                 status=404
#             )

#         # ── Vérifier les champs envoyés ─────────────────────────────────
#         unknown_fields = set(data.keys()) - self.PATCHABLE_FIELDS - self.READONLY_FIELDS
#         readonly_sent  = set(data.keys()) & self.READONLY_FIELDS

#         if readonly_sent:
#             return Response(
#                 {
#                     'error': 'Champs système non modifiables via ce endpoint.',
#                     'champs': list(readonly_sent),
#                 },
#                 status=400
#             )

#         if unknown_fields:
#             return Response(
#                 {
#                     'error': 'Champs inconnus.',
#                     'champs': list(unknown_fields),
#                 },
#                 status=400
#             )

#         # ── Appliquer les modifications ─────────────────────────────────
#         updated_fields = []

#         for field, value in data.items():
#             if field not in self.PATCHABLE_FIELDS:
#                 continue

#             if field in self.DECIMAL_FIELDS:
#                 value = self._to_decimal_or_none(value)

#             # Mapping spécial : 'perimetre' → 'perm' sur le modèle
#             model_field = 'perm' if field == 'perimetre' else field

#             setattr(actif, model_field, value)
#             updated_fields.append(model_field)

#         if not updated_fields:
#             return Response(
#                 {'error': 'Aucun champ valide à mettre à jour.'},
#                 status=400
#             )

#         # ✅ FIX : actif.save() est maintenant APRÈS le return guard, pas avant
#         actif.save(update_fields=updated_fields)

#         serializer = BudgetRecordSerializer(
#             actif,
#             context={'request': request}
#         )

#         return Response(
#             {
#                 'success': True,
#                 'message': f'{len(updated_fields)} champ(s) mis à jour sur la version active.',
#                 'version': actif.version,
#                 'champs_modifies': updated_fields,
#                 'data': serializer.data,
#             },
#             status=200,
#         )
class PatchProjetAdminView(APIView):
    """
    PATCH /api/budget/admin/patch-projet/{code_division}/

    Permet à l'admin de modifier un ou plusieurs champs directement
    sur la version active, sans créer de nouvelle version.

    Exemple body :
        { "libelle": "Nouveau libellé" }
        { "region": "Nord", "direction": "DR01", "famille": "F2" }
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAdmin]

    PATCHABLE_FIELDS = {
        # Identitaires
        'region', 'direction', 'perimetre', 'famille', 'activite',
        'libelle', 'type_projet', 'code_division',
        'annee_debut_pmt', 'annee_fin_pmt',
        'description_technique', 'opportunite_projet',
        'version_comment',
        # Financiers — prévisions
        'prev_n_plus2_total', 'prev_n_plus2_dont_dex',
        'prev_n_plus3_total', 'prev_n_plus3_dont_dex',
        'prev_n_plus4_total', 'prev_n_plus4_dont_dex',
        'prev_n_plus5_total', 'prev_n_plus5_dont_dex',
        # Financiers — mois
        'janvier_total',   'janvier_dont_dex',
        'fevrier_total',   'fevrier_dont_dex',
        'mars_total',      'mars_dont_dex',
        'avril_total',     'avril_dont_dex',
        'mai_total',       'mai_dont_dex',
        'juin_total',      'juin_dont_dex',
        'juillet_total',   'juillet_dont_dex',
        'aout_total',      'aout_dont_dex',
        'septembre_total', 'septembre_dont_dex',
        'octobre_total',   'octobre_dont_dex',
        'novembre_total',  'novembre_dont_dex',
        'decembre_total',  'decembre_dont_dex',
        # Réalisations
        'realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex',
        'real_s1_n_total',  'real_s1_n_dont_dex',
        'prev_s2_n_total',  'prev_s2_n_dont_dex',
    }

    READONLY_FIELDS = {
        'id', 'version', 'is_active', 'parent_id',
        'created_by', 'region_id', 'structure_id', 'direction_id', 'departement_id',
        'upload',
        'prev_n_plus1_total', 'prev_n_plus1_dont_dex',
        'reste_a_realiser_total', 'reste_a_realiser_dont_dex',
        'prev_cloture_n_total', 'prev_cloture_n_dont_dex',
        'cout_initial_total', 'cout_initial_dont_dex',
    }

    DECIMAL_FIELDS = {f for f in PATCHABLE_FIELDS if '_total' in f or '_dont_dex' in f}

    @staticmethod
    def _to_decimal_or_none(val):
        if val in (None, '', 'null', 'None'):
            return None
        try:
            return Decimal(str(val))
        except (ValueError, TypeError):
            return None

    def patch(self, request, code_division):
        data = request.data

        if not data:
            return Response(
                {'error': 'Aucun champ fourni.'},
                status=400
            )

        # Récupérer la version active
        actif = (
            BudgetRecord.objects.filter(
                code_division=code_division, is_active=True
            ).first()
            or BudgetRecord.objects.filter(
                code_division=code_division
            ).order_by('-version').first()
        )
        if not actif:
            return Response(
                {'error': f'Projet {code_division} introuvable.'},
                status=404
            )

        # Vérifier les champs envoyés
        unknown_fields = set(data.keys()) - self.PATCHABLE_FIELDS - self.READONLY_FIELDS
        readonly_sent = set(data.keys()) & self.READONLY_FIELDS

        if readonly_sent:
            return Response(
                {
                    'error': 'Champs système non modifiables via ce endpoint.',
                    'champs': list(readonly_sent),
                },
                status=400
            )

        if unknown_fields:
            return Response(
                {
                    'error': 'Champs inconnus.',
                    'champs': list(unknown_fields),
                },
                status=400
            )

        # Appliquer les modifications
        updated_fields = []

        for field, value in data.items():
            if field not in self.PATCHABLE_FIELDS:
                continue

            if field in self.DECIMAL_FIELDS:
                value = self._to_decimal_or_none(value)

            # Mapping spécial : 'perimetre' → 'perm' sur le modèle
            if field == 'perimetre':
                model_field = 'perm'
            else:
                model_field = field

            setattr(actif, model_field, value)
            updated_fields.append(model_field)

        if not updated_fields:
            return Response(
                {'error': 'Aucun champ valide à mettre à jour.'},
                status=400
            )

        actif.save(update_fields=updated_fields)

        serializer = BudgetRecordSerializer(
            actif,
            context={'request': request}
        )

        return Response(
            {
                'success': True,
                'message': f'{len(updated_fields)} champ(s) mis à jour sur la version active.',
                'version': actif.version,
                'champs_modifies': updated_fields,
                'data': serializer.data,
            },
            status=200,
        )


# class PatchProjetAdminView(APIView):
#     """
#     PATCH /api/budget/admin/patch-projet/{code_division}/

#     Permet à l'admin de modifier un ou plusieurs champs directement
#     sur la version active, sans créer de nouvelle version.

#     Exemple body :
#         { "libelle": "Nouveau libellé" }
#         { "region": "Nord", "famille": "F2" }
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAdmin]

#     # Champs autorisés au PATCH (tout sauf les champs système)
#     PATCHABLE_FIELDS = {
#         # Identitaires
#         'region', 'perimetre', 'famille', 'activite',
#         'libelle', 'type_projet', 'code_division',
#         'annee_debut_pmt', 'annee_fin_pmt',
#         'description_technique', 'opportunite_projet',
#         'version_comment',
#         # Financiers — prévisions
#         'prev_n_plus2_total', 'prev_n_plus2_dont_dex',
#         'prev_n_plus3_total', 'prev_n_plus3_dont_dex',
#         'prev_n_plus4_total', 'prev_n_plus4_dont_dex',
#         'prev_n_plus5_total', 'prev_n_plus5_dont_dex',
#         # Financiers — mois
#         'janvier_total',   'janvier_dont_dex',
#         'fevrier_total',   'fevrier_dont_dex',
#         'mars_total',      'mars_dont_dex',
#         'avril_total',     'avril_dont_dex',
#         'mai_total',       'mai_dont_dex',
#         'juin_total',      'juin_dont_dex',
#         'juillet_total',   'juillet_dont_dex',
#         'aout_total',      'aout_dont_dex',
#         'septembre_total', 'septembre_dont_dex',
#         'octobre_total',   'octobre_dont_dex',
#         'novembre_total',  'novembre_dont_dex',
#         'decembre_total',  'decembre_dont_dex',
#         # Réalisations
#         'realisation_cumul_n_mins1_total', 'realisation_cumul_n_mins1_dont_dex',
#         'real_s1_n_total',  'real_s1_n_dont_dex',
#         'prev_s2_n_total',  'prev_s2_n_dont_dex',
#     }

#     # Champs système — jamais modifiables via ce endpoint
#     READONLY_FIELDS = {
#         'id', 'version', 'is_active', 'parent_id',
#         'created_by', 'region_id', 'structure_id', 'upload',
#         # Calculés automatiquement
#         'prev_n_plus1_total', 'prev_n_plus1_dont_dex',
#         'reste_a_realiser_total', 'reste_a_realiser_dont_dex',
#         'prev_cloture_n_total', 'prev_cloture_n_dont_dex',
#         'cout_initial_total', 'cout_initial_dont_dex',
#     }

#     DECIMAL_FIELDS = {f for f in PATCHABLE_FIELDS if '_total' in f or '_dont_dex' in f}

#     @staticmethod
#     def _to_decimal_or_none(val):
#         if val in (None, '', 'null', 'None'):
#             return None
#         try:
#             return Decimal(str(val))
#         except (ValueError, TypeError):
#             return None

#     def patch(self, request, code_division):
#         data = request.data

#         if not data:
#             return Response(
#                 {'error': 'Aucun champ fourni.'},
#                 status=400
#             )

#         # ── Récupérer la version active ─────────────────────────────────
#         actif = (
#             BudgetRecord.objects.filter(
#                 code_division=code_division, is_active=True
#             ).first()
#             or BudgetRecord.objects.filter(
#                 code_division=code_division
#             ).order_by('-version').first()
#         )
#         if not actif:
#             return Response(
#                 {'error': f'Projet {code_division} introuvable.'},
#                 status=404
#             )

#         # ── Vérifier les champs envoyés ─────────────────────────────────
#         unknown_fields  = set(data.keys()) - self.PATCHABLE_FIELDS - self.READONLY_FIELDS
#         readonly_sent   = set(data.keys()) & self.READONLY_FIELDS
#         invalid_fields  = set(data.keys()) - self.PATCHABLE_FIELDS

#         if readonly_sent:
#             return Response(
#                 {
#                     'error': 'Champs système non modifiables via ce endpoint.',
#                     'champs': list(readonly_sent),
#                 },
#                 status=400
#             )

#         if unknown_fields:
#             return Response(
#                 {
#                     'error': 'Champs inconnus.',
#                     'champs': list(unknown_fields),
#                 },
#                 status=400
#             )

#         # ── Appliquer les modifications ─────────────────────────────────
#         updated_fields = []

#         for field, value in data.items():
#             if field not in self.PATCHABLE_FIELDS:
#                 continue

#             # Conversion Decimal pour les champs financiers
#             if field in self.DECIMAL_FIELDS:
#                 value = self._to_decimal_or_none(value)

#             # Mapping spécial : 'perimetre' dans la requête → 'perm' sur le modèle
#             model_field = 'perm' if field == 'perimetre' else field

#             setattr(actif, model_field, value)
#             updated_fields.append(model_field)

#         if not updated_fields:
#             return Response(
#                 {'error': 'Aucun champ valide à mettre à jour.'},
#                 status=400
#             )
#             actif.save(update_fields=updated_fields)
    
#     # ✅ CORRECTION ICI - Ajoutez le contexte avec la requête
#         serializer = BudgetRecordSerializer(
#             actif, 
#             context={'request': request}  # ← CRITICAL : Passe le token
#         )
        
#         return Response(
#             {
#                 'success': True,
#                 'message': f'{len(updated_fields)} champ(s) mis à jour sur la version active.',
#                 'version': actif.version,
#                 'champs_modifies': updated_fields,
#                 'data': serializer.data,  # ← Utilisez le serializer avec contexte
#             },
#             status=200,
#         )

        # actif.save(update_fields=updated_fields)

        # return Response(
        #     {
        #         'success':         True,
        #         'message':         f'{len(updated_fields)} champ(s) mis à jour sur la version active.',
        #         'version':         actif.version,
        #         'champs_modifies': updated_fields,
        #         'data':            BudgetRecordSerializer(actif).data,
        #     },
        #     status=200,
        # )
# ─────────────────────────────────────────
# HELPER — récupérer record par id
# ─────────────────────────────────────────

def get_record_or_404(record_id):
    try:
        return BudgetRecord.objects.get(id=record_id)
    except BudgetRecord.DoesNotExist:
        return None


# ─────────────────────────────────────────
# WORKFLOW DE VALIDATION
# ─────────────────────────────────────────


from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response


def get_record_or_404(record_id):
    try:
        return BudgetRecord.objects.get(id=record_id)
    except BudgetRecord.DoesNotExist:
        return None


# # ================================================================== #
# #  1. SOUMETTRE  (ResponsableStructure)
# # ================================================================== #
# class SoumettreProjetView(APIView):
#     """
#     POST /recap/budget/soumettre/<id>/
#     Condition : statut = brouillon
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsResponsableStructure]

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         if record.statut != 'brouillon':
#             return Response({
#                 'error': f"Impossible de soumettre — statut actuel : '{record.statut}'"
#             }, status=400)

#         record.statut = 'soumis'
#         record.save(update_fields=['statut'])

#         return Response({
#             'success': True,
#             'message': 'Projet soumis pour validation',
#             'statut':  record.statut,
#         })


# # ================================================================== #
# #  2. DIRECTEUR RÉGION  (valider / rejeter)
# # ================================================================== #

# class ValiderDirecteurRegionView(APIView):
#     """
#     POST /recap/budget/valider/directeur-region/<id>/
#     Condition : statut = soumis
#                 OU reserve_agent / reserve_chef / reserve_directeur
#                    (retour de réserve vers DR)
#     Actions   : valider | rejeter
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteurRegion]

#     STATUTS_AUTORISÉS = {
#         'soumis',
#         'reserve_agent',
#         'reserve_chef',
#         'reserve_directeur',
#     }

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'rejeter'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'rejeter'"},
#                 status=400
#             )

#         # Vérification que le commentaire est présent pour un rejet
#         if action == 'rejeter' and not commentaire:
#             return Response({
#                 'error': "Le commentaire est obligatoire pour un rejet"
#             }, status=400)

#         if record.statut not in self.STATUTS_AUTORISÉS:
#             return Response({
#                 'error': (
#                     f"Statut '{record.statut}' non autorisé pour cette action. "
#                     f"Statuts acceptés : {', '.join(self.STATUTS_AUTORISÉS)}"
#                 )
#             }, status=400)

#         if action == 'valider':
#             record.statut                           = 'valide_directeur_region'
#             record.valide_par_directeur_region      = request.user.nom_complet
#             record.date_validation_directeur_region = timezone.now()
#             record.commentaire_directeur_region     = commentaire or None
#             message = 'Projet validé par le directeur région'
#         else:  # action == 'rejeter'
#             record.statut                           = 'rejete'
#             record.rejete_par                       = request.user.nom_complet
#             record.date_rejet                       = timezone.now()
#             record.motif_rejet                      = commentaire
#             # ✅ AJOUT: Remplir aussi le commentaire_directeur_region
#             record.commentaire_directeur_region     = commentaire
#             message = 'Projet rejeté par le directeur région'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut':  record.statut,
#             'commentaire': commentaire  # Optionnel: retourner le commentaire
#         })

# # ================================================================== #
# #  3. AGENT  (valider / réserver)
# #     Voit : valide_directeur_region
# #     Ne peut PAS rejeter
# # ================================================================== #
# class ValiderAgentView(APIView):
#     """
#     POST /recap/budget/valider/agent/<id>/
#     Condition : statut = valide_directeur_region
#     Actions   : valider | reserver (avec commentaire obligatoire)
#     Réserver  → retourne au DR (statut = reserve_agent)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsAgent]

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'reserver'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'reserver'"},
#                 status=400
#             )

#         if record.statut != 'valide_directeur_region':
#             return Response({
#                 'error': (
#                     f"Le projet doit être 'valide_directeur_region' "
#                     f"— statut actuel : '{record.statut}'"
#                 )
#             }, status=400)

#         if action == 'reserver' and not commentaire:
#             return Response(
#                 {'error': "Un commentaire est obligatoire pour réserver."},
#                 status=400
#             )

#         if action == 'valider':
#             record.statut = 'valide_agent'
#             message       = 'Projet validé par l\'agent'
#         else:
#             # Réserver → retourne au DR avec commentaire
#             record.statut                           = 'reserve_agent'
#             record.commentaire_agent    = (
#                 f"{commentaire}"
#             )
#             message = 'Projet réservé — retourné au directeur région'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut':  record.statut,
#         })


# # ================================================================== #
# #  4. CHEF  (valider / réserver)
# #     Voit : valide_agent + reserve_agent
# #     Ne peut PAS rejeter
# # ================================================================== #
# class ValiderChefView(APIView):
#     """
#     POST /recap/budget/valider/chef/<id>/
#     Condition : statut = valide_agent OU reserve_agent
#     Actions   : valider | reserver (avec commentaire obligatoire)
#     Réserver  → retourne au DR (statut = reserve_chef)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsChef]

#     STATUTS_AUTORISÉS = {'valide_agent', 'reserve_agent'}

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'reserver'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'reserver'"},
#                 status=400
#             )

#         if record.statut not in self.STATUTS_AUTORISÉS:
#             return Response({
#                 'error': (
#                     f"Statut '{record.statut}' non autorisé. "
#                     f"Acceptés : {', '.join(self.STATUTS_AUTORISÉS)}"
#                 )
#             }, status=400)

#         if action == 'reserver' and not commentaire:
#             return Response(
#                 {'error': "Un commentaire est obligatoire pour réserver."},
#                 status=400
#             )

#         if action == 'valider':
#             record.statut               = 'valide_chef'
#             record.valide_par_chef      = request.user.nom_complet
#             record.date_validation_chef = timezone.now()
#             record.commentaire_chef     = commentaire
#             message = 'Projet validé par le chef'
#         else:
#             record.statut           = 'reserve_chef'
#             record.commentaire_chef = f"{commentaire}"
#             message = 'Projet réservé — retourné au directeur région'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut':  record.statut,
#         })



# # ================================================================== #
# #  5. DIRECTEUR  (valider / réserver)
# #     Condition : valide_chef
# #     Ne peut PAS rejeter
# # ================================================================== #
# class ValiderDirecteurView(APIView):
#     """
#     POST /recap/budget/valider/directeur/<id>/
#     Condition : statut = valide_chef
#     Actions   : valider | reserver (avec commentaire obligatoire)
#     Réserver  → retourne au DR (statut = reserve_directeur)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteur]

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'reserver'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'reserver'"},
#                 status=400
#             )

#         if record.statut != 'valide_chef':
#             return Response({
#                 'error': (
#                     f"Le projet doit être 'valide_chef' "
#                     f"— statut actuel : '{record.statut}'"
#                 )
#             }, status=400)

#         if action == 'reserver' and not commentaire:
#             return Response(
#                 {'error': "Un commentaire est obligatoire pour réserver."},
#                 status=400
#             )

#         if action == 'valider':
#             record.statut                    = 'valide_directeur'
#             record.valide_par_directeur      = request.user.nom_complet
#             record.date_validation_directeur = timezone.now()
#             record.commentaire_directeur     = commentaire
#             message = 'Projet validé par le directeur'
#         else:
#             record.statut                = 'reserve_directeur'
#             record.commentaire_directeur = f"[Réservé par directeur] {commentaire}"
#             message = 'Projet réservé — retourné au directeur région'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut':  record.statut,
#         })


# # ================================================================== #
# #  6. DIVISIONNAIRE  (valider / rejeter)
# #     Condition : valide_directeur
# #     Seul à pouvoir rejeter après le DR
# # ================================================================== #
# class ValiderDivisionnnaireView(APIView):
#     """
#     POST /recap/budget/valider/divisionnaire/<id>/
#     Condition : statut = valide_directeur
#     Actions   : valider | rejeter
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDivisionnaire]

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'rejeter'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'rejeter'"},
#                 status=400
#             )

#         if record.statut != 'valide_directeur':
#             return Response({
#                 'error': (
#                     f"Le projet doit être 'valide_directeur' "
#                     f"— statut actuel : '{record.statut}'"
#                 )
#             }, status=400)

#         if action == 'valider':
#             record.statut                        = 'valide_divisionnaire'
#             record.valide_par_divisionnaire      = request.user.nom_complet
#             record.date_validation_divisionnaire = timezone.now()
#             record.commentaire_divisionnaire     = commentaire
#             message = 'Projet validé par le divisionnaire — validation complète ✅'
#         else:
#             record.statut      = 'rejete'
#             record.rejete_par  = request.user.nom_complet
#             record.date_rejet  = timezone.now()
#             record.motif_rejet = commentaire
#             message = 'Projet rejeté par le divisionnaire'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut':  record.statut,
#         })

# #validation par total:
# class ValiderTousProjetsDivisionnaireView(APIView):
#     """
#     GET  /recap/budget/valider/divisionnaire/tous/
#          → Retourne tous les projets 'valide_directeur' de l'année prochaine + leur total

#     POST /recap/budget/valider/divisionnaire/tous/
#          → action: 'valider' | 'rejeter'
#          → Change le statut de TOUS les projets 'valide_directeur' en masse
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDivisionnaire]

#     def get(self, request):
#         next_year = datetime.now().year + 1

#         projets = BudgetRecord.objects.filter(
#             statut='valide_directeur',
#             annee_debut_pmt=next_year,
#         )

#         if not projets.exists():
#             return Response({
#                 'success': True,
#                 'count': 0,
#                 'projets': [],
#                 'total': {},
#                 'message': f'Aucun projet à valider pour {next_year}',
#             })

#         # Total de toutes les colonnes
#         total = projets.aggregate(**build_aggregation())

#         serializer = BudgetRecordSerializer(
#             projets, many=True, context={'request': request}
#         )

#         return Response({
#             'success':    True,
#             'count':      projets.count(),
#             'annee':      next_year,
#             'projets':    serializer.data,
#             'total':      total,
#         })

#     def post(self, request):
#         next_year   = datetime.now().year + 1
#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'rejeter'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'rejeter'"},
#                 status=400
#             )

#         projets = BudgetRecord.objects.filter(
#             statut='valide_directeur',
#             annee_debut_pmt=next_year,
#         )

#         if not projets.exists():
#             return Response({
#                 'error': f'Aucun projet avec statut valide_directeur pour {next_year}',
#             }, status=404)

#         count = projets.count()
#         now   = timezone.now()
#         nom   = request.user.nom_complet

#         if action == 'valider':
#             projets.update(
#                 statut                        = 'valide_divisionnaire',
#                 valide_par_divisionnaire      = nom,
#                 date_validation_divisionnaire = now,
#                 commentaire_divisionnaire     = commentaire,
#             )
#             message = f'{count} projet(s) validés par le divisionnaire ✅'

#         else:  # rejeter
#             projets.update(
#                 statut     = 'rejete',
#                 rejete_par  = nom,
#                 date_rejet  = now,
#                 motif_rejet = commentaire,
#             )
#             message = f'{count} projet(s) rejetés par le divisionnaire ❌'

#         return Response({
#             'success': True,
#             'message': message,
#             'action':  action,
#             'count':   count,
#             'annee':   next_year,
#         })
# views.py - WORKFLOW DE VALIDATION AVEC NOUVEAUX STATUTS

from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime
import requests


def get_record_or_404(record_id):
    try:
        return BudgetRecord.objects.get(id=record_id)
    except BudgetRecord.DoesNotExist:
        return None


# ================================================================== #
#  1. SOUMETTRE (ResponsableStructure)
# ================================================================== #
class SoumettreProjetView(APIView):
    """
    POST /recap/budget/soumettre/<id>/
    Condition : statut = brouillon (ni statut_workflow ni statut_final)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsResponsableStructure]

    def post(self, request, record_id):
        record = get_record_or_404(record_id)
        if not record:
            return Response({'error': 'Projet introuvable'}, status=404)

        # Vérifier que le projet est en brouillon
        if record.statut_workflow is not None or record.statut_final is not None:
            return Response({
                'error': f"Impossible de soumettre — statut actuel : '{record.statut}'"
            }, status=400)

        record.statut_workflow = 'soumis'
        record.save()

        return Response({
            'success': True,
            'message': 'Projet soumis pour validation',
            'statut_workflow': record.statut_workflow,
            'statut_final': record.statut_final,
        })


# ================================================================== #
#  2. DIRECTEUR RÉGION (valider / rejeter)
# ================================================================== #
# class ValiderDirecteurRegionView(APIView):
#     """
#     POST /recap/budget/valider/directeur-region/<id>/
#     Condition : statut_workflow = soumis
#                 OU statut_workflow = reserve_chef (retour de réserve)
#                 OU statut_workflow = reserve_directeur (retour de réserve)
#     Actions   : valider | rejeter
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteurRegion]

#     STATUTS_AUTORISÉS = {
#         'soumis',
#         'reserve_chef',
#         'reserve_directeur',
#     }

#     def post(self, request, record_id):
#         record = get_record_or_404(record_id)
#         if not record:
#             return Response({'error': 'Projet introuvable'}, status=404)

#         action      = request.data.get('action')
#         commentaire = request.data.get('commentaire', '')

#         if action not in ('valider', 'rejeter'):
#             return Response(
#                 {'error': "action doit être 'valider' ou 'rejeter'"},
#                 status=400
#             )

#         if action == 'rejeter' and not commentaire:
#             return Response({
#                 'error': "Le commentaire est obligatoire pour un rejet"
#             }, status=400)

#         if record.statut_workflow not in self.STATUTS_AUTORISÉS:
#             return Response({
#                 'error': (
#                     f"Statut '{record.statut_workflow}' non autorisé pour cette action. "
#                     f"Statuts acceptés : {', '.join(self.STATUTS_AUTORISÉS)}"
#                 )
#             }, status=400)

#         if action == 'valider':
#             record.statut_final = 'valide_directeur_region'
#             record.statut_workflow = None
#             record.valide_par_directeur_region = request.user.nom_complet
#             record.date_validation_directeur_region = timezone.now()
#             record.commentaire_directeur_region = commentaire or None
#             message = 'Projet validé par le directeur région'
#         else:  # action == 'rejeter'
#             record.statut_final = 'rejete_directeur_region'
#             record.statut_workflow = None
#             record.rejete_par_directeur_region = request.user.nom_complet
#             record.date_rejet_directeur_region = timezone.now()
#             record.motif_rejet_directeur_region = commentaire
#             # Champs génériques pour compatibilité
#             record.rejete_par = request.user.nom_complet
#             record.date_rejet = timezone.now()
#             record.motif_rejet = commentaire
#             record.commentaire_directeur_region = commentaire
#             message = 'Projet rejeté par le directeur région'

#         record.save()
#         return Response({
#             'success': True,
#             'message': message,
#             'statut_final': record.statut_final,
#             'commentaire': commentaire
#         })
from rest_framework.views import APIView
from rest_framework.response import Response
from django.utils import timezone

class ValiderDirecteurRegionView(APIView):
    """
    POST /recap/budget/valider/directeur-region/<id>/

    Autorisé si :
        - statut_workflow = soumis
        - OU statut_workflow = reserve_directeur
        - OU statut_final = annule_divisionnaire
        - OU statut_final = rejete_divisionnaire

    Actions : valider | rejeter
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    # statuts workflow autorisés
    ALLOWED_WORKFLOW = {
        'soumis',
        'reserve_directeur',
    }

    # statuts finaux autorisés
    ALLOWED_FINAL = {
        'annule_divisionnaire',
        'rejete_divisionnaire',
    }

    def post(self, request, record_id):
        record = get_record_or_404(record_id)

        if not record:
            return Response({'error': 'Projet introuvable'}, status=404)

        action = request.data.get('action')
        commentaire = request.data.get('commentaire', '')

        # validation action
        if action not in ('valider', 'rejeter'):
            return Response(
                {'error': "action doit être 'valider' ou 'rejeter'"},
                status=400
            )

        # commentaire obligatoire pour rejet
        if action == 'rejeter' and not commentaire:
            return Response(
                {'error': "Le commentaire est obligatoire pour un rejet"},
                status=400
            )

        # ==========================
        # VALIDATION STATUTS
        # ==========================

        statut_workflow_ok = record.statut_workflow in self.ALLOWED_WORKFLOW
        statut_final_ok = record.statut_final in self.ALLOWED_FINAL

        if not (statut_workflow_ok or statut_final_ok):
            return Response({
                'error': (
                    "Statut non autorisé pour le directeur région. "
                    f"Workflow: {record.statut_workflow}, Final: {record.statut_final}"
                )
            }, status=400)

        # ==========================
        # ACTION = VALIDER
        # ==========================
        if action == 'valider':
            record.statut_final = 'valide_directeur_region'
            record.statut_workflow = None

            record.valide_par_directeur_region = request.user.nom_complet
            record.date_validation_directeur_region = timezone.now()
            record.commentaire_directeur_region = commentaire or None

            message = 'Projet validé par le directeur région'

        # ==========================
        # ACTION = REJETER
        # ==========================
        else:
            record.statut_final = 'rejete_directeur_region'
            record.statut_workflow = None

            record.rejete_par_directeur_region = request.user.nom_complet
            record.date_rejet_directeur_region = timezone.now()
            record.motif_rejet_directeur_region = commentaire

            # compatibilité champs génériques
            record.rejete_par = request.user.nom_complet
            record.date_rejet = timezone.now()
            record.motif_rejet = commentaire
            record.commentaire_directeur_region = commentaire

            message = 'Projet rejeté par le directeur région'

        record.save()

        return Response({
            'success': True,
            'message': message,
            'statut_final': record.statut_final,
            'statut_workflow': record.statut_workflow,
            'commentaire': commentaire
        })

# ================================================================== #
#  3. CHEF (pre_approuve / reserve)
#     Voit : valide_directeur_region
# ================================================================== #
class ValiderChefView(APIView):
    """
    POST /recap/budget/valider/chef/<id>/
    Condition : statut_final = valide_directeur_region
    Actions   : pre_approuver | reserver (avec commentaire obligatoire pour reserve)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsChef]

    def post(self, request, record_id):
        record = get_record_or_404(record_id)
        if not record:
            return Response({'error': 'Projet introuvable'}, status=404)

        action      = request.data.get('action')
        commentaire = request.data.get('commentaire', '')

        if action not in ('pre_approuver', 'reserver'):
            return Response(
                {'error': "action doit être 'pre_approuver' ou 'reserver'"},
                status=400
            )

        if record.statut_final != 'valide_directeur_region':
            return Response({
                'error': (
                    f"Le projet doit avoir statut_final='valide_directeur_region' "
                    f"— statut actuel : '{record.statut_final}'"
                )
            }, status=400)

        if action == 'reserver' and not commentaire:
            return Response(
                {'error': "Un commentaire est obligatoire pour réserver."},
                status=400
            )

        if action == 'pre_approuver':
            record.statut_workflow = 'pre_approuve_chef'
            record.statut_final = None
            record.preapprouve_par_chef = request.user.nom_complet
            record.date_preapprouve_chef = timezone.now()
            record.commentaire_preapprouve_chef = commentaire
            message = 'Projet pré-approuvé par le chef'
        else:  # action == 'reserver'
            record.statut_workflow = 'reserve_chef'
            record.statut_final = None
            record.reserve_par_chef = request.user.nom_complet
            record.date_reserve_chef = timezone.now()
            record.commentaire_reserve_chef = commentaire
            message = 'Projet réservé par le chef — retourné au directeur région'

        record.save()
        return Response({
            'success': True,
            'message': message,
            'statut_workflow': record.statut_workflow,
        })


# ================================================================== #
#  4. DIRECTEUR NATIONAL (approuve / reserve)
#     Voit : pre_approuve_chef, reserve_chef
# ================================================================== #
class ValiderDirecteurView(APIView):
    """
    POST /recap/budget/valider/directeur/<id>/
    Condition : statut_workflow = pre_approuve_chef OU reserve_chef
    Actions   : approuver | reserver (avec commentaire obligatoire pour reserve)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsDirecteur]

    STATUTS_AUTORISÉS = {'pre_approuve_chef', 'reserve_chef'}

    def post(self, request, record_id):
        record = get_record_or_404(record_id)
        if not record:
            return Response({'error': 'Projet introuvable'}, status=404)

        action      = request.data.get('action')
        commentaire = request.data.get('commentaire', '')

        if action not in ('approuver', 'reserver'):
            return Response(
                {'error': "action doit être 'approuver' ou 'reserver'"},
                status=400
            )

        if record.statut_workflow not in self.STATUTS_AUTORISÉS:
            return Response({
                'error': (
                    f"Statut '{record.statut_workflow}' non autorisé. "
                    f"Acceptés : {', '.join(self.STATUTS_AUTORISÉS)}"
                )
            }, status=400)

        if action == 'reserver' and not commentaire:
            return Response(
                {'error': "Un commentaire est obligatoire pour réserver."},
                status=400
            )

        if action == 'approuver':
            record.statut_workflow = 'approuve_directeur'
            record.statut_final = None
            record.approuve_par_directeur = request.user.nom_complet
            record.date_approuve_directeur = timezone.now()
            record.commentaire_approuve_directeur = commentaire
            message = 'Projet approuvé par le directeur'
        else:  # action == 'reserve'
            record.statut_workflow = 'reserve_directeur'
            record.statut_final = None
            record.reserve_par_directeur = request.user.nom_complet
            record.date_reserve_directeur = timezone.now()
            record.commentaire_reserve_directeur = commentaire
            message = 'Projet réservé par le directeur — retourné au directeur région'

        record.save()
        return Response({
            'success': True,
            'message': message,
            'statut_workflow': record.statut_workflow,
        })


# ================================================================== #
#  5. DIVISIONNAIRE (valide / rejete / annule)
#     Condition : approuve_directeur
# ================================================================== #
class ValiderDivisionnaireView(APIView):
    """
    POST /recap/budget/valider/divisionnaire/<id>/
    Condition : statut_workflow = approuve_directeur
    Actions   : valider | rejeter | annuler
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsDivisionnaire]

    def post(self, request, record_id):
        record = get_record_or_404(record_id)
        if not record:
            return Response({'error': 'Projet introuvable'}, status=404)

        action      = request.data.get('action')
        commentaire = request.data.get('commentaire', '')

        if action not in ('valider', 'rejeter', 'annuler'):
            return Response(
                {'error': "action doit être 'valider', 'rejeter' ou 'annuler'"},
                status=400
            )

        if record.statut_workflow != 'approuve_directeur':
            return Response({
                'error': (
                    f"Le projet doit avoir statut_workflow='approuve_directeur' "
                    f"— statut actuel : '{record.statut_workflow}'"
                )
            }, status=400)

        if action in ('rejeter', 'annuler') and not commentaire:
            return Response({
                'error': f"Un commentaire est obligatoire pour {action}."
            }, status=400)

        if action == 'valider':
            record.statut_final = 'valide_divisionnaire'
            record.statut_workflow = None
            record.valide_par_divisionnaire = request.user.nom_complet
            record.date_validation_divisionnaire = timezone.now()
            record.commentaire_divisionnaire = commentaire
            message = 'Projet validé par le divisionnaire — validation complète ✅'
        
        elif action == 'rejeter':
            record.statut_final = 'rejete_divisionnaire'
            record.statut_workflow = None
            record.rejete_par_divisionnaire = request.user.nom_complet
            record.date_rejet_divisionnaire = timezone.now()
            record.motif_rejet_divisionnaire = commentaire
            # Champs génériques pour compatibilité
            record.rejete_par = request.user.nom_complet
            record.date_rejet = timezone.now()
            record.motif_rejet = commentaire
            message = 'Projet rejeté par le divisionnaire ❌'
        
        else:  # action == 'annuler'
            record.statut_final = 'annule_divisionnaire'
            record.statut_workflow = None
            record.annule_par_divisionnaire = request.user.nom_complet
            record.date_annulation_divisionnaire = timezone.now()
            record.motif_annulation_divisionnaire = commentaire
            message = 'Projet annulé par le divisionnaire 🚫'

        record.save()
        return Response({
            'success': True,
            'message': message,
            'statut_final': record.statut_final,
        })


# ================================================================== #
#  6. VALIDATION EN MASSE PAR DIVISIONNAIRE
# ================================================================== #
class ValiderTousProjetsDivisionnaireView(APIView):
    """
    GET  /recap/budget/valider/divisionnaire/tous/
         → Retourne tous les projets 'approuve_directeur' de l'année prochaine

    POST /recap/budget/valider/divisionnaire/tous/
         → action: 'valide' | 'rejete' | 'annule'
         → Change le statut de TOUS les projets 'approuve_directeur' en masse
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsDivisionnaire]

    def get(self, request):
        next_year = datetime.now().year + 1

        projets = BudgetRecord.objects.filter(
            statut_workflow='approuve_directeur',
            annee_debut_pmt=next_year,
        )

        if not projets.exists():
            return Response({
                'success': True,
                'count': 0,
                'projets': [],
                'total': {},
                'message': f'Aucun projet à valider pour {next_year}',
            })

        total = projets.aggregate(**build_aggregation())
        serializer = BudgetRecordSerializer(projets, many=True, context={'request': request})

        return Response({
            'success': True,
            'count': projets.count(),
            'annee': next_year,
            'projets': serializer.data,
            'total': total,
        })

    def post(self, request):
        next_year = datetime.now().year + 1
        action = request.data.get('action')
        commentaire = request.data.get('commentaire', '')

        if action not in ('valide', 'rejete', 'annule'):
            return Response(
                {'error': "action doit être 'valide', 'rejete' ou 'annule'"},
                status=400
            )

        projets = BudgetRecord.objects.filter(
            statut_workflow='approuve_directeur',
            annee_debut_pmt=next_year,
        )

        if not projets.exists():
            return Response({
                'error': f'Aucun projet avec statut_workflow=approuve_directeur pour {next_year}',
            }, status=404)

        count = projets.count()
        now = timezone.now()
        nom = request.user.nom_complet

        if action == 'valide':
            projets.update(
                statut_final='valide_divisionnaire',
                statut_workflow=None,
                valide_par_divisionnaire=nom,
                date_validation_divisionnaire=now,
                commentaire_divisionnaire=commentaire,
            )
            message = f'{count} projet(s) validés par le divisionnaire ✅'

        elif action == 'rejete':
            projets.update(
                statut_final='rejete_divisionnaire',
                statut_workflow=None,
                rejete_par_divisionnaire=nom,
                date_rejet_divisionnaire=now,
                motif_rejet_divisionnaire=commentaire,
                rejete_par=nom,
                date_rejet=now,
                motif_rejet=commentaire,
            )
            message = f'{count} projet(s) rejetés par le divisionnaire ❌'

        else:  # annule
            projets.update(
                statut_final='annule_divisionnaire',
                statut_workflow=None,
                annule_par_divisionnaire=nom,
                date_annulation_divisionnaire=now,
                motif_annulation_divisionnaire=commentaire,
            )
            message = f'{count} projet(s) annulés par le divisionnaire 🚫'

        return Response({
            'success': True,
            'message': message,
            'action': action,
            'count': count,
            'annee': next_year,
        })



# ─────────────────────────────────────────
# LISTES DES PROJETS PAR STATUT ET RÔLE
# ─────────────────────────────────────────

class MesProjetsBrouillonView(APIView):
    """
    GET /recap/mes-projets/brouillon/
    Responsable structure → voit tous ses projets en brouillon
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsResponsableStructure]

    def get(self, request):
        structure_id = getattr(request.user, 'structure_id', None)

        if not structure_id:
            return Response({'error': 'structure_id introuvable dans le token'}, status=400)

        qs = BudgetRecord.objects.filter(
            statut='brouillon',
            structure_id=structure_id
        ).order_by('-id')

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({'success': True, 'total': qs.count(), 'data': serializer.data})


class ProjetsSoumisParRegionView(APIView):
    """
    GET /recap/projets/soumis/region/
    Directeur région → voit tous les projets soumis dans sa région
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id introuvable dans le token'}, status=400)

        service_url = get_service_param_url()
        token = request.headers.get('Authorization', '')

        try:
            region_resp = requests.get(
                f"{service_url}/params/regions/id/{region_id}",
                headers={'Authorization': token},
                timeout=5
            )
            if region_resp.status_code != 200:
                code_region = region_id
                nom_region  = region_id
            else:
                region_data = region_resp.json().get('data', {})
                code_region = region_data.get('code_region', region_id)
                nom_region  = region_data.get('nom_region', region_id)
        except Exception:
            code_region = region_id
            nom_region  = region_id

        qs = BudgetRecord.objects.filter(
            statut='soumis',
            region_id=region_id
        ).order_by('-id')

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'region': {'id': region_id, 'code': code_region, 'nom': nom_region},
            'total':  qs.count(),
            'data':   serializer.data
        })


class ProjetsValidesDirecteurRegionView(APIView):
    """
    GET /recap/projets/valides/directeur-region/
    Chef → voit tous les projets validés par les directeurs région
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):
        qs = BudgetRecord.objects.filter(statut='valide_directeur_region').order_by('-id')

        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)

        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({'success': True, 'total': qs.count(), 'data': serializer.data})


class ProjetsValidesChefView(APIView):
    """
    GET /recap/projets/valides/chef/
    Directeur → voit tous les projets validés par les chefs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):
        qs = BudgetRecord.objects.filter(statut='valide_chef').order_by('-id')

        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)

        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({'success': True, 'total': qs.count(), 'data': serializer.data})


class ProjetsValidesDirecteurView(APIView):
    """
    GET /recap/projets/valides/directeur/
    Divisionnaire → voit tous les projets validés par les directeurs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDivisionnaire]

    def get(self, request):
        qs = BudgetRecord.objects.filter(statut='valide_directeur').order_by('-id')

        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)

        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({'success': True, 'total': qs.count(), 'data': serializer.data})


class ProjetsSoumisParRegionFiltreView(APIView):
    """
    GET /recap/projets/soumis/region/filtre/?upload_id=1&annee=2026
    Directeur région → version avec filtres optionnels
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id introuvable dans le token'}, status=400)

        service_url = get_service_param_url()
        token = request.headers.get('Authorization', '')

        try:
            region_resp = requests.get(
                f"{service_url}/params/regions/id/{region_id}",
                headers={'Authorization': token},
                timeout=5
            )
            if region_resp.status_code != 200:
                code_region = region_id
                nom_region  = region_id
            else:
                region_data = region_resp.json().get('data', {})
                code_region = region_data.get('code_region', region_id)
                nom_region  = region_data.get('nom_region', region_id)
        except Exception:
            code_region = region_id
            nom_region  = region_id

        qs = BudgetRecord.objects.filter(statut='soumis', region_id=region_id)

        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        annee = request.query_params.get('annee')
        if annee:
            qs = qs.filter(annee=annee)

        activite = request.query_params.get('activite')
        if activite:
            qs = qs.filter(activite=activite)

        famille = request.query_params.get('famille')
        if famille:
            qs = qs.filter(famille=famille)

        qs = qs.order_by('-id')

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'region': {'id': region_id, 'code': code_region, 'nom': nom_region},
            'filtres': {
                'upload_id': upload_id,
                'annee':     annee,
                'activite':  activite,
                'famille':   famille,
            },
            'total': qs.count(),
            'data':  serializer.data
        })

# ─────────────────────────────────────────
# LISTES DES PROJETS PAR STATUT ET RÔLE (CORRIGÉ AVEC NOUVEAUX CHAMPS)
# ─────────────────────────────────────────

class MesProjetsBrouillonView(APIView):
    """
    GET /recap/mes-projets/brouillon/
    Responsable structure → voit tous ses projets en brouillon
    Filtre par structure_id depuis le token
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsResponsableStructure]

    def get(self, request):
        # Récupérer l'ID de la structure depuis le token
        structure_id = getattr(request.user, 'structure_id', None)
        
        if not structure_id:
            return Response({'error': 'structure_id introuvable dans le token'}, status=400)

        # Filtrer par structure_id (champ direct dans BudgetRecord)
        qs = BudgetRecord.objects.filter(
            statut='brouillon',
            structure_id=structure_id
        ).order_by('-id')
        
        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'total': qs.count(),
            'data': serializer.data
        })


class ProjetsSoumisParRegionView(APIView):
    """
    GET /recap/projets/soumis/region/
    Directeur région → voit tous les projets soumis dans sa région
    Filtre par region_id depuis le token
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        # Récupérer region_id depuis le token (c'est l'ObjectId MongoDB)
        region_id = getattr(request.user, 'region_id', None)
        
        if not region_id:
            return Response({'error': 'region_id introuvable dans le token'}, status=400)

        # Résoudre le code_region depuis region_id (ObjectId) via le service param
        service_url = get_service_param_url()
        token = request.headers.get('Authorization', '')
        
        try:
            region_resp = requests.get(
                f"{service_url}/params/regions/id/{region_id}",
                headers={'Authorization': token},
                timeout=5
            )
            
            if region_resp.status_code != 200:
                # Fallback: utiliser region_id directement comme code_region
                code_region = region_id
                nom_region = region_id
            else:
                region_data = region_resp.json().get('data', {})
                code_region = region_data.get('code_region', region_id)
                nom_region = region_data.get('nom_region', region_id)
            
        except Exception as e:
            # Fallback en cas d'erreur
            code_region = region_id
            nom_region = region_id

        # Filtrer les projets soumis dans cette région
        # Utiliser region_id (ObjectId) OU region (code_region)
        qs = BudgetRecord.objects.filter(statut='soumis').order_by('-id')
        
        # Filtrer par region_id si disponible
        qs = qs.filter(region_id=region_id)
        
        # Alternative: filtrer aussi par code_region si nécessaire
        # qs = qs.filter(models.Q(region_id=region_id) | models.Q(region=code_region))

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'region': {
                'id': region_id,
                'code': code_region,
                'nom': nom_region
            },
            'total': qs.count(),
            'data': serializer.data
        })


class ProjetsValidesDirecteurRegionView(APIView):
    """
    GET /recap/projets/valides/directeur-region/
    Chef → voit tous les projets validés par les directeurs région
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut='valide_directeur_region'
        ).order_by('-id')
        
        # Filtre optionnel par région
        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)
        
        # Filtre optionnel par upload
        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'total': qs.count(),
            'data': serializer.data
        })


class ProjetsValidesChefView(APIView):
    """
    GET /recap/projets/valides/chef/
    Directeur → voit tous les projets validés par les chefs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut='valide_chef'
        ).order_by('-id')
        
        # Filtre optionnel par région
        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)
        
        # Filtre optionnel par upload
        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'total': qs.count(),
            'data': serializer.data
        })


class ProjetsValidesDirecteurView(APIView):
    """
    GET /recap/projets/valides/directeur/
    Divisionnaire → voit tous les projets validés par les directeurs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDivisionnaire]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut='valide_directeur'
        ).order_by('-id')
        
        # Filtre optionnel par région
        region_id = request.query_params.get('region_id')
        if region_id:
            qs = qs.filter(region_id=region_id)
        
        # Filtre optionnel par upload
        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'total': qs.count(),
            'data': serializer.data
        })


# Version alternative avec filtres supplémentaires
class ProjetsSoumisParRegionFiltreView(APIView):
    """
    GET /recap/projets/soumis/region/filtre/?upload_id=1&annee=2026
    Directeur région → version avec filtres optionnels
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)
        
        if not region_id:
            return Response({'error': 'region_id introuvable dans le token'}, status=400)

        # Résoudre le code_region depuis region_id
        service_url = get_service_param_url()
        token = request.headers.get('Authorization', '')
        
        try:
            region_resp = requests.get(
                f"{service_url}/params/regions/id/{region_id}",
                headers={'Authorization': token},
                timeout=5
            )
            
            if region_resp.status_code != 200:
                code_region = region_id
                nom_region = region_id
            else:
                region_data = region_resp.json().get('data', {})
                code_region = region_data.get('code_region', region_id)
                nom_region = region_data.get('nom_region', region_id)
        except Exception as e:
            code_region = region_id
            nom_region = region_id

        # Base queryset filtrée par region_id
        qs = BudgetRecord.objects.filter(
            statut='soumis',
            region_id=region_id
        )

        # Filtres optionnels
        upload_id = request.query_params.get('upload_id')
        if upload_id:
            qs = qs.filter(upload_id=upload_id)

        annee = request.query_params.get('annee')
        if annee:
            qs = qs.filter(annee=annee)

        activite = request.query_params.get('activite')
        if activite:
            qs = qs.filter(activite=activite)

        famille = request.query_params.get('famille')
        if famille:
            qs = qs.filter(famille=famille)

        qs = qs.order_by('-id')

        serializer = BudgetRecordSerializer(qs, many=True)
        return Response({
            'success': True,
            'region': {
                'id': region_id,
                'code': code_region,
                'nom': nom_region
            },
            'filtres': {
                'upload_id': upload_id,
                'annee': annee,
                'activite': activite,
                'famille': famille,
            },
            'total': qs.count(),
            'data': serializer.data
        })
    

# ================================================================== #
#  Listes les projets
# ================================================================== #
# ================================================================== #
#  RESPONSABLE STRUCTURE
#  Filtre auto : structure_id du token
# ================================================================== #
class ListeProjetsResponsableView(APIView):
    """
    GET /recap/budget/projets/responsable/
    ?statut=brouillon|soumis|...
    ?type_projet=nouveau|en_cours
    ?code_division=PROJ001
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsResponsableStructure]

    def get(self, request):
        structure_id = getattr(request.user, 'structure_id', None)

        if not structure_id:
            return Response(
                {'error': "Votre token ne contient pas de structure_id."},
                status=403
            )

        qs = BudgetRecord.objects.filter(
            structure_id=structure_id,
            # is_active=True,
        )

        statut        = request.query_params.get('statut')
        type_projet   = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if statut:
            qs = qs.filter(statut=statut)
        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        qs = qs.order_by('-id')

        from django.db.models import Count
        compteurs = {
            item['statut']: item['total']
            for item in qs.values('statut').annotate(total=Count('id'))
        }

        return Response({
            'count':                qs.count(),
            'compteurs_par_statut': compteurs,
            # 'projets':              BudgetRecordSerializer(qs, many=True).data,
            'projets': BudgetRecordSerializer(qs, many=True, context={'request': request}).data,
        })


# ================================================================== #
#  DIRECTEUR RÉGION
#  Filtre auto : region_id du token
# ================================================================== #
# class ListeProjetsDirecteurRegionView(APIView):
#     """
#     GET /recap/budget/projets/directeur-region/
#     ?statut=soumis|reserve_agent|reserve_chef|reserve_directeur|tous
#     ?type_projet=nouveau|en_cours
#     ?code_division=PROJ001
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteurRegion]

#     STATUTS_PAR_DEFAUT = [
#         'soumis',
#         'reserve_agent',
#         'reserve_chef',
#         'reserve_directeur',
#     ]

#     def get(self, request):
#         region_id = getattr(request.user, 'region_id', None)

#         if not region_id:
#             return Response(
#                 {'error': "Votre token ne contient pas de region_id."},
#                 status=403
#             )

#         qs = BudgetRecord.objects.filter(
#             region_id=region_id,
#             is_active=True,
#         )

#         statut        = request.query_params.get('statut')
#         type_projet   = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')

#         if statut and statut != 'tous':
#             qs = qs.filter(statut=statut)
#         else:
#             qs = qs.filter(statut__in=self.STATUTS_PAR_DEFAUT)

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         qs = qs.order_by('-id')

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         return Response({
#             'count':                qs.count(),
#             'compteurs_par_statut': compteurs,
#             'projets':              BudgetRecordSerializer(qs, many=True).data,
#         })


# import logging
# from django.db.models import Count, Q

# # Configuration du logger
# logger = logging.getLogger(__name__)

# class ListeProjetsDirecteurRegionView(APIView):
#     """
#     GET /recap/budget/projets/directeur-region/
#     ?statut=soumis|reserve_agent|reserve_chef|reserve_directeur|valide_directeur_region|tous
#     ?type_projet=nouveau|en_cours
#     ?code_division=PROJ001
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteurRegion]

#     # ✅ MODIFIÉ : Ajout de 'valide_directeur_region'
#     STATUTS_PAR_DEFAUT = [
#         'soumis',
#         'reserve_agent',
#         'reserve_chef',
#         'reserve_directeur',
#         'valide_directeur_region',  # ← NOUVEAU statut
#     ]

#     def get(self, request):
#         print("=" * 80)
#         print("[DEBUG] === ListeProjetsDirecteurRegionView.get() ===")
        
#         # 1. Récupérer region_id du token
#         region_id = getattr(request.user, 'region_id', None)
#         print(f"[DEBUG] region_id from token: {region_id}")
#         print(f"[DEBUG] region_id type: {type(region_id)}")
#         print(f"[DEBUG] request.user: {request.user}")
#         print(f"[DEBUG] request.user role: {getattr(request.user, 'role', 'unknown')}")

#         if not region_id:
#             print("[DEBUG] ❌ Aucun region_id trouvé dans le token")
#             return Response(
#                 {'error': "Votre token ne contient pas de region_id."},
#                 status=403
#             )

#         # 2. Vérifier tous les projets actifs dans la base
#         total_all_active = BudgetRecord.objects.filter(is_active=True).count()
#         print(f"[DEBUG] Total projets actifs dans la base: {total_all_active}")
        
#         # 3. Vérifier les projets avec ce region_id
#         projets_avec_region_id = BudgetRecord.objects.filter(region_id=region_id, is_active=True)
#         print(f"[DEBUG] Projets avec region_id='{region_id}': {projets_avec_region_id.count()}")
        
#         # Afficher les 5 premiers projets trouvés
#         for p in projets_avec_region_id[:5]:
#             print(f"[DEBUG]   - ID:{p.id} | code:{p.code_division} | statut:'{p.statut}' | region:'{p.region}' | region_id:'{p.region_id}'")

#         # 4. Vérifier les projets qui ont region = 'REG001' mais region_id NULL
#         projets_region_null = BudgetRecord.objects.filter(
#             Q(region='REG001') & Q(region_id__isnull=True),
#             is_active=True
#         )
#         print(f"[DEBUG] Projets avec region='REG001' mais region_id NULL: {projets_region_null.count()}")
#         for p in projets_region_null[:5]:
#             print(f"[DEBUG]   - ID:{p.id} | code:{p.code_division} | statut:'{p.statut}' | region:'{p.region}' | region_id:'{p.region_id}'")

#         # 5. Construction du queryset principal
#         qs = BudgetRecord.objects.filter(
#             region_id=region_id,
#             is_active=True,
#         )
#         print(f"[DEBUG] Queryset initial (region_id={region_id}): {qs.count()} projets")

#         # 6. Récupérer les paramètres de requête
#         statut        = request.query_params.get('statut')
#         type_projet   = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')
        
#         print(f"[DEBUG] Paramètres reçus:")
#         print(f"[DEBUG]   - statut: {statut}")
#         print(f"[DEBUG]   - type_projet: {type_projet}")
#         print(f"[DEBUG]   - code_division: {code_division}")

#         # 7. Application du filtre statut
#         if statut:
#             if statut == 'tous':
#                 statuts_inclus = [
#                     'soumis', 'reserve_agent', 'reserve_chef', 
#                     'reserve_directeur', 'valide_directeur_region',
#                     'valide_chef', 'valide_directeur', 'valide_divisionnaire', 
#                     'rejete', 'brouillon', 'valide_agent'
#                 ]
#                 print(f"[DEBUG] Filtre 'tous' - statuts inclus: {statuts_inclus}")
#                 qs = qs.filter(statut__in=statuts_inclus)
#             else:
#                 print(f"[DEBUG] Filtre spécifique - statut='{statut}'")
#                 qs = qs.filter(statut=statut)
#         else:
#             print(f"[DEBUG] Pas de filtre statut - utilisation des STATUTS_PAR_DEFAUT: {self.STATUTS_PAR_DEFAUT}")
#             qs = qs.filter(statut__in=self.STATUTS_PAR_DEFAUT)

#         print(f"[DEBUG] Après filtre statut: {qs.count()} projets")

#         # 8. Application des autres filtres
#         if type_projet:
#             print(f"[DEBUG] Filtre type_projet='{type_projet}'")
#             qs = qs.filter(type_projet=type_projet)
        
#         if code_division:
#             print(f"[DEBUG] Filtre code_division__icontains='{code_division}'")
#             qs = qs.filter(code_division__icontains=code_division)

#         qs = qs.order_by('-id')
#         print(f"[DEBUG] Après tous les filtres: {qs.count()} projets")

#         # 9. Afficher la liste des statuts trouvés
#         statuts_trouves = qs.values_list('statut', flat=True).distinct()
#         print(f"[DEBUG] Statuts trouvés dans le résultat: {list(statuts_trouves)}")

#         # 10. Compter par statut
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }
#         print(f"[DEBUG] Compteurs par statut: {compteurs}")

#         # 11. Afficher les détails du projet spécifique (ID 58 si présent)
#         projet_58 = qs.filter(id=58).first()
#         if projet_58:
#             print(f"[DEBUG] ✅ Projet ID 58 trouvé dans le résultat!")
#             print(f"[DEBUG]   - code_division: {projet_58.code_division}")
#             print(f"[DEBUG]   - statut: {projet_58.statut}")
#             print(f"[DEBUG]   - region_id: {projet_58.region_id}")
#             print(f"[DEBUG]   - is_active: {projet_58.is_active}")
#         else:
#             print(f"[DEBUG] ❌ Projet ID 58 NON trouvé dans le résultat")
#             # Vérifier pourquoi le projet 58 n'est pas trouvé
#             p58 = BudgetRecord.objects.filter(id=58).first()
#             if p58:
#                 print(f"[DEBUG] Projet 58 existe en base:")
#                 print(f"[DEBUG]   - statut: {p58.statut}")
#                 print(f"[DEBUG]   - region_id: {p58.region_id} (type: {type(p58.region_id)})")
#                 print(f"[DEBUG]   - is_active: {p58.is_active}")
#                 print(f"[DEBUG]   - region_id du token: {region_id} (type: {type(region_id)})")
#                 if p58.region_id != region_id:
#                     print(f"[DEBUG] ❌ Mismatch: region_id projet != region_id token")
#                 if p58.statut not in self.STATUTS_PAR_DEFAUT and statut != p58.statut:
#                     print(f"[DEBUG] ❌ Statut '{p58.statut}' non inclus dans le filtre")
#             else:
#                 print(f"[DEBUG] Projet 58 n'existe pas en base")

#         print("=" * 80)
#         print(f"[DEBUG] Réponse finale - count: {qs.count()}")
#         print("=" * 80)

#         return Response({
#             'count':                qs.count(),
#             'compteurs_par_statut': compteurs,
#             'projets':              BudgetRecordSerializer(qs, many=True).data,
#         })
# class ListeProjetsDirecteurRegionView(APIView):
#     """
#     GET /recap/budget/projets/directeur-region/
#     ?statut=soumis|reserve_agent|reserve_chef|reserve_directeur|valide_directeur_region|tous
#     ?type_projet=nouveau|en_cours
#     ?code_division=PROJ001
#     ?inclure_inactifs=true   # Force l'inclusion des projets inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes     = [IsDirecteurRegion]

#     STATUTS_PAR_DEFAUT = [
#         'soumis',
#         'reserve_agent',
#         'reserve_chef',
#         'reserve_directeur',
#         'valide_directeur_region',
#     ]
    
#     # Statuts pour lesquels on inclut automatiquement les inactifs
#     STATUTS_AVEC_HISTORIQUE = [
#         'valide_directeur_region',
#         'valide_chef', 
#         'valide_directeur',
#         'valide_divisionnaire',
#         'rejete'
#     ]

#     def get(self, request):
#         print("=" * 80)
#         print("[DEBUG] === ListeProjetsDirecteurRegionView.get() ===")
        
#         region_id = getattr(request.user, 'region_id', None)
#         print(f"[DEBUG] region_id from token: {region_id}")

#         if not region_id:
#             return Response(
#                 {'error': "Votre token ne contient pas de region_id."},
#                 status=403
#             )

#         statut        = request.query_params.get('statut')
#         type_projet   = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')
        
#         # ✅ Règle intelligente pour inclure/inclure les inactifs
#         inclure_inactifs_param = request.query_params.get('inclure_inactifs', 'auto').lower()
        
#         if inclure_inactifs_param == 'true':
#             inclure_inactifs = True
#             print(f"[DEBUG] Force inclusion des inactifs via paramètre")
#         elif inclure_inactifs_param == 'false':
#             inclure_inactifs = False
#             print(f"[DEBUG] Force exclusion des inactifs via paramètre")
#         else:
#             # Mode auto : inclure les inactifs seulement pour certains statuts
#             if statut and statut in self.STATUTS_AVEC_HISTORIQUE:
#                 inclure_inactifs = True
#                 print(f"[DEBUG] Mode auto: inclusion des inactifs pour statut='{statut}'")
#             elif not statut and any(s in self.STATUTS_AVEC_HISTORIQUE for s in self.STATUTS_PAR_DEFAUT):
#                 inclure_inactifs = True
#                 print(f"[DEBUG] Mode auto: inclusion des inactifs pour statuts par défaut")
#             else:
#                 inclure_inactifs = False
#                 print(f"[DEBUG] Mode auto: exclusion des inactifs")

#         # Construction du queryset
#         if inclure_inactifs:
#             qs = BudgetRecord.objects.filter(region_id=region_id)
#             print(f"[DEBUG] ✅ Incluant les projets inactifs (is_active=True ou False)")
#         else:
#             qs = BudgetRecord.objects.filter(region_id=region_id, is_active=True)
#             print(f"[DEBUG] ❌ Excluant les projets inactifs (is_active=True uniquement)")

#         # Application des filtres
#         if statut:
#             if statut == 'tous':
#                 qs = qs.filter(
#                     statut__in=[
#                         'soumis', 'reserve_agent', 'reserve_chef', 
#                         'reserve_directeur', 'valide_directeur_region',
#                         'valide_chef', 'valide_directeur', 'valide_divisionnaire', 
#                         'rejete', 'brouillon', 'valide_agent'
#                     ]
#                 )
#             else:
#                 qs = qs.filter(statut=statut)
#         else:
#             qs = qs.filter(statut__in=self.STATUTS_PAR_DEFAUT)

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         qs = qs.order_by('-id')

#         # Logging pour le projet 58
#         projet_58 = qs.filter(id=58).first()
#         if projet_58:
#             print(f"[DEBUG] ✅ Projet 58 trouvé! is_active={projet_58.is_active}, statut={projet_58.statut}")
#         else:
#             p58 = BudgetRecord.objects.filter(id=58).first()
#             if p58:
#                 print(f"[DEBUG] ❌ Projet 58 EXCLU: is_active={p58.is_active}, statut={p58.statut}")
#                 print(f"[DEBUG]    - inclure_inactifs={inclure_inactifs}")
#                 print(f"[DEBUG]    - statut demandé={statut}")
#                 if p58.statut == statut and not inclure_inactifs and not p58.is_active:
#                     print(f"[DEBUG]    ➜ Solution: Ajoutez &inclure_inactifs=true à l'URL")

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         print(f"[DEBUG] Total final: {qs.count()} projets")
#         print("=" * 80)

#         return Response({
#             'count':                qs.count(),
#             'compteurs_par_statut': compteurs,
#             'filtres_appliques': {
#                 'inclure_inactifs': inclure_inactifs,
#                 'statut': statut,
#                 'type_projet': type_projet,
#                 'code_division': code_division,
#             },
#             'projets': BudgetRecordSerializer(qs, many=True).data,
#         })







# ================================================================== #
#  DIRECTEUR RÉGION - LISTES SPÉCIFIQUES (gets)
# ================================================================== #

class ListeProjetsSoumisDRView(APIView):
    """
    GET /recap/budget/directeur-region/soumis/
    Projets soumis (actifs + inactifs)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)
        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        # ✅ Inclut TOUS les projets (is_active=True ou False)
        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_workflow='soumis'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })



class ListeProjetsValidesDRView(APIView):
    """
    GET /recap/budget/directeur-region/valides/
    
    Retourne tous les projets que le directeur région a validés,
    indépendamment de leur statut actuel.
    
    Critère : valide_par_directeur_region n'est pas NULL
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)
        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        # ✅ Filtre sur le champ de validation, pas sur le statut
        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            valide_par_directeur_region__isnull=False  # ← La clé !
        ).order_by('-date_validation_directeur_region')  # Tri par date de validation

        # Sérialisation avec contexte pour avoir les noms
        serializer = BudgetRecordSerializer(
            qs, 
            many=True, 
            context={'request': request}
        )

        return Response({
            'count': qs.count(),
            'projets': serializer.data
        })


class ListeProjetsRejetesDRView(APIView):
    """
    GET /recap/budget/directeur-region/rejetes/
    Projets rejetés (actifs + inactifs)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)
        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        # ✅ Inclut TOUS les projets (is_active=True ou False)
        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_final='rejete_directeur_region'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


class ListeProjetsReserveDirecteurDRView(APIView):
    """
    GET /recap/budget/directeur-region/reserve-directeur/
    Projets réservés par le directeur (actifs + inactifs)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)
        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        # ✅ Inclut TOUS les projets (is_active=True ou False)
        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_workflow='reserve_directeur'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


class ListeProjetsTousDRView(APIView):
    """
    GET /recap/budget/directeur-region/tous/
    Tous les projets de la région (actifs + inactifs)
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    ALLOWED_WORKFLOW = [
        'reserve_directeur',
    ]

    ALLOWED_FINAL = [
        'valide_divisionnaire',
        'rejete_divisionnaire',
        'annule_divisionnaire',
        'valide_directeur_region',
        'rejete_directeur_region',
    ]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        from django.db.models import Q

        qs = BudgetRecord.objects.filter(
            region_id=region_id
        ).filter(
            Q(statut_workflow__in=self.ALLOWED_WORKFLOW) |
            Q(statut_final__in=self.ALLOWED_FINAL)
        ).order_by('-id')

        from django.db.models import Count

        compteurs = {
            item['statut_workflow'] or item['statut_final']: item['total']
            for item in qs.values('statut_workflow', 'statut_final')
            .annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'compteurs_par_statut': compteurs,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
    
class ListeProjetsValideDivisionnaireView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_final='valide_divisionnaire'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
class ListeProjetsRejeteDivisionnaireView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_final='rejete_divisionnaire'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
class ListeProjetsAnnuleDivisionnaireView(APIView):
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        qs = BudgetRecord.objects.filter(
            region_id=region_id,
            statut_final='annule_divisionnaire'
        ).order_by('-id')

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
class ListeProjetsrevoirDRView(APIView):
    """
    GET /recap/budget/directeur-region/revoir/
    Tous les projets de la région (actifs + inactifs)
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteurRegion]

    ALLOWED_WORKFLOW = [
        'reserve_directeur',
    ]

    ALLOWED_FINAL = [
        'rejete_divisionnaire',
        'annule_divisionnaire',
    ]

    def get(self, request):
        region_id = getattr(request.user, 'region_id', None)

        if not region_id:
            return Response({'error': 'region_id manquant'}, status=403)

        from django.db.models import Q

        qs = BudgetRecord.objects.filter(
            region_id=region_id
        ).filter(
            Q(statut_workflow__in=self.ALLOWED_WORKFLOW) |
            Q(statut_final__in=self.ALLOWED_FINAL)
        ).order_by('-id')

        from django.db.models import Count

        compteurs = {
            item['statut_workflow'] or item['statut_final']: item['total']
            for item in qs.values('statut_workflow', 'statut_final')
            .annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'compteurs_par_statut': compteurs,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })

# class ListeProjetsHistoriqueDRView(APIView):
#     """
#     GET /recap/budget/directeur-region/historique/
#     TOUS les projets sans filtre de statut (actifs + inactifs)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDirecteurRegion]

#     def get(self, request):
#         region_id = getattr(request.user, 'region_id', None)
#         if not region_id:
#             return Response({'error': 'region_id manquant'}, status=403)

#         # ✅ Inclut TOUS les projets, TOUS les statuts, is_active=True ou False
#         qs = BudgetRecord.objects.filter(
#             region_id=region_id
#         ).order_by('-id')

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         return Response({
#             'count': qs.count(),
#             'compteurs_par_statut': compteurs,
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })





# class ListeProjetsBrouillonDRView(APIView):
#     """
#     GET /recap/budget/directeur-region/brouillon/
#     Projets en brouillon (actifs + inactifs)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDirecteurRegion]

#     def get(self, request):
#         region_id = getattr(request.user, 'region_id', None)
#         if not region_id:
#             return Response({'error': 'region_id manquant'}, status=403)

#         # ✅ Inclut TOUS les projets (is_active=True ou False)
#         qs = BudgetRecord.objects.filter(
#             region_id=region_id,
#             statut='brouillon'
#         ).order_by('-id')

#         return Response({
#             'count': qs.count(),
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })

# ================================================================== #
#  CHEF les gets 
# ================================================================== #



class ListeProjetsChefView(APIView):
    """
    GET /recap/budget/chef/valider-DR/

    Récupère uniquement les projets validés par le directeur région
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_final='valide_directeur_region'
        ).order_by('-id')

        serializer = BudgetRecordSerializer(
            qs,
            many=True,
            context={'request': request}
        )

        return Response({
            'success': True,
            'count': qs.count(),
            'projets': serializer.data
        })

class ListeProjetsChefValidesView(APIView):
    """
    GET /recap/budget/chef/pre_approuve/
    Projets déjà pré-approuvés par le chef
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):

        qs = BudgetRecord.objects.filter(
            preapprouve_par_chef__isnull=False
        ).order_by('-date_preapprouve_chef')  # ✅ FIX ICI

        # filtres optionnels
        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)

        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        serializer = BudgetRecordSerializer(
            qs,
            many=True,
            context={'request': request}
        )

        stats = {
            'total': qs.count(),
            # 'par_statut_final': {
            #     'valide_directeur_region': qs.filter(statut_final='valide_directeur_region').count(),
            #     'rejete_directeur_region': qs.filter(statut_final='rejete_directeur_region').count(),
            #     'valide_divisionnaire': qs.filter(statut_final='valide_divisionnaire').count(),
            #     'rejete_divisionnaire': qs.filter(statut_final='rejete_divisionnaire').count(),
            #     'annule_divisionnaire': qs.filter(statut_final='annule_divisionnaire').count(),
            # }
        }

        return Response({
            'success': True,
            'stats': stats,
            'count': qs.count(),
            'projets': serializer.data
        })

class ListeProjetsChefReserveChefView(APIView):
    """
    GET /recap/budget/chef/reserve-chef/
    Projets réservés par le chef (retournés à l'agent)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_workflow='reserve_chef'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


class ListeProjetsChefTousView(APIView):
    """
    GET /recap/budget/chef/tous/

    Affiche uniquement :
    - pré-approuvés chef
    - réservés chef
    - validés directeur région
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsChef]

    def get(self, request):

        qs = BudgetRecord.objects.filter(
            # 🔵 chef actions
            # preapprouve_par_chef__isnull=False
            statut_workflow='pre_approuve_chef'
        ) | BudgetRecord.objects.filter(
            # reserve_par_chef__isnull=False
            statut_workflow='reserve_chef'
        ) | BudgetRecord.objects.filter(
            statut_final='valide_directeur_region'
        )

        qs = qs.distinct().order_by('-id')

        # filtres optionnels
        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)

        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        # 📊 stats propres
        stats = {
            "total": qs.count(),

            "preapprouves_chef": qs.filter(preapprouve_par_chef__isnull=False).count(),
            "reserves_chef": qs.filter(reserve_par_chef__isnull=False).count(),
            "valide_directeur_region": qs.filter(statut_final='valide_directeur_region').count(),
        }

        serializer = BudgetRecordSerializer(
            qs,
            many=True,
            context={'request': request}
        )

        return Response({
            "success": True,
            "stats": stats,
            "count": qs.count(),
            "projets": serializer.data
        })


# class ListeProjetsChefHistoriqueView(APIView):
#     """
#     GET /recap/budget/chef/historique/
#     TOUS les projets (tous statuts) pour historique
#     Inclut actifs + inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsChef]

#     def get(self, request):
#         qs = BudgetRecord.objects.all().order_by('-id')

#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         return Response({
#             'count': qs.count(),
#             'compteurs_par_statut': compteurs,
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })








# ================================================================== #
#  DIRECTEUR - LISTES SPÉCIFIQUES
# ================================================================== #

class ListeProjetsDirecteurValidesChefView(APIView):
    """
    GET /recap/budget/directeur/valides-chef/
    Projets validés par le chef (à valider par le directeur)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut='valide_chef'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


# class ListeProjetsDirecteurReserveChefView(APIView):
#     """
#     GET /recap/budget/directeur/reserve-chef/
#     Projets réservés par le chef (retournés au directeur)
#     Inclut actifs + inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDirecteur]

#     def get(self, request):
#         qs = BudgetRecord.objects.filter(
#             statut='reserve_chef'
#         ).order_by('-id')

#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         return Response({
#             'count': qs.count(),
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })

class ListeProjetsDirecteurView(APIView):
    """
    GET /recap/budget/directeur/projets/
    
    Récupère TOUS les projets:
        - statut_workflow = 'pre_approuve_chef' (pre_approuve par le chef)
        - statut_workflow = 'reserve_chef' (réservés par le chef)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):
        # Tous les projets (validés + réservés par le chef)
        qs = BudgetRecord.objects.filter(
            statut_workflow__in=['pre_approuve_chef', 'reserve_chef']
        ).order_by('-id')
        
        # Filtres optionnels
        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')
        
        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)
        
        # Statistiques
        stats = {
            'total': qs.count(),
            'pre_approuve_chef': qs.filter(statut_workflow='pre_approuve_chef').count(),
            'reserve_chef': qs.filter(statut_workflow='reserve_chef').count(),
        }
        
        # Sérialisation
        serializer = BudgetRecordSerializer(
            qs, 
            many=True, 
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'stats': stats,
            'count': qs.count(),
            'projets': serializer.data
        })

# class ListeProjetsDirecteurValidesView(APIView):
#     """
#     GET /recap/budget/directeur/valides/
    
#     Projets déjà validés par le directeur (historique complet)
#     Filtre sur valide_par_directeur NOT NULL (indépendant du statut actuel)
#     Inclut actifs + inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDirecteur]

#     def get(self, request):
#         # ✅ Filtre sur le champ de validation, pas sur le statut
#         qs = BudgetRecord.objects.filter(
#             approuve_par_directeur__isnull=False  # ← La clé !
#         ).order_by('-date_approuve_directeur')  # Tri par date de validation
        
#         # Filtres optionnels
#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')
        
#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)
        
#         # Statistiques optionnelles
#         stats = {
#             'total': qs.count(),
#             'par_statut_actuel': {
#                 # 'valide_divisionnaire': qs.filter(statut='valide_divisionnaire').count(),
#                 'approuve_directeur': qs.filter(statut_workflow='approuve_directeur').count(),
#                 'rejete': qs.filter(statut='rejete').count(),
    
#             }
#         }
        
#         # Sérialisation
#         serializer = BudgetRecordSerializer(
#             qs, 
#             many=True, 
#             context={'request': request}
#         )
        
#         return Response({
#             'success': True,
#             'stats': stats,
#             'count': qs.count(),
#             'projets': serializer.data
#         })
class ListeProjetsDirecteurValidesView(APIView):
    """
    GET /recap/budget/directeur/valides/
    Historique des projets approuvés par le directeur
    """

    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):

        qs = BudgetRecord.objects.filter(
            approuve_par_directeur__isnull=False
        ).order_by('-date_approuve_directeur')

        # filtres optionnels
        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)

        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        # ✅ STATS CORRIGÉES
        stats = {
            "total": qs.count(),

            "approuves_directeur": qs.filter(approuve_par_directeur__isnull=False).count(),

            # "reserves_directeur": qs.filter(reserve_par_directeur__isnull=False).count(),

            # "valide_directeur_region": qs.filter(statut_final='valide_directeur_region').count(),

            # "rejete_directeur_region": qs.filter(statut_final='rejete_directeur_region').count(),
        }

        serializer = BudgetRecordSerializer(
            qs,
            many=True,
            context={'request': request}
        )

        return Response({
            "success": True,
            "stats": stats,
            "count": qs.count(),
            "projets": serializer.data
        })
class ListeProjetsDirecteurReserveDirecteurView(APIView):
    """
    GET /recap/budget/directeur/reserve-directeur/
    Projets réservés par le directeur (retournés au chef)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_workflow='reserve_directeur'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        return Response({
            'count': qs.count(),
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


class ListeProjetsDirecteurTousView(APIView):
    """
    GET /recap/budget/directeur/tous/
    Tous les projets du directeur (valides chef + réservés chef + valides directeur + réservés directeur)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDirecteur]

    STATUTS_DIRECTEUR = [
        # 'pre_approuve_chef',          # projets à valider
        # 'reserve_chef',         # projets réservés par chef
        'approuve_directeur',     # projets déjà validés par directeur
        'reserve_directeur',    # projets réservés par directeur
    ]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_workflow__in=self.STATUTS_DIRECTEUR
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        from django.db.models import Count
        compteurs = {
            item['statut_workflow']: item['total']
            for item in qs.values('statut_workflow').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'compteurs_par_statut': compteurs,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


# class ListeProjetsDirecteurHistoriqueView(APIView):
#     """
#     GET /recap/budget/directeur/historique/
#     TOUS les projets (tous statuts) pour historique
#     Inclut actifs + inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDirecteur]

#     def get(self, request):
#         qs = BudgetRecord.objects.all().order_by('-id')

#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         return Response({
#             'count': qs.count(),
#             'compteurs_par_statut': compteurs,
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })
    

# ================================================================== #
#  DIVISIONNAIRE - LISTES SPÉCIFIQUES
# ================================================================== #

class ListeProjetsDivisionnaireValidesDirecteurView(APIView):
    """
    GET /recap/budget/divisionnaire/valides-directeur/
    Projets validés par le directeur (à valider par le divisionnaire)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDivisionnaire]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_workflow='approuve_directeur'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        from django.db.models import Count
        par_region = {
            item['region_id']: item['total']
            for item in qs.values('region_id').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'par_region': par_region,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })


# class ListeProjetsDivisionnaireView(APIView):
#     """
#     GET /recap/budget/divisionnaire/projets/
    
#     Récupère TOUS les projets:
#         - statut_workflow='approuve_directeur'(validés par le directeur)
        
    
#     Inclut comptage par région
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDivisionnaire]

#     def get(self, request):
#         from django.db.models import Count
        
#         # Tous les projets (validés + réservés par le directeur)
#         qs = BudgetRecord.objects.filter(
#             statut__in=['valide_directeur', 'reserve_directeur']
#         ).order_by('-id')
        
#         # Filtres optionnels
#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')
        
#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)
        
#         # Comptage par région
#         par_region = {
#             item['region_id']: item['total']
#             for item in qs.values('region_id').annotate(total=Count('id'))
#         }
        
#         # Statistiques
#         stats = {
#             'total': qs.count(),
#             'valides_directeur': qs.filter(statut='valide_directeur').count(),
#             'reserve_directeur': qs.filter(statut='reserve_directeur').count(),
#             'par_region': par_region,
#         }
        
#         # Sérialisation
#         serializer = BudgetRecordSerializer(
#             qs, 
#             many=True, 
#             context={'request': request}
#         )
        
#         return Response({
#             'success': True,
#             'stats': stats,
#             'count': qs.count(),
#             'par_region': par_region,
#             'projets': serializer.data
#         })

class ListeProjetsDivisionnaireValidesView(APIView):
    """
    GET /recap/budget/divisionnaire/valides/
    Projets déjà validés par le divisionnaire (validation finale)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_final='valide_divisionnaire'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        from django.db.models import Count
        par_region = {
            item['region_id']: item['total']
            for item in qs.values('region_id').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'par_region': par_region,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
from django.db.models import Q, Count

class ListeProjetsDivisionnaireTerminesView(APIView):
    """
    GET /recap/budget/divisionnaire/termines/
    
    Projets terminés par le divisionnaire :
        - valide_divisionnaire
        - rejete_divisionnaire
        - annule_divisionnaire
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    STATUTS_TERMINE = [
        'valide_divisionnaire',
        'rejete_divisionnaire',
        'annule_divisionnaire',
    ]

    def get(self, request):

        qs = BudgetRecord.objects.filter(
            statut_final__in=self.STATUTS_TERMINE
        ).order_by('-date_validation_divisionnaire')

        # filtres optionnels
        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')
        statut_filter = request.query_params.get('statut_final')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)

        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        if statut_filter in self.STATUTS_TERMINE:
            qs = qs.filter(statut_final=statut_filter)

        # stats
        stats = {
            'total': qs.count(),
            'valide_divisionnaire': qs.filter(statut_final='valide_divisionnaire').count(),
            'rejete_divisionnaire': qs.filter(statut_final='rejete_divisionnaire').count(),
            'annule_divisionnaire': qs.filter(statut_final='annule_divisionnaire').count(),
        }

        serializer = BudgetRecordSerializer(
            qs,
            many=True,
            context={'request': request}
        )

        return Response({
            'success': True,
            'stats': stats,
            'count': qs.count(),
            'projets': serializer.data
        })
class ListeProjetsDivisionnaireRejetesView(APIView):
    """
    GET /recap/budget/divisionnaire/rejetes/
    Projets rejetés (le divisionnaire peut rejeter)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_final='rejete_divisionnaire'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        from django.db.models import Count
        par_region = {
            item['region_id']: item['total']
            for item in qs.values('region_id').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'par_region': par_region,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
class ListeProjetsDivisionnaireAnnulesView(APIView):
    """
    GET /recap/budget/divisionnaire/annules/
    Projets annules (le divisionnaire peut annuler)
    Inclut actifs + inactifs
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = BudgetRecord.objects.filter(
            statut_final='annule_divisionnaire'
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)
        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        from django.db.models import Count
        par_region = {
            item['region_id']: item['total']
            for item in qs.values('region_id').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'par_region': par_region,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
from django.db.models import Q, Count


class ListeProjetsDivisionnaireTousView(APIView):
    """
    GET /recap/budget/divisionnaire/tous/
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsDivisionnaire]

    STATUTS_FINAL = [
        'valide_divisionnaire',
        'rejete_divisionnaire',
        'annule_divisionnaire',
    ]

    def get(self, request):

        qs = BudgetRecord.objects.filter(
            Q(statut_workflow='approuve_directeur') |
            Q(statut_final__in=self.STATUTS_FINAL)
        ).order_by('-id')

        type_projet = request.query_params.get('type_projet')
        code_division = request.query_params.get('code_division')

        if type_projet:
            qs = qs.filter(type_projet=type_projet)

        if code_division:
            qs = qs.filter(code_division__icontains=code_division)

        # ✅ IMPORTANT: utiliser statut_final (pas statut)
        compteurs = {
            item['statut_final']: item['total']
            for item in qs.values('statut_final').annotate(total=Count('id'))
            if item['statut_final'] is not None
        }

        par_region = {
            item['region_id']: item['total']
            for item in qs.values('region_id').annotate(total=Count('id'))
        }

        return Response({
            'count': qs.count(),
            'compteurs_par_statut_final': compteurs,
            'par_region': par_region,
            'projets': BudgetRecordSerializer(qs, many=True).data
        })
# class ListeProjetsDivisionnaireHistoriqueView(APIView):
#     """
#     GET /recap/budget/divisionnaire/historique/
#     TOUS les projets (tous statuts) pour historique
#     Inclut actifs + inactifs
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsDivisionnaire]

#     def get(self, request):
#         qs = BudgetRecord.objects.all().order_by('-id')

#         type_projet = request.query_params.get('type_projet')
#         code_division = request.query_params.get('code_division')

#         if type_projet:
#             qs = qs.filter(type_projet=type_projet)
#         if code_division:
#             qs = qs.filter(code_division__icontains=code_division)

#         from django.db.models import Count
#         compteurs = {
#             item['statut']: item['total']
#             for item in qs.values('statut').annotate(total=Count('id'))
#         }

#         par_region = {
#             item['region_id']: item['total']
#             for item in qs.values('region_id').annotate(total=Count('id'))
#         }

#         return Response({
#             'count': qs.count(),
#             'compteurs_par_statut': compteurs,
#             'par_region': par_region,
#             'projets': BudgetRecordSerializer(qs, many=True).data
#         })






# ================================================================== #
#  HISTORIQUE PAR CODE_DIVISION
#  Accessible par tous les rôles
# ================================================================== #
class HistoriqueProjetView(APIView):
    """
    GET /recap/budget/historique/<code_division>/
    GET /recap/budget/historique/<code_division>/actif/
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes     = [IsUser]

    def get(self, request, code_division, mode=None):
        qs = BudgetRecord.objects.filter(
            code_division=code_division
        ).order_by('-version')

        if not qs.exists():
            return Response(
                {'error': f'Projet {code_division} introuvable.'},
                status=404
            )

        if mode == 'actif':
            actif = qs.filter(is_active=True).first() or qs.first()
            return Response({
                'code_division':  code_division,
                'version_active': BudgetRecordSerializer(actif).data,
            })

        actif    = qs.filter(is_active=True).first() or qs.first()
        derniere = qs.first()
        premiere = qs.last()

        return Response({
            'code_division':    code_division,
            'total_versions':   qs.count(),
            'version_active':   actif.version,
            'derniere_version': derniere.version,
            'premiere_version': premiere.version,
            'historique':       BudgetRecordSerializer(qs, many=True).data,
        })


class ProjetAvecVersionPrecedenteView(APIView):
    """
    GET /api/budget/projet/{code_division}/with-previous/
    
    Récupère le projet actif et sa version précédente (version N-1)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, code_division):
        # Récupérer la version active
        projet_actif = BudgetRecord.objects.filter(
            code_division=code_division, 
            is_active=True
        ).first()
        
        if not projet_actif:
            return Response(
                {'error': f'Projet {code_division} introuvable.'},
                status=404
            )
        
        # 🔥 CORRECTION : Récupérer la version précédente par le numéro de version
        # Au lieu d'utiliser parent_id, on cherche la version avec le numéro (version_actuelle - 1)
        version_precedente = BudgetRecord.objects.filter(
            code_division=code_division,
            version=projet_actif.version - 1  # Version N-1
        ).first()
        
        # Si pas trouvée, essayer de la trouver via parent_id (fallback)
        if not version_precedente and projet_actif.parent_id:
            version_precedente = BudgetRecord.objects.filter(
                id=projet_actif.parent_id
            ).first()
        
        # Sérialiser
        serializer_actif = BudgetRecordSerializer(
            projet_actif, 
            context={'request': request}
        )
        
        serializer_precedent = None
        if version_precedente:
            serializer_precedent = BudgetRecordSerializer(
                version_precedente, 
                context={'request': request}
            )
        
        # Retourner la réponse
        response_data = {
            'success': True,
            'code_division': code_division,
            'projet_actuel': {
                'version': projet_actif.version,
                'data': serializer_actif.data
            }
        }
        
        if serializer_precedent:
            response_data['version_precedente'] = {
                'version': version_precedente.version,
                'data': serializer_precedent.data
            }
        else:
            response_data['version_precedente'] = None
            response_data['message'] = 'Ce projet est la version initiale (aucune version précédente)'
        
        return Response(response_data, status=200)

# # ================================================================== #
# # EXPORT EXCEL - PROJETS VALIDÉS PAR DIVISIONNAIRE
# # ================================================================== #

# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from django.http import HttpResponse
# from datetime import datetime
# import requests
# from django.db.models import Sum, Q

# class ExportProjetsValidesDivisionnaireView(APIView):
#     """
#     GET /recap/budget/export/valides-divisionnaire/
    
#     Exporte les projets validés par le divisionnaire avec annee_debut_pmt = année_actuelle + 1
#     Format identique au template recap.xls
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     # Mapping des activités (ajustez selon vos besoins)
#     ACTIVITE_MAPPING = {
#         'A': 'Activité A - Production',
#         'B': 'Activité B - Transport',
#         # Ajoutez vos mappings ici
#     }
    
#     def get(self, request):
#         # Calculer l'année cible = année actuelle + 1
#         annee_cible = datetime.now().year + 1
        
#         print(f"[DEBUG] Export projets validés divisionnaire - Année cible: {annee_cible}")
        
#         # Récupérer les projets validés par le divisionnaire avec l'année cible
#         qs = BudgetRecord.objects.filter(
#             statut='valide_divisionnaire',
#             annee_debut_pmt=annee_cible,
#             is_active=True
#         ).order_by('region', 'code_division')
        
#         print(f"[DEBUG] Nombre de projets trouvés: {qs.count()}")
        
#         # Récupérer les mappings
#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')
        
#         region_mapping = self._get_region_mapping(service_url, token)
#         famille_mapping = self._get_famille_mapping(service_url, token)
        
#         # Créer le classeur Excel
#         wb = openpyxl.Workbook()
        
#         # Supprimer la feuille par défaut
#         wb.remove(wb.active)
        
#         # Créer la feuille principale
#         sheet_name = f"Projets_Valides_Divisionnaire_{annee_cible}"
#         ws = wb.create_sheet(sheet_name[:31])  # Excel limite à 31 caractères
        
#         # Appliquer le formatage
#         self._write_header(ws, annee_cible)
#         self._write_data_rows(ws, qs, region_mapping, famille_mapping)
#         self._apply_column_widths(ws)
        
#         # Créer une feuille récapitulative
#         ws_recap = wb.create_sheet("Recapitulatif")
#         self._write_recap_sheet(ws_recap, qs, region_mapping, annee_cible)
        
#         # Préparer la réponse
#         filename = f"projets_valides_divisionnaire_{annee_cible}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
#         wb.save(response)
#         return response
    
#     def _write_header(self, ws, annee_cible):
#         """Écrit les en-têtes du fichier Excel"""
        
#         # Styles
#         header_font = Font(bold=True, size=11, color="FFFFFF")
#         header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#         subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
#         title_font = Font(bold=True, size=14)
        
#         # Ligne 1: Titre principal
#         ws.merge_cells('A1:F1')
#         ws['A1'] = "SONATRACH - ACTIVITE AMONT - DIVISION PRODUCTION"
#         ws['A1'].font = Font(bold=True, size=14)
#         ws['A1'].alignment = Alignment(horizontal='center')
        
#         # Ligne 2: Sous-titre
#         ws.merge_cells('A2:F2')
#         ws['A2'] = f"PROJETS VALIDÉS PAR LE DIVISIONNAIRE - PMT {annee_cible}/{annee_cible+4}"
#         ws['A2'].font = Font(bold=True, size=12)
#         ws['A2'].alignment = Alignment(horizontal='center')
        
#         # Ligne 3: Date d'export
#         ws.merge_cells('A3:F3')
#         ws['A3'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
#         ws['A3'].alignment = Alignment(horizontal='center')
        
#         # Ligne 4: Unité
#         ws.merge_cells('A4:F4')
#         ws['A4'] = "Unité : Millier DA"
#         ws['A4'].font = Font(bold=True)
#         ws['A4'].alignment = Alignment(horizontal='center')
        
#         # Ligne 5: En-têtes principaux
#         main_headers = [
#             ('A', 'B'), ('C',), ('D',), ('E',), ('F', 'G'), ('H', 'I'), 
#             ('J', 'K'), ('L', 'M'), ('N', 'O'), ('P', 'Q'), ('R', 'S'),
#             ('T', 'U'), ('V', 'W'), ('X', 'Y'), ('Z', 'AA'), ('AB', 'AC'),
#             ('AD', 'AE'), ('AF', 'AG'), ('AH', 'AI'), ('AJ', 'AK'), ('AL', 'AM')
#         ]
        
#         main_texts = [
#             "Activité / Région / Périmètre / Famille",
#             "N°:Cpte Analy.",
#             "Libellés",
#             "Coût Global Initial\nPMT",
#             "Réalisations Cumulées",
#             "Prévisions de clôture N",
#             "Prévisions N+1",
#             "Reste à Réaliser",
#             "Prévisions N+2",
#             "Prévisions N+3",
#             "Prévisions N+4",
#             "Prévisions N+5",
#             "Janvier",
#             "Février",
#             "Mars",
#             "Avril",
#             "Mai",
#             "Juin",
#             "Juillet",
#             "Août",
#             "Septembre",
#             "Octobre",
#             "Novembre",
#             "Décembre"
#         ]
        
#         row_num = 5
#         col_idx = 0
#         for start_col, end_col in main_headers:
#             if start_col != end_col:
#                 ws.merge_cells(f'{start_col}{row_num}:{end_col}{row_num}')
#             ws[f'{start_col}{row_num}'] = main_texts[col_idx] if col_idx < len(main_texts) else ""
#             ws[f'{start_col}{row_num}'].font = Font(bold=True)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#             ws[f'{start_col}{row_num}'].fill = header_fill
#             ws[f'{start_col}{row_num}'].font = header_font
#             col_idx += 1
        
#         # Ligne 6: Sous-en-têtes Total / Dont DEX
#         row_num = 6
#         sous_headers = [
#             ('A', ''), ('B', ''), ('C', ''), ('D', ''), ('E', ''),
#             ('F', 'Total'), ('G', 'Dont DEX'),
#             ('H', 'Total'), ('I', 'Dont DEX'),
#             ('J', 'Total'), ('K', 'Dont DEX'),
#             ('L', 'Total'), ('M', 'Dont DEX'),
#             ('N', 'Total'), ('O', 'Dont DEX'),
#             ('P', 'Total'), ('Q', 'Dont DEX'),
#             ('R', 'Total'), ('S', 'Dont DEX'),
#             ('T', 'Total'), ('U', 'Dont DEX'),
#             ('V', 'Total'), ('W', 'Dont DEX'),
#             ('X', 'Total'), ('Y', 'Dont DEX'),
#             ('Z', 'Total'), ('AA', 'Dont DEX'),
#             ('AB', 'Total'), ('AC', 'Dont DEX'),
#             ('AD', 'Total'), ('AE', 'Dont DEX'),
#             ('AF', 'Total'), ('AG', 'Dont DEX'),
#             ('AH', 'Total'), ('AI', 'Dont DEX'),
#             ('AJ', 'Total'), ('AK', 'Dont DEX'),
#             ('AL', 'Total'), ('AM', 'Dont DEX')
#         ]
        
#         for col_letter, header_text in sous_headers:
#             if header_text:
#                 ws[f'{col_letter}{row_num}'] = header_text
#                 ws[f'{col_letter}{row_num}'].font = Font(bold=True, size=10)
#                 ws[f'{col_letter}{row_num}'].alignment = Alignment(horizontal='center')
#                 ws[f'{col_letter}{row_num}'].fill = subheader_fill
        
#         # Ajuster la hauteur des lignes d'en-tête
#         for r in range(1, 7):
#             ws.row_dimensions[r].height = 30
    
#     def _write_data_rows(self, ws, qs, region_mapping, famille_mapping):
#         """Écrit les lignes de données"""
#         row_num = 7
        
#         # Définir les styles pour les bordures
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )
        
#         for record in qs:
#             # Données de base
#             activite = self.ACTIVITE_MAPPING.get(record.activite, record.activite or '')
#             region = region_mapping.get(record.region, record.region or '')
#             perimetre = record.perm or ''
#             famille = famille_mapping.get(record.famille, record.famille or '')
#             code_analytique = record.code_division or ''
#             libelle = record.libelle or ''
            
#             # Données financières
#             cout_total = self._format_number(record.cout_initial_total)
#             cout_dex = self._format_number(record.cout_initial_dont_dex)
            
#             realisation_total = self._format_number(record.realisation_cumul_n_mins1_total)
#             realisation_dex = self._format_number(record.realisation_cumul_n_mins1_dont_dex)
            
#             prev_cloture_total = self._format_number(record.prev_cloture_n_total)
#             prev_cloture_dex = self._format_number(record.prev_cloture_n_dont_dex)
            
#             prev_n1_total = self._format_number(record.prev_n_plus1_total)
#             prev_n1_dex = self._format_number(record.prev_n_plus1_dont_dex)
            
#             rar_total = self._format_number(record.reste_a_realiser_total)
#             rar_dex = self._format_number(record.reste_a_realiser_dont_dex)
            
#             # Données annuelles N+2 à N+5
#             prev_n2_total = self._format_number(record.prev_n_plus2_total)
#             prev_n2_dex = self._format_number(record.prev_n_plus2_dont_dex)
#             prev_n3_total = self._format_number(record.prev_n_plus3_total)
#             prev_n3_dex = self._format_number(record.prev_n_plus3_dont_dex)
#             prev_n4_total = self._format_number(record.prev_n_plus4_total)
#             prev_n4_dex = self._format_number(record.prev_n_plus4_dont_dex)
#             prev_n5_total = self._format_number(record.prev_n_plus5_total)
#             prev_n5_dex = self._format_number(record.prev_n_plus5_dont_dex)
            
#             # Données mensuelles
#             mois_data = []
#             for mois in ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']:
#                 total = getattr(record, f'{mois}_total', None)
#                 dex = getattr(record, f'{mois}_dont_dex', None)
#                 mois_data.append((self._format_number(total), self._format_number(dex)))
            
#             # Construire la ligne
#             row_data = [
#                 f"{activite} / {region} / {perimetre} / {famille}",
#                 code_analytique,
#                 libelle,
#                 cout_total, cout_dex,
#                 realisation_total, realisation_dex,
#                 prev_cloture_total, prev_cloture_dex,
#                 prev_n1_total, prev_n1_dex,
#                 rar_total, rar_dex,
#                 prev_n2_total, prev_n2_dex,
#                 prev_n3_total, prev_n3_dex,
#                 prev_n4_total, prev_n4_dex,
#                 prev_n5_total, prev_n5_dex,
#             ]
            
#             # Ajouter les 12 mois
#             for mois_total, mois_dex in mois_data:
#                 row_data.append(mois_total)
#                 row_data.append(mois_dex)
            
#             # Écrire la ligne
#             for col_idx, value in enumerate(row_data, start=1):
#                 cell = ws.cell(row=row_num, column=col_idx, value=value)
#                 cell.border = thin_border
                
#                 # Alignement différent pour les colonnes texte et chiffres
#                 if col_idx <= 3:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')
#                 else:
#                     cell.alignment = Alignment(horizontal='right', vertical='center')
#                     if value and isinstance(value, (int, float)):
#                         cell.number_format = '#,##0.00'
            
#             # Appliquer une couleur de fond alternée
#             if row_num % 2 == 0:
#                 for col_idx in range(1, len(row_data) + 1):
#                     ws.cell(row=row_num, column=col_idx).fill = PatternFill(
#                         start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
#                     )
            
#             row_num += 1
    
#     def _write_recap_sheet(self, ws, qs, region_mapping, annee_cible):
#         """Écrit la feuille récapitulative"""
        
#         # Titre
#         ws['A1'] = f"RÉCAPITULATIF DES PROJETS VALIDÉS PAR LE DIVISIONNAIRE - PMT {annee_cible}"
#         ws['A1'].font = Font(bold=True, size=14)
#         ws.merge_cells('A1:G1')
        
#         ws['A3'] = "Date d'export:"
#         ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
#         ws['A4'] = "Nombre total de projets:"
#         ws['B4'] = qs.count()
        
#         # Entêtes du tableau
#         headers = ['Code Division', 'Libellé', 'Région', 'Activité', 
#                    'Famille', 'Coût Total (kDA)', 'Dont DEX (kDA)']
        
#         for col_idx, header in enumerate(headers, start=1):
#             cell = ws.cell(row=6, column=col_idx, value=header)
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center')
        
#         # Données
#         row_num = 7
#         total_general = 0
#         total_dex_general = 0
        
#         # Grouper par région
#         regions_data = {}
        
#         for record in qs:
#             region_nom = region_mapping.get(record.region, record.region or 'Non spécifié')
#             cout_total = float(record.cout_initial_total or 0)
#             cout_dex = float(record.cout_initial_dont_dex or 0)
            
#             total_general += cout_total
#             total_dex_general += cout_dex
            
#             if region_nom not in regions_data:
#                 regions_data[region_nom] = {'total': 0, 'dex': 0, 'count': 0}
#             regions_data[region_nom]['total'] += cout_total
#             regions_data[region_nom]['dex'] += cout_dex
#             regions_data[region_nom]['count'] += 1
            
#             ws.cell(row=row_num, column=1, value=record.code_division or '')
#             ws.cell(row=row_num, column=2, value=record.libelle or '')
#             ws.cell(row=row_num, column=3, value=region_nom)
#             ws.cell(row=row_num, column=4, value=record.activite or '')
#             ws.cell(row=row_num, column=5, value=record.famille or '')
            
#             cell_total = ws.cell(row=row_num, column=6, value=cout_total)
#             cell_total.number_format = '#,##0.00'
#             cell_dex = ws.cell(row=row_num, column=7, value=cout_dex)
#             cell_dex.number_format = '#,##0.00'
            
#             row_num += 1
        
#         # Ligne de séparation
#         ws.merge_cells(f'A{row_num}:G{row_num}')
#         ws.cell(row=row_num, column=1, value="").border = Border(top=Side(style='thin'))
#         row_num += 1
        
#         # Sous-totaux par région
#         ws.cell(row=row_num, column=1, value="RÉCAPITULATIF PAR RÉGION")
#         ws.merge_cells(f'A{row_num}:G{row_num}')
#         ws.cell(row=row_num, column=1).font = Font(bold=True)
#         row_num += 1
        
#         for region_nom, data in sorted(regions_data.items()):
#             ws.cell(row=row_num, column=1, value=region_nom)
#             ws.cell(row=row_num, column=2, value=f"{data['count']} projet(s)")
#             cell_total = ws.cell(row=row_num, column=3, value=data['total'])
#             cell_total.number_format = '#,##0.00'
#             cell_dex = ws.cell(row=row_num, column=4, value=data['dex'])
#             cell_dex.number_format = '#,##0.00'
#             row_num += 1
        
#         row_num += 1
        
#         # Total général
#         ws.cell(row=row_num, column=1, value="TOTAL GÉNÉRAL")
#         ws.cell(row=row_num, column=1).font = Font(bold=True)
#         cell_total = ws.cell(row=row_num, column=2, value=total_general)
#         cell_total.font = Font(bold=True)
#         cell_total.number_format = '#,##0.00'
#         cell_dex = ws.cell(row=row_num, column=3, value=total_dex_general)
#         cell_dex.font = Font(bold=True)
#         cell_dex.number_format = '#,##0.00'
        
#         # Ajuster les largeurs
#         for col in range(1, 8):
#             ws.column_dimensions[get_column_letter(col)].width = 22
    
#     def _apply_column_widths(self, ws):
#         """Applique les largeurs de colonnes"""
#         column_widths = {
#             'A': 45,  # Activité/Région/Périmètre/Famille
#             'B': 18,  # Code analytique
#             'C': 40,  # Libellés
#             'D': 15,  # Coût Global
#             'E': 15,  # Dont DEX
#             'F': 15, 'G': 15,  # Réalisations
#             'H': 15, 'I': 15,  # Prévisions clôture
#             'J': 15, 'K': 15,  # Prévisions N+1
#             'L': 15, 'M': 15,  # Reste à réaliser
#             'N': 15, 'O': 15,  # Prévisions N+2
#             'P': 15, 'Q': 15,  # Prévisions N+3
#             'R': 15, 'S': 15,  # Prévisions N+4
#             'T': 15, 'U': 15,  # Prévisions N+5
#         }
        
#         # Largeurs pour les mois (colonnes V à AM)
#         for i in range(22, 40):  # V à AM
#             col_letter = get_column_letter(i)
#             column_widths[col_letter] = 12
        
#         for col_letter, width in column_widths.items():
#             try:
#                 ws.column_dimensions[col_letter].width = width
#             except:
#                 pass
    
#     def _get_region_mapping(self, service_url, token):
#         """Récupère le mapping des régions"""
#         mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 data = response.json().get('data', [])
#                 for item in data:
#                     code = item.get('code_region')
#                     nom = item.get('nom_region')
#                     if code:
#                         mapping[str(code)] = nom
#         except Exception as e:
#             print(f"Erreur récupération régions: {e}")
#         return mapping
    
#     def _get_famille_mapping(self, service_url, token):
#         """Récupère le mapping des familles"""
#         mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 data = response.json().get('data', [])
#                 for item in data:
#                     code = item.get('code_famille')
#                     nom = item.get('nom_famille')
#                     if code:
#                         mapping[str(code)] = nom
#         except Exception as e:
#             print(f"Erreur récupération familles: {e}")
#         return mapping
    
#     def _format_number(self, value):
#         """Formate un nombre"""
#         if value is None:
#             return 0
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return 0


# # ================================================================== #
# # EXPORT AVEC FILTRES PERSONNALISÉS
# # ================================================================== #

# class ExportProjetsFiltresView(APIView):
#     """
#     GET /recap/budget/export/filtres/
    
#     Export avec filtres personnalisables :
#     - statut: valide_divisionnaire, valide_directeur, etc.
#     - annee_debut_pmt: année spécifique
#     - region: code région
#     - activite: code activité
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         # Récupérer les filtres
#         statut = request.query_params.get('statut')
#         annee_debut_pmt = request.query_params.get('annee_debut_pmt')
#         region = request.query_params.get('region')
#         activite = request.query_params.get('activite')
#         famille = request.query_params.get('famille')
        
#         # Construire le queryset
#         qs = BudgetRecord.objects.filter(is_active=True)
        
#         if statut:
#             qs = qs.filter(statut=statut)
        
#         if annee_debut_pmt:
#             try:
#                 annee_int = int(annee_debut_pmt)
#                 qs = qs.filter(annee_debut_pmt=annee_int)
#             except ValueError:
#                 pass
        
#         if region:
#             qs = qs.filter(region=region)
        
#         if activite:
#             qs = qs.filter(activite=activite)
        
#         if famille:
#             qs = qs.filter(famille=famille)
        
#         qs = qs.order_by('region', 'code_division')
        
#         # Récupérer les mappings
#         service_url = get_service_param_url()
#         token = request.headers.get('Authorization', '')
        
#         region_mapping = self._get_region_mapping(service_url, token)
#         famille_mapping = self._get_famille_mapping(service_url, token)
        
#         # Mapping des activités
#         activite_mapping = {
#             'A': 'Production',
#             'B': 'Transport',
#             'C': 'Stockage',
#             'D': 'Distribution',
#             # Ajoutez vos mappings
#         }
        
#         # Créer le classeur
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Projets_Filtres"
        
#         # En-têtes
#         headers = [
#             'Code Division', 'Libellé', 'Région', 'Activité', 'Famille',
#             'Périmètre', 'Année Début PMT', 'Année Fin PMT',
#             'Coût Total (kDA)', 'Dont DEX (kDA)', 'Statut', 'Version'
#         ]
        
#         for col_idx, header in enumerate(headers, start=1):
#             cell = ws.cell(row=1, column=col_idx, value=header)
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center')
        
#         # Données
#         row_num = 2
#         total_general = 0
        
#         for record in qs:
#             ws.cell(row=row_num, column=1, value=record.code_division or '')
#             ws.cell(row=row_num, column=2, value=record.libelle or '')
#             ws.cell(row=row_num, column=3, value=region_mapping.get(record.region, record.region or ''))
#             ws.cell(row=row_num, column=4, value=activite_mapping.get(record.activite, record.activite or ''))
#             ws.cell(row=row_num, column=5, value=famille_mapping.get(record.famille, record.famille or ''))
#             ws.cell(row=row_num, column=6, value=record.perm or '')
#             ws.cell(row=row_num, column=7, value=record.annee_debut_pmt or '')
#             ws.cell(row=row_num, column=8, value=record.annee_fin_pmt or '')
            
#             cout_total = float(record.cout_initial_total or 0)
#             total_general += cout_total
            
#             cell_total = ws.cell(row=row_num, column=9, value=cout_total)
#             cell_total.number_format = '#,##0.00'
            
#             cell_dex = ws.cell(row=row_num, column=10, value=float(record.cout_initial_dont_dex or 0))
#             cell_dex.number_format = '#,##0.00'
            
#             ws.cell(row=row_num, column=11, value=record.statut or '')
#             ws.cell(row=row_num, column=12, value=record.version or 1)
            
#             row_num += 1
        
#         # Total
#         ws.cell(row=row_num, column=8, value="TOTAL:")
#         ws.cell(row=row_num, column=8).font = Font(bold=True)
#         cell_total = ws.cell(row=row_num, column=9, value=total_general)
#         cell_total.font = Font(bold=True)
#         cell_total.number_format = '#,##0.00'
        
#         # Ajuster les largeurs
#         for col in range(1, 13):
#             ws.column_dimensions[get_column_letter(col)].width = 20
        
#         # Réponse
#         filename = f"projets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
#         wb.save(response)
#         return response
    
#     def _get_region_mapping(self, service_url, token):
#         """Récupère le mapping des régions"""
#         mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/regions", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 data = response.json().get('data', [])
#                 for item in data:
#                     code = item.get('code_region')
#                     nom = item.get('nom_region')
#                     if code:
#                         mapping[str(code)] = nom
#         except Exception as e:
#             print(f"Erreur récupération régions: {e}")
#         return mapping
    
#     def _get_famille_mapping(self, service_url, token):
#         """Récupère le mapping des familles"""
#         mapping = {}
#         try:
#             headers = {'Authorization': f'Bearer {token}'} if token else {}
#             response = requests.get(f"{service_url}/params/familles", headers=headers, timeout=5)
#             if response.status_code == 200:
#                 data = response.json().get('data', [])
#                 for item in data:
#                     code = item.get('code_famille')
#                     nom = item.get('nom_famille')
#                     if code:
#                         mapping[str(code)] = nom
#         except Exception as e:
#             print(f"Erreur récupération familles: {e}")
#         return mapping
# ================================================================== #
# EXPORT EXCEL - FORMAT IDENTIQUE AU TEMPLATE recap.xls (CORRIGÉ)
# ================================================================== #

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import requests

from .mappings import ACTIVITE_MAPPING


# class ExportProjetsValidesDivisionnaireView(APIView):
#     """
#     GET /recap/budget/export/valides-divisionnaire/
    
#     Exporte les projets validés par le divisionnaire avec annee_debut_pmt = année_actuelle + 1
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         annee_cible = datetime.now().year -2
        
#         print(f"[DEBUG] Export - Année cible: {annee_cible}")
        
#         qs = BudgetRecord.objects.filter(
#             statut='valide_divisionnaire',
#             annee_debut_pmt=annee_cible,
#             # is_active=True
#         ).order_by('region', 'code_division')
        
#         print(f"[DEBUG] Projets trouvés: {qs.count()}")
        
#         wb = openpyxl.Workbook()
#         wb.remove(wb.active)
        
#         ws = wb.create_sheet("invest 2015-2019 Vers Stag")
        
#         self._write_template_header(ws, annee_cible)
#         self._write_template_data_rows(ws, qs)
#         self._apply_template_column_widths(ws)
        
#         ws_recap = wb.create_sheet("Récapitulatif")
#         self._write_recap_sheet(ws_recap, qs, annee_cible)
        
#         filename = f"projets_valides_divisionnaire_{annee_cible}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
#         wb.save(response)
#         return response
    
#     def _write_template_header(self, ws, annee_cible):
#         """Écrit les en-têtes exactement comme dans le template recap.xls"""
        
#         header_font = Font(bold=True, size=11)
#         title_font = Font(bold=True, size=14)
#         subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
#         # Lignes 1-6: Titres
#         ws.merge_cells('A1:E1')
#         ws['A1'] = "SONATRACH - ACTIVITE AMONT - DIVISION PRODUCTION"
#         ws['A1'].font = title_font
#         ws['A1'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A2:E2')
#         ws['A2'] = "PLAN ANNUEL 2015 ET PMT 2015/2019"
#         ws['A2'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A3:E3')
#         ws['A3'] = "ECHEANCIER DETAILLE PAR PROJETS"
#         ws['A3'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A4:E4')
#         ws['A4'] = "EXPLICATION DES VARIATIONS DES COÛTS"
#         ws['A4'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A5:E5')
#         ws['A5'] = "DEGLOBALISATION DE LA PREVISION ANNUELLE 2015"
#         ws['A5'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A6:E6')
#         ws['A6'] = "Unité : Millier DA"
#         ws['A6'].font = Font(bold=True)
#         ws['A6'].alignment = Alignment(horizontal='center')
        
#         # Ligne 7: En-têtes principaux
#         header_fusions = [
#             ('A', 'D'), ('E', 'E'), ('F', 'F'), ('G', 'H'), ('I', 'J'),
#             ('K', 'L'), ('M', 'N'), ('O', 'P'), ('Q', 'R'), ('S', 'T'),
#             ('U', 'V'), ('W', 'X'), ('Y', 'Z'), ('AA', 'AB'), ('AC', 'AD'),
#             ('AE', 'AF'), ('AG', 'AH'), ('AI', 'AJ'), ('AK', 'AL'), ('AM', 'AN'),
#             ('AO', 'AP'), ('AQ', 'AR'), ('AS', 'AT'), ('AU', 'AV')
#         ]
        
#         header_texts = [
#             "Activité / Région / PERIMETRE / Famille", "N°:Cpte Analy.", "Libellés",
#             "Coût Global Initial PMT 2015/2019", "Réalisations Cumulées à fin 2013 au coût réel",
#             "Prévisions de clôture 2014", "Prévisions 2015", "Reste à Réaliser 2016/2019",
#             "Prévisions 2016", "Prévisions 2017", "Prévisions 2018", "Prévisions 2019",
#             "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
#             "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
#         ]
        
#         row_num = 7
#         for idx, (start_col, end_col) in enumerate(header_fusions):
#             if start_col != end_col:
#                 ws.merge_cells(f'{start_col}{row_num}:{end_col}{row_num}')
#             cell = ws[f'{start_col}{row_num}']
#             cell.value = header_texts[idx] if idx < len(header_texts) else ""
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
#         # Ligne 8: Sous-en-têtes Total / dont Dev.
#         row_num = 8
        
#         for col in ['A', 'B', 'C', 'D', 'E', 'F']:
#             ws.merge_cells(f'{col}{row_num}:{col}{row_num+1}')
        
#         dex_columns = [
#             ('G', 'H'), ('I', 'J'), ('K', 'L'), ('M', 'N'), ('O', 'P'),
#             ('Q', 'R'), ('S', 'T'), ('U', 'V'), ('W', 'X')
#         ]
        
#         for start_col, end_col in dex_columns:
#             ws[f'{start_col}{row_num}'] = "Total"
#             ws[f'{end_col}{row_num}'] = "dont Dev."
#             ws[f'{start_col}{row_num}'].font = Font(bold=True, size=10)
#             ws[f'{end_col}{row_num}'].font = Font(bold=True, size=10)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{start_col}{row_num}'].fill = subheader_fill
#             ws[f'{end_col}{row_num}'].fill = subheader_fill
        
#         # Mois - VERSION CORRIGÉE pour gérer 'AA', 'AB', etc.
#         mois_pairs = [
#             ('Y', 'Z'), ('AA', 'AB'), ('AC', 'AD'), ('AE', 'AF'),
#             ('AG', 'AH'), ('AI', 'AJ'), ('AK', 'AL'), ('AM', 'AN'),
#             ('AO', 'AP'), ('AQ', 'AR'), ('AS', 'AT'), ('AU', 'AV')
#         ]
        
#         for start_col, end_col in mois_pairs:
#             ws[f'{start_col}{row_num}'] = "Total"
#             ws[f'{end_col}{row_num}'] = "dont Dev."
#             ws[f'{start_col}{row_num}'].font = Font(bold=True, size=9)
#             ws[f'{end_col}{row_num}'].font = Font(bold=True, size=9)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{start_col}{row_num}'].fill = subheader_fill
#             ws[f'{end_col}{row_num}'].fill = subheader_fill
        
#         # Ligne 9: sous-en-têtes "Prévisions" pour les mois
#         row_num = 9
#         for start_col, end_col in mois_pairs:
#             ws[f'{start_col}{row_num}'] = "Prévisions"
#             ws[f'{end_col}{row_num}'] = ""
#             ws[f'{start_col}{row_num}'].font = Font(size=8)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
        
#         for r in range(1, 10):
#             ws.row_dimensions[r].height = 30
#         ws.row_dimensions[7].height = 40
    
#     def _write_template_data_rows(self, ws, qs):
#         """Écrit les données"""
        
#         row_num = 10
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )
        
#         for record in qs:
#             # Récupérer les valeurs
#             activite = record.activite or ''
#             region = record.region or ''
#             perimetre = record.perm or ''
#             famille = record.famille or ''
#             code_analytique = record.code_division or ''
#             libelle = record.libelle or ''
            
#             # Données financières
#             cout_total = self._format_number(record.cout_initial_total)
#             cout_dex = self._format_number(record.cout_initial_dont_dex)
#             realisation_total = self._format_number(record.realisation_cumul_n_mins1_total)
#             realisation_dex = self._format_number(record.realisation_cumul_n_mins1_dont_dex)
            
#             # ⚠️ IMPORTANT: Réal. 1er Semestre et Prév. 2è Semestre
#             real_s1_total = self._format_number(record.real_s1_n_total)
#             real_s1_dex = self._format_number(record.real_s1_n_dont_dex)
#             prev_s2_total = self._format_number(record.prev_s2_n_total)
#             prev_s2_dex = self._format_number(record.prev_s2_n_dont_dex)
            
#             # Prévisions de clôture 2014 = Real S1 + Prev S2
#             prev_cloture_total = self._format_number(record.prev_cloture_n_total)
#             prev_cloture_dex = self._format_number(record.prev_cloture_n_dont_dex)
            
#             prev_n1_total = self._format_number(record.prev_n_plus1_total)
#             prev_n1_dex = self._format_number(record.prev_n_plus1_dont_dex)
#             rar_total = self._format_number(record.reste_a_realiser_total)
#             rar_dex = self._format_number(record.reste_a_realiser_dont_dex)
#             prev_n2_total = self._format_number(record.prev_n_plus2_total)
#             prev_n2_dex = self._format_number(record.prev_n_plus2_dont_dex)
#             prev_n3_total = self._format_number(record.prev_n_plus3_total)
#             prev_n3_dex = self._format_number(record.prev_n_plus3_dont_dex)
#             prev_n4_total = self._format_number(record.prev_n_plus4_total)
#             prev_n4_dex = self._format_number(record.prev_n_plus4_dont_dex)
#             prev_n5_total = self._format_number(record.prev_n_plus5_total)
#             prev_n5_dex = self._format_number(record.prev_n_plus5_dont_dex)
            
#             # Données mensuelles
#             mois_data = []
#             for mois in ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']:
#                 total = getattr(record, f'{mois}_total', None)
#                 dex = getattr(record, f'{mois}_dont_dex', None)
#                 mois_data.append((self._format_number(total), self._format_number(dex)))
            
#             # Construction de la ligne selon le template Excel
#             # Colonnes A à AV comme dans votre fichier recap.xls
#             row_data = [
#                 # Colonnes A-F: identifiants
#                 activite,           # A
#                 region,             # B  
#                 perimetre,          # C
#                 famille,            # D
#                 code_analytique,    # E
#                 libelle,            # F
                
#                 # Colonnes G-H: Coût Global Initial
#                 cout_total, cout_dex,
                
#                 # Colonnes I-J: Réalisations Cumulées
#                 realisation_total, realisation_dex,
                
#                 # ⚠️ Colonnes K-L: Réal. 1er Semestre (c'est ici !)
#                 real_s1_total, real_s1_dex,
                
#                 # ⚠️ Colonnes M-N: Prév. 2è Semestre (c'est ici !)
#                 prev_s2_total, prev_s2_dex,
                
#                 # Colonnes O-P: Prévisions de clôture 2014
#                 prev_cloture_total, prev_cloture_dex,
                
#                 # Colonnes Q-R: Prévisions 2015
#                 prev_n1_total, prev_n1_dex,
                
#                 # Colonnes S-T: Reste à Réaliser
#                 rar_total, rar_dex,
                
#                 # Colonnes U-V: Prévisions 2016
#                 prev_n2_total, prev_n2_dex,
                
#                 # Colonnes W-X: Prévisions 2017
#                 prev_n3_total, prev_n3_dex,
                
#                 # Colonnes Y-Z: Prévisions 2018
#                 prev_n4_total, prev_n4_dex,
                
#                 # Colonnes AA-AB: Prévisions 2019
#                 prev_n5_total, prev_n5_dex,
#             ]
            
#             # Ajouter les 12 mois (colonnes AC à AV)
#             for mois_total, mois_dex in mois_data:
#                 row_data.append(mois_total)
#                 row_data.append(mois_dex)
            
#             # Écrire la ligne
#             for col_idx, value in enumerate(row_data, start=1):
#                 col_letter = get_column_letter(col_idx)
#                 cell = ws[f'{col_letter}{row_num}']
#                 cell.value = value
#                 cell.border = thin_border
                
#                 # Alignement
#                 if col_idx <= 6:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')
#                 else:
#                     cell.alignment = Alignment(horizontal='right', vertical='center')
#                     if value and isinstance(value, (int, float)):
#                         cell.number_format = '#,##0.00'
            
#             row_num += 1
    
#     # def _write_template_data_rows(self, ws, qs):
#     #     """Écrit les données"""
        
#     #     row_num = 10
#     #     thin_border = Border(
#     #         left=Side(style='thin'),
#     #         right=Side(style='thin'),
#     #         top=Side(style='thin'),
#     #         bottom=Side(style='thin')
#     #     )
        
#     #     for record in qs:
#     #         # Récupérer les valeurs
#     #         activite = record.activite or ''
#     #         region = record.region or ''
#     #         perimetre = record.perm or ''
#     #         famille = record.famille or ''
#     #         code_analytique = record.code_division or ''
#     #         libelle = record.libelle or ''
            
#     #         # Données financières
#     #         cout_total = self._format_number(record.cout_initial_total)
#     #         cout_dex = self._format_number(record.cout_initial_dont_dex)
#     #         realisation_total = self._format_number(record.realisation_cumul_n_mins1_total)
#     #         realisation_dex = self._format_number(record.realisation_cumul_n_mins1_dont_dex)
#     #         real_s1_total = self._format_number(record.real_s1_n_total)
#     #         real_s1_dex = self._format_number(record.real_s1_n_dont_dex)
#     #         prev_s2_total = self._format_number(record.prev_s2_n_total)
#     #         prev_s2_dex = self._format_number(record.prev_s2_n_dont_dex)
#     #         prev_cloture_total = self._format_number(record.prev_cloture_n_total)
#     #         prev_cloture_dex = self._format_number(record.prev_cloture_n_dont_dex)
#     #         prev_n1_total = self._format_number(record.prev_n_plus1_total)
#     #         prev_n1_dex = self._format_number(record.prev_n_plus1_dont_dex)
#     #         rar_total = self._format_number(record.reste_a_realiser_total)
#     #         rar_dex = self._format_number(record.reste_a_realiser_dont_dex)
#     #         prev_n2_total = self._format_number(record.prev_n_plus2_total)
#     #         prev_n2_dex = self._format_number(record.prev_n_plus2_dont_dex)
#     #         prev_n3_total = self._format_number(record.prev_n_plus3_total)
#     #         prev_n3_dex = self._format_number(record.prev_n_plus3_dont_dex)
#     #         prev_n4_total = self._format_number(record.prev_n_plus4_total)
#     #         prev_n4_dex = self._format_number(record.prev_n_plus4_dont_dex)
#     #         prev_n5_total = self._format_number(record.prev_n_plus5_total)
#     #         prev_n5_dex = self._format_number(record.prev_n_plus5_dont_dex)
            
#     #         # Données mensuelles
#     #         mois_data = []
#     #         for mois in ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#     #                     'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']:
#     #             total = getattr(record, f'{mois}_total', None)
#     #             dex = getattr(record, f'{mois}_dont_dex', None)
#     #             mois_data.append((self._format_number(total), self._format_number(dex)))
            
#     #         # Construction de la ligne
#     #         row_data = [
#     #             activite, region, perimetre, famille, code_analytique, libelle,
#     #             cout_total, cout_dex,
#     #             realisation_total, realisation_dex,
#     #             real_s1_total, real_s1_dex,
#     #             prev_s2_total, prev_s2_dex,
#     #             prev_cloture_total, prev_cloture_dex,
#     #             prev_n1_total, prev_n1_dex,
#     #             rar_total, rar_dex,
#     #             prev_n2_total, prev_n2_dex,
#     #             prev_n3_total, prev_n3_dex,
#     #             prev_n4_total, prev_n4_dex,
#     #             prev_n5_total, prev_n5_dex,
#     #         ]
            
#     #         # Ajouter les mois
#     #         for mois_total, mois_dex in mois_data:
#     #             row_data.append(mois_total)
#     #             row_data.append(mois_dex)
            
#     #         # Écrire la ligne
#     #         for col_idx, value in enumerate(row_data, start=1):
#     #             col_letter = get_column_letter(col_idx)
#     #             cell = ws[f'{col_letter}{row_num}']
#     #             cell.value = value
#     #             cell.border = thin_border
                
#     #             if col_idx <= 6:
#     #                 cell.alignment = Alignment(horizontal='left', vertical='center')
#     #             else:
#     #                 cell.alignment = Alignment(horizontal='right', vertical='center')
#     #                 if value and isinstance(value, (int, float)):
#     #                     cell.number_format = '#,##0.00'
            
#     #         row_num += 1
    
#     def _write_recap_sheet(self, ws, qs, annee_cible):
#         """Écrit la feuille récapitulative"""
        
#         ws['A1'] = f"RÉCAPITULATIF - PROJETS VALIDÉS PAR DIVISIONNAIRE"
#         ws['A1'].font = Font(bold=True, size=14)
#         ws.merge_cells('A1:H1')
        
#         ws['A3'] = f"PMT: {annee_cible} - {annee_cible+4}"
#         ws['A4'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
#         ws['A5'] = f"Nombre de projets: {qs.count()}"
        
#         headers = ['Activité', 'Région', 'Périmètre', 'Famille', 
#                    'Code Division', 'Libellé', 'Coût Total', 'Dont DEX']
        
#         for col_idx, header in enumerate(headers, start=1):
#             cell = ws.cell(row=7, column=col_idx, value=header)
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center')
        
#         row_num = 8
#         total_general = 0
        
#         for record in qs:
#             cout_total = float(record.cout_initial_total or 0)
#             total_general += cout_total
            
#             ws.cell(row=row_num, column=1, value=record.activite or '')
#             ws.cell(row=row_num, column=2, value=record.region or '')
#             ws.cell(row=row_num, column=3, value=record.perm or '')
#             ws.cell(row=row_num, column=4, value=record.famille or '')
#             ws.cell(row=row_num, column=5, value=record.code_division or '')
#             ws.cell(row=row_num, column=6, value=record.libelle or '')
            
#             cell_total = ws.cell(row=row_num, column=7, value=cout_total)
#             cell_total.number_format = '#,##0.00'
#             cell_dex = ws.cell(row=row_num, column=8, value=float(record.cout_initial_dont_dex or 0))
#             cell_dex.number_format = '#,##0.00'
            
#             row_num += 1
        
#         ws.cell(row=row_num, column=6, value="TOTAL:")
#         ws.cell(row=row_num, column=6).font = Font(bold=True)
#         cell_total = ws.cell(row=row_num, column=7, value=total_general)
#         cell_total.font = Font(bold=True)
#         cell_total.number_format = '#,##0.00'
        
#         for col in range(1, 9):
#             ws.column_dimensions[get_column_letter(col)].width = 18
    
#     def _apply_template_column_widths(self, ws):
#         """Applique les largeurs de colonnes"""
#         column_widths = {
#             'A': 12, 'B': 15, 'C': 20, 'D': 10, 'E': 18, 'F': 40,
#             'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15,
#             'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 15, 'R': 15,
#             'S': 15, 'T': 15, 'U': 15, 'V': 15, 'W': 15, 'X': 15,
#         }
        
#         for i in range(25, 48):
#             col_letter = get_column_letter(i)
#             column_widths[col_letter] = 12
        
#         for col_letter, width in column_widths.items():
#             try:
#                 ws.column_dimensions[col_letter].width = width
#             except:
#                 pass
    
#     def _format_number(self, value):
#         if value is None:
#             return 0
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return 0
# ================================================================== #
# EXPORT EXCEL - FORMAT IDENTIQUE AU TEMPLATE recap.xls
# ================================================================== #

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import requests

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from .models import BudgetRecord
from .remote_auth import RemoteJWTAuthentication
from .mappings import ACTIVITE_MAPPING


# class ExportProjetsValidesDivisionnaireView(APIView):
#     """
#     GET /recap/budget/export/valides-divisionnaire/
    
#     Exporte les projets validés par le divisionnaire 
#     avec filtre optionnel par année (?annee=2024)
#     """
#     authentication_classes = [RemoteJWTAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         # Récupérer l'année depuis les paramètres GET
#         annee_param = request.query_params.get('annee')
        
#         if annee_param:
#             annee_cible = int(annee_param)
#         else:
#             annee_cible = datetime.now().year - 2
        
#         print(f"[DEBUG] Export - Année cible: {annee_cible}")
        
#         # Récupérer les projets (sans filtre is_active)
#         qs = BudgetRecord.objects.filter(
#             statut='valide_divisionnaire',
#             annee_debut_pmt=annee_cible
#         ).order_by('region', 'code_division')
        
#         print(f"[DEBUG] Projets trouvés: {qs.count()}")
        
#         # Afficher les détails des projets trouvés
#         for record in qs:
#             print(f"[DEBUG] Projet: {record.code_division} - is_active: {record.is_active}")
        
#         # Créer le classeur Excel
#         wb = openpyxl.Workbook()
#         wb.remove(wb.active)
        
#         ws = wb.create_sheet("invest 2015-2019 Vers Stag")
        
#         self._write_template_header(ws, annee_cible)
#         self._write_template_data_rows(ws, qs)
#         self._apply_template_column_widths(ws)
        
#         ws_recap = wb.create_sheet("Récapitulatif")
#         self._write_recap_sheet(ws_recap, qs, annee_cible)
        
#         filename = f"projets_valides_divisionnaire_{annee_cible}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
#         wb.save(response)
#         return response
    
#     def _write_template_header(self, ws, annee_cible):
#         """Écrit les en-têtes exactement comme dans le template recap.xls"""
        
#         header_font = Font(bold=True, size=11)
#         title_font = Font(bold=True, size=14)
#         subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
#         # ═══════════════════════════════════════════════════════════════════
#         # LIGNES 1 à 6: Titres
#         # ═══════════════════════════════════════════════════════════════════
        
#         ws.merge_cells('A1:E1')
#         ws['A1'] = "SONATRACH - ACTIVITE AMONT - DIVISION PRODUCTION"
#         ws['A1'].font = title_font
#         ws['A1'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A2:E2')
#         ws['A2'] = "PLAN ANNUEL 2015 ET PMT 2015/2019"
#         ws['A2'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A3:E3')
#         ws['A3'] = "ECHEANCIER DETAILLE PAR PROJETS"
#         ws['A3'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A4:E4')
#         ws['A4'] = "EXPLICATION DES VARIATIONS DES COÛTS"
#         ws['A4'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A5:E5')
#         ws['A5'] = "DEGLOBALISATION DE LA PREVISION ANNUELLE 2015"
#         ws['A5'].alignment = Alignment(horizontal='center')
        
#         ws.merge_cells('A6:E6')
#         ws['A6'] = "Unité : Millier DA"
#         ws['A6'].font = Font(bold=True)
#         ws['A6'].alignment = Alignment(horizontal='center')
        
#         # ═══════════════════════════════════════════════════════════════════
#         # LIGNE 7: En-têtes principaux
#         # ═══════════════════════════════════════════════════════════════════
        
#         header_fusions = [
#             ('A', 'A'),    # Activité 
#             ('B', 'B'),    # Région
#             ('C', 'C'),    # PERIMETRE
#             ('D', 'D'),    # Famille
#             ('E', 'E'),    # N°:Cpte Analy.
#             ('F', 'F'),    # Libellés
#             ('G', 'H'),    # Coût Global Initial PMT 2015/2019
#             ('I', 'J'),    # Réalisations Cumulées à fin 2013 au coût réel
#             ('K', 'L'),    # Réal. 1er Semestre
#             ('M', 'N'),    # Prév. 2è Semestre
#             ('O', 'P'),    # Prévisions de clôture 2014
#             ('Q', 'R'),    # Prévisions 2015
#             ('S', 'T'),    # Reste à Réaliser 2016/2019
#             ('U', 'V'),    # Prévisions 2016
#             ('W', 'X'),    # Prévisions 2017
#             ('Y', 'Z'),    # Prévisions 2018
#             ('AA', 'AB'),  # Prévisions 2019
#             ('AC', 'AD'),  # Janvier
#             ('AE', 'AF'),  # Février
#             ('AG', 'AH'),  # Mars
#             ('AI', 'AJ'),  # Avril
#             ('AK', 'AL'),  # Mai
#             ('AM', 'AN'),  # Juin
#             ('AO', 'AP'),  # Juillet
#             ('AQ', 'AR'),  # Août
#             ('AS', 'AT'),  # Septembre
#             ('AU', 'AV'),  # Octobre
#             ('AW', 'AX'),  # Novembre
#             ('AY', 'AZ'),  # Décembre
#         ]
        
#         header_texts = [
#             "Activité / Région / PERIMETRE / Famille",
#             "N°:Cpte Analy.",
#             "Libellés",
#             "Coût Global Initial PMT 2015/2019",
#             "Réalisations Cumulées à fin 2013 au coût réel",
#             "Réal. 1er Semestre",
#             "Prév. 2è Semestre",
#             "Prévisions de clôture 2014",
#             "Prévisions 2015",
#             "Reste à Réaliser 2016/2019",
#             "Prévisions 2016",
#             "Prévisions 2017",
#             "Prévisions 2018",
#             "Prévisions 2019",
#             "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
#             "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
#         ]
        
#         row_num = 7
#         for idx, (start_col, end_col) in enumerate(header_fusions):
#             if start_col != end_col:
#                 ws.merge_cells(f'{start_col}{row_num}:{end_col}{row_num}')
#             cell = ws[f'{start_col}{row_num}']
#             cell.value = header_texts[idx] if idx < len(header_texts) else ""
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
#         # ═══════════════════════════════════════════════════════════════════
#         # LIGNE 8: Sous-en-têtes "Total" / "dont Dev."
#         # ═══════════════════════════════════════════════════════════════════
        
#         row_num = 8
        
#         # Colonnes A à F: fusionnées (pas de sous-en-tête)
#         for col in ['A', 'B', 'C', 'D', 'E', 'F']:
#             ws.merge_cells(f'{col}{row_num}:{col}{row_num+1}')
        
#         # Colonnes avec Total / dont Dev. (G à AB)
#         dex_columns = [
#             ('G', 'H'), ('I', 'J'), ('K', 'L'), ('M', 'N'), ('O', 'P'),
#             ('Q', 'R'), ('S', 'T'), ('U', 'V'), ('W', 'X'), ('Y', 'Z'),
#             ('AA', 'AB')
#         ]
        
#         for start_col, end_col in dex_columns:
#             ws[f'{start_col}{row_num}'] = "Total"
#             ws[f'{end_col}{row_num}'] = "dont Dev."
#             ws[f'{start_col}{row_num}'].font = Font(bold=True, size=10)
#             ws[f'{end_col}{row_num}'].font = Font(bold=True, size=10)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{start_col}{row_num}'].fill = subheader_fill
#             ws[f'{end_col}{row_num}'].fill = subheader_fill
        
#         # Colonnes des mois (AC à AZ)
#         mois_pairs = [
#             ('AC', 'AD'), ('AE', 'AF'), ('AG', 'AH'), ('AI', 'AJ'),
#             ('AK', 'AL'), ('AM', 'AN'), ('AO', 'AP'), ('AQ', 'AR'),
#             ('AS', 'AT'), ('AU', 'AV'), ('AW', 'AX'), ('AY', 'AZ')
#         ]
        
#         for start_col, end_col in mois_pairs:
#             ws[f'{start_col}{row_num}'] = "Total"
#             ws[f'{end_col}{row_num}'] = "dont Dev."
#             ws[f'{start_col}{row_num}'].font = Font(bold=True, size=9)
#             ws[f'{end_col}{row_num}'].font = Font(bold=True, size=9)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
#             ws[f'{start_col}{row_num}'].fill = subheader_fill
#             ws[f'{end_col}{row_num}'].fill = subheader_fill
        
#         # ═══════════════════════════════════════════════════════════════════
#         # LIGNE 9: Sous-en-têtes "Prévisions" pour les mois
#         # ═══════════════════════════════════════════════════════════════════
        
#         row_num = 9
#         for start_col, end_col in mois_pairs:
#             ws[f'{start_col}{row_num}'] = "Prévisions"
#             ws[f'{end_col}{row_num}'] = ""
#             ws[f'{start_col}{row_num}'].font = Font(size=8)
#             ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
        
#         # Ajuster les hauteurs
#         for r in range(1, 10):
#             ws.row_dimensions[r].height = 30
#         ws.row_dimensions[7].height = 40
    
#     def _write_template_data_rows(self, ws, qs):
#         """Écrit les données"""
        
#         row_num = 10
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )
        
#         for record in qs:
#             # Identifiants (colonnes A-F)
#             activite = record.activite or ''
#             region = record.region or ''
#             perimetre = record.perm or ''
#             famille = record.famille or ''
#             code_analytique = record.code_division or ''
#             libelle = record.libelle or ''
            
#             # Données financières (colonnes G à AB)
#             cout_total = self._format_number(record.cout_initial_total)
#             cout_dex = self._format_number(record.cout_initial_dont_dex)
#             realisation_total = self._format_number(record.realisation_cumul_n_mins1_total)
#             realisation_dex = self._format_number(record.realisation_cumul_n_mins1_dont_dex)
#             real_s1_total = self._format_number(record.real_s1_n_total)
#             real_s1_dex = self._format_number(record.real_s1_n_dont_dex)
#             prev_s2_total = self._format_number(record.prev_s2_n_total)
#             prev_s2_dex = self._format_number(record.prev_s2_n_dont_dex)
#             prev_cloture_total = self._format_number(record.prev_cloture_n_total)
#             prev_cloture_dex = self._format_number(record.prev_cloture_n_dont_dex)
#             prev_n1_total = self._format_number(record.prev_n_plus1_total)
#             prev_n1_dex = self._format_number(record.prev_n_plus1_dont_dex)
#             rar_total = self._format_number(record.reste_a_realiser_total)
#             rar_dex = self._format_number(record.reste_a_realiser_dont_dex)
#             prev_n2_total = self._format_number(record.prev_n_plus2_total)
#             prev_n2_dex = self._format_number(record.prev_n_plus2_dont_dex)
#             prev_n3_total = self._format_number(record.prev_n_plus3_total)
#             prev_n3_dex = self._format_number(record.prev_n_plus3_dont_dex)
#             prev_n4_total = self._format_number(record.prev_n_plus4_total)
#             prev_n4_dex = self._format_number(record.prev_n_plus4_dont_dex)
#             prev_n5_total = self._format_number(record.prev_n_plus5_total)
#             prev_n5_dex = self._format_number(record.prev_n_plus5_dont_dex)
            
#             # Données mensuelles (colonnes AC à AZ)
#             mois_data = []
#             for mois in ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
#                         'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']:
#                 total = getattr(record, f'{mois}_total', None)
#                 dex = getattr(record, f'{mois}_dont_dex', None)
#                 mois_data.append((self._format_number(total), self._format_number(dex)))
            
#             # Construction de la ligne complète (colonnes A à AZ)
#             row_data = [
#                 # A-F: Identifiants
#                 activite, region, perimetre, famille, code_analytique, libelle,
                
#                 # G-H: Coût Global Initial
#                 cout_total, cout_dex,
                
#                 # I-J: Réalisations Cumulées
#                 realisation_total, realisation_dex,
                
#                 # K-L: Réal. 1er Semestre
#                 real_s1_total, real_s1_dex,
                
#                 # M-N: Prév. 2è Semestre
#                 prev_s2_total, prev_s2_dex,
                
#                 # O-P: Prévisions de clôture 2014
#                 prev_cloture_total, prev_cloture_dex,
                
#                 # Q-R: Prévisions 2015
#                 prev_n1_total, prev_n1_dex,
                
#                 # S-T: Reste à Réaliser
#                 rar_total, rar_dex,
                
#                 # U-V: Prévisions 2016
#                 prev_n2_total, prev_n2_dex,
                
#                 # W-X: Prévisions 2017
#                 prev_n3_total, prev_n3_dex,
                
#                 # Y-Z: Prévisions 2018
#                 prev_n4_total, prev_n4_dex,
                
#                 # AA-AB: Prévisions 2019
#                 prev_n5_total, prev_n5_dex,
#             ]
            
#             # Ajouter les 12 mois (AC à AZ)
#             for mois_total, mois_dex in mois_data:
#                 row_data.append(mois_total)
#                 row_data.append(mois_dex)
            
#             # Écrire la ligne
#             for col_idx, value in enumerate(row_data, start=1):
#                 col_letter = get_column_letter(col_idx)
#                 cell = ws[f'{col_letter}{row_num}']
#                 cell.value = value
#                 cell.border = thin_border
                
#                 if col_idx <= 6:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')
#                 else:
#                     cell.alignment = Alignment(horizontal='right', vertical='center')
#                     if value and isinstance(value, (int, float)):
#                         cell.number_format = '#,##0.00'
            
#             # Couleur de fond alternée
#             if row_num % 2 == 0:
#                 for col_letter in self._get_all_columns(52):
#                     try:
#                         ws[f'{col_letter}{row_num}'].fill = PatternFill(
#                             start_color="F2F2F6", end_color="F2F2F6", fill_type="solid"
#                         )
#                     except:
#                         pass
            
#             row_num += 1
    
#     def _write_recap_sheet(self, ws, qs, annee_cible):
#         """Écrit la feuille récapitulative"""
        
#         ws['A1'] = f"RÉCAPITULATIF - PROJETS VALIDÉS PAR LE DIVISIONNAIRE"
#         ws['A1'].font = Font(bold=True, size=14)
#         ws.merge_cells('A1:H1')
        
#         ws['A3'] = f"PMT: {annee_cible} - {annee_cible+4}"
#         ws['A4'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
#         ws['A5'] = f"Nombre de projets: {qs.count()}"
        
#         headers = ['Activité', 'Région', 'Périmètre', 'Famille', 
#                    'Code Division', 'Libellé', 'Coût Total', 'Dont DEX']
        
#         for col_idx, header in enumerate(headers, start=1):
#             cell = ws.cell(row=7, column=col_idx, value=header)
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center')
        
#         row_num = 8
#         total_general = 0
        
#         for record in qs:
#             cout_total = float(record.cout_initial_total or 0)
#             total_general += cout_total
            
#             ws.cell(row=row_num, column=1, value=record.activite or '')
#             ws.cell(row=row_num, column=2, value=record.region or '')
#             ws.cell(row=row_num, column=3, value=record.perm or '')
#             ws.cell(row=row_num, column=4, value=record.famille or '')
#             ws.cell(row=row_num, column=5, value=record.code_division or '')
#             ws.cell(row=row_num, column=6, value=record.libelle or '')
            
#             cell_total = ws.cell(row=row_num, column=7, value=cout_total)
#             cell_total.number_format = '#,##0.00'
#             cell_dex = ws.cell(row=row_num, column=8, value=float(record.cout_initial_dont_dex or 0))
#             cell_dex.number_format = '#,##0.00'
            
#             row_num += 1
        
#         ws.cell(row=row_num, column=6, value="TOTAL:")
#         ws.cell(row=row_num, column=6).font = Font(bold=True)
#         cell_total = ws.cell(row=row_num, column=7, value=total_general)
#         cell_total.font = Font(bold=True)
#         cell_total.number_format = '#,##0.00'
        
#         for col in range(1, 9):
#             ws.column_dimensions[get_column_letter(col)].width = 20
    
#     def _apply_template_column_widths(self, ws):
#         """Applique les largeurs de colonnes"""
#         column_widths = {
#             'A': 15, 'B': 15, 'C': 25, 'D': 12, 'E': 18, 'F': 45,
#             'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15,
#             'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 15, 'R': 15,
#             'S': 15, 'T': 15, 'U': 15, 'V': 15, 'W': 15, 'X': 15,
#             'Y': 15, 'Z': 15, 'AA': 15, 'AB': 15,
#         }
        
#         for i in range(29, 53):
#             col_letter = get_column_letter(i)
#             column_widths[col_letter] = 12
        
#         for col_letter, width in column_widths.items():
#             try:
#                 ws.column_dimensions[col_letter].width = width
#             except:
#                 pass
    
#     def _get_all_columns(self, max_col):
#         """Retourne la liste des lettres de colonnes jusqu'à max_col"""
#         columns = []
#         for i in range(1, max_col + 1):
#             columns.append(get_column_letter(i))
#         return columns
    
#     def _format_number(self, value):
#         if value is None:
#             return 0
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return 0
# ================================================================== #
# EXPORT EXCEL - FORMAT IDENTIQUE AU TEMPLATE recap.xls
# ================================================================== #

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import requests

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from .models import BudgetRecord
from .remote_auth import RemoteJWTAuthentication
from .mappings import ACTIVITE_MAPPING


class ExportProjetsValidesDivisionnaireView(APIView):
    """
    GET /recap/budget/export/valides-divisionnaire/
    
    Exporte les projets validés par le divisionnaire 
    avec filtre optionnel par année (?annee=2024)
    """
    authentication_classes = [RemoteJWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Récupérer l'année depuis les paramètres GET
        annee_param = request.query_params.get('annee')
        
        if annee_param:
            annee_cible = int(annee_param)
        else:
            annee_cible = datetime.now().year + 1
        
        print(f"[DEBUG] Export - Année cible: {annee_cible}")
        
        # Récupérer les projets (sans filtre is_active)
        qs = BudgetRecord.objects.filter(
            statut_final='valide_divisionnaire',
            annee_debut_pmt=annee_cible
        ).order_by('region', 'code_division')
        
        print(f"[DEBUG] Projets trouvés: {qs.count()}")
        
        # Afficher les détails des projets trouvés
        for record in qs:
            print(f"[DEBUG] Projet: {record.code_division} - is_active: {record.is_active}")
        
        # Créer le classeur Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet("invest 2015-2019 Vers Stag")
        
        self._write_template_header(ws, annee_cible)
        self._write_template_data_rows(ws, qs)
        self._apply_template_column_widths(ws)
        
        ws_recap = wb.create_sheet("Récapitulatif")
        self._write_recap_sheet(ws_recap, qs, annee_cible)
        
        filename = f"projets_valides_divisionnaire_{annee_cible}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def _write_template_header(self, ws, annee_cible):
        """Écrit les en-têtes exactement comme dans le template recap.xls"""
        
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # ═══════════════════════════════════════════════════════════════════
        # LIGNES 1 à 6: Titres
        # ═══════════════════════════════════════════════════════════════════
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "SONATRACH - ACTIVITE AMONT - DIVISION PRODUCTION"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = "PLAN ANNUEL 2015 ET PMT 2015/2019"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:F3')
        ws['A3'] = "ECHEANCIER DETAILLE PAR PROJETS"
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A4:F4')
        ws['A4'] = "EXPLICATION DES VARIATIONS DES COÛTS"
        ws['A4'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A5:F5')
        ws['A5'] = "DEGLOBALISATION DE LA PREVISION ANNUELLE 2015"
        ws['A5'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A6:F6')
        ws['A6'] = "Unité : Millier DA"
        ws['A6'].font = Font(bold=True)
        ws['A6'].alignment = Alignment(horizontal='center')
        
        # ═══════════════════════════════════════════════════════════════════
        # LIGNE 7: En-têtes principaux (colonnes séparées A,B,C,D,E,F)
        # ═══════════════════════════════════════════════════════════════════
        
        header_fusions = [
            ('A', 'A'),    # Activité 
            ('B', 'B'),    # Région
            ('C', 'C'),    # PERIMETRE
            ('D', 'D'),    # Famille
            ('E', 'E'),    # N°:Cpte Analy.
            ('F', 'F'),    # Libellés
            ('G', 'H'),    # Coût Global Initial PMT 2015/2019
            ('I', 'J'),    # Réalisations Cumulées à fin 2013 au coût réel
            ('K', 'L'),    # Réal. 1er Semestre
            ('M', 'N'),    # Prév. 2è Semestre
            ('O', 'P'),    # Prévisions de clôture 2014
            ('Q', 'R'),    # Prévisions 2015
            ('S', 'T'),    # Reste à Réaliser 2016/2019
            ('U', 'V'),    # Prévisions 2016
            ('W', 'X'),    # Prévisions 2017
            ('Y', 'Z'),    # Prévisions 2018
            ('AA', 'AB'),  # Prévisions 2019
            ('AC', 'AD'),  # Janvier
            ('AE', 'AF'),  # Février
            ('AG', 'AH'),  # Mars
            ('AI', 'AJ'),  # Avril
            ('AK', 'AL'),  # Mai
            ('AM', 'AN'),  # Juin
            ('AO', 'AP'),  # Juillet
            ('AQ', 'AR'),  # Août
            ('AS', 'AT'),  # Septembre
            ('AU', 'AV'),  # Octobre
            ('AW', 'AX'),  # Novembre
            ('AY', 'AZ'),  # Décembre
        ]
        
        header_texts = [
            "Activité",
            "Région",
            "PERIMETRE",
            "Famille",
            "N°:Cpte Analy.",
            "Libellés",
            "Coût Global Initial PMT 2015/2019",
            "Réalisations Cumulées à fin 2013 au coût réel",
            "Réal. 1er Semestre",
            "Prév. 2è Semestre",
            "Prévisions de clôture 2014",
            "Prévisions 2015",
            "Reste à Réaliser 2016/2019",
            "Prévisions 2016",
            "Prévisions 2017",
            "Prévisions 2018",
            "Prévisions 2019",
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]
        
        row_num = 7
        for idx, (start_col, end_col) in enumerate(header_fusions):
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{row_num}:{end_col}{row_num}')
            cell = ws[f'{start_col}{row_num}']
            cell.value = header_texts[idx] if idx < len(header_texts) else ""
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ═══════════════════════════════════════════════════════════════════
        # LIGNE 8: Sous-en-têtes "Total" / "dont Dev."
        # ═══════════════════════════════════════════════════════════════════
        
        row_num = 8
        
        # Colonnes A à F: simples (pas de fusion, juste une ligne)
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row_num}'] = ""
        
        # Colonnes avec Total / dont Dev. (G à AB)
        dex_columns = [
            ('G', 'H'), ('I', 'J'), ('K', 'L'), ('M', 'N'), ('O', 'P'),
            ('Q', 'R'), ('S', 'T'), ('U', 'V'), ('W', 'X'), ('Y', 'Z'),
            ('AA', 'AB')
        ]
        
        for start_col, end_col in dex_columns:
            ws[f'{start_col}{row_num}'] = "Total"
            ws[f'{end_col}{row_num}'] = "dont Dev."
            ws[f'{start_col}{row_num}'].font = Font(bold=True, size=10)
            ws[f'{end_col}{row_num}'].font = Font(bold=True, size=10)
            ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
            ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
            ws[f'{start_col}{row_num}'].fill = subheader_fill
            ws[f'{end_col}{row_num}'].fill = subheader_fill
        
        # Colonnes des mois (AC à AZ)
        mois_pairs = [
            ('AC', 'AD'), ('AE', 'AF'), ('AG', 'AH'), ('AI', 'AJ'),
            ('AK', 'AL'), ('AM', 'AN'), ('AO', 'AP'), ('AQ', 'AR'),
            ('AS', 'AT'), ('AU', 'AV'), ('AW', 'AX'), ('AY', 'AZ')
        ]
        
        for start_col, end_col in mois_pairs:
            ws[f'{start_col}{row_num}'] = "Total"
            ws[f'{end_col}{row_num}'] = "dont Dev."
            ws[f'{start_col}{row_num}'].font = Font(bold=True, size=9)
            ws[f'{end_col}{row_num}'].font = Font(bold=True, size=9)
            ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
            ws[f'{end_col}{row_num}'].alignment = Alignment(horizontal='center')
            ws[f'{start_col}{row_num}'].fill = subheader_fill
            ws[f'{end_col}{row_num}'].fill = subheader_fill
        
        # ═══════════════════════════════════════════════════════════════════
        # LIGNE 9: Sous-en-têtes "Prévisions" pour les mois
        # ═══════════════════════════════════════════════════════════════════
        
        row_num = 9
        for start_col, end_col in mois_pairs:
            ws[f'{start_col}{row_num}'] = "Prévisions"
            ws[f'{end_col}{row_num}'] = ""
            ws[f'{start_col}{row_num}'].font = Font(size=8)
            ws[f'{start_col}{row_num}'].alignment = Alignment(horizontal='center')
        
        # Ajuster les hauteurs
        for r in range(1, 10):
            ws.row_dimensions[r].height = 30
        ws.row_dimensions[7].height = 40
    
    def _write_template_data_rows(self, ws, qs):
        """Écrit les données"""
        
        row_num = 10
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for record in qs:
            # Identifiants (colonnes A-F) - une valeur par colonne
            activite = record.activite or ''
            region = record.region or ''
            perimetre = record.perm or ''
            famille = record.famille or ''
            code_analytique = record.code_division or ''
            libelle = record.libelle or ''
            
            # Données financières (colonnes G à AB)
            cout_total = self._format_number(record.cout_initial_total)
            cout_dex = self._format_number(record.cout_initial_dont_dex)
            realisation_total = self._format_number(record.realisation_cumul_n_mins1_total)
            realisation_dex = self._format_number(record.realisation_cumul_n_mins1_dont_dex)
            real_s1_total = self._format_number(record.real_s1_n_total)
            real_s1_dex = self._format_number(record.real_s1_n_dont_dex)
            prev_s2_total = self._format_number(record.prev_s2_n_total)
            prev_s2_dex = self._format_number(record.prev_s2_n_dont_dex)
            prev_cloture_total = self._format_number(record.prev_cloture_n_total)
            prev_cloture_dex = self._format_number(record.prev_cloture_n_dont_dex)
            prev_n1_total = self._format_number(record.prev_n_plus1_total)
            prev_n1_dex = self._format_number(record.prev_n_plus1_dont_dex)
            rar_total = self._format_number(record.reste_a_realiser_total)
            rar_dex = self._format_number(record.reste_a_realiser_dont_dex)
            prev_n2_total = self._format_number(record.prev_n_plus2_total)
            prev_n2_dex = self._format_number(record.prev_n_plus2_dont_dex)
            prev_n3_total = self._format_number(record.prev_n_plus3_total)
            prev_n3_dex = self._format_number(record.prev_n_plus3_dont_dex)
            prev_n4_total = self._format_number(record.prev_n_plus4_total)
            prev_n4_dex = self._format_number(record.prev_n_plus4_dont_dex)
            prev_n5_total = self._format_number(record.prev_n_plus5_total)
            prev_n5_dex = self._format_number(record.prev_n_plus5_dont_dex)
            
            # Données mensuelles (colonnes AC à AZ)
            mois_data = []
            for mois in ['janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
                        'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre']:
                total = getattr(record, f'{mois}_total', None)
                dex = getattr(record, f'{mois}_dont_dex', None)
                mois_data.append((self._format_number(total), self._format_number(dex)))
            
            # Construction de la ligne complète (colonnes A à AZ)
            row_data = [
                # A-F: Identifiants (une colonne par valeur)
                activite,      # A
                region,        # B
                perimetre,     # C
                famille,       # D
                code_analytique, # E
                libelle,       # F
                
                # G-H: Coût Global Initial
                cout_total, cout_dex,
                
                # I-J: Réalisations Cumulées
                realisation_total, realisation_dex,
                
                # K-L: Réal. 1er Semestre
                real_s1_total, real_s1_dex,
                
                # M-N: Prév. 2è Semestre
                prev_s2_total, prev_s2_dex,
                
                # O-P: Prévisions de clôture 2014
                prev_cloture_total, prev_cloture_dex,
                
                # Q-R: Prévisions 2015
                prev_n1_total, prev_n1_dex,
                
                # S-T: Reste à Réaliser
                rar_total, rar_dex,
                
                # U-V: Prévisions 2016
                prev_n2_total, prev_n2_dex,
                
                # W-X: Prévisions 2017
                prev_n3_total, prev_n3_dex,
                
                # Y-Z: Prévisions 2018
                prev_n4_total, prev_n4_dex,
                
                # AA-AB: Prévisions 2019
                prev_n5_total, prev_n5_dex,
            ]
            
            # Ajouter les 12 mois (AC à AZ)
            for mois_total, mois_dex in mois_data:
                row_data.append(mois_total)
                row_data.append(mois_dex)
            
            # Écrire la ligne
            for col_idx, value in enumerate(row_data, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f'{col_letter}{row_num}']
                cell.value = value
                cell.border = thin_border
                
                if col_idx <= 6:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
            
            row_num += 1
    
    def _write_recap_sheet(self, ws, qs, annee_cible):
        """Écrit la feuille récapitulative"""
        
        ws['A1'] = f"RÉCAPITULATIF - PROJETS VALIDÉS PAR LE DIVISIONNAIRE"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        ws['A3'] = f"PMT: {annee_cible} - {annee_cible+4}"
        ws['A4'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A5'] = f"Nombre de projets: {qs.count()}"
        
        headers = ['Activité', 'Région', 'Périmètre', 'Famille', 
                   'Code Division', 'Libellé', 'Coût Total', 'Dont DEX']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row_num = 8
        total_general = 0
        
        for record in qs:
            cout_total = float(record.cout_initial_total or 0)
            total_general += cout_total
            
            ws.cell(row=row_num, column=1, value=record.activite or '')
            ws.cell(row=row_num, column=2, value=record.region or '')
            ws.cell(row=row_num, column=3, value=record.perm or '')
            ws.cell(row=row_num, column=4, value=record.famille or '')
            ws.cell(row=row_num, column=5, value=record.code_division or '')
            ws.cell(row=row_num, column=6, value=record.libelle or '')
            
            cell_total = ws.cell(row=row_num, column=7, value=cout_total)
            cell_total.number_format = '#,##0.00'
            cell_dex = ws.cell(row=row_num, column=8, value=float(record.cout_initial_dont_dex or 0))
            cell_dex.number_format = '#,##0.00'
            
            row_num += 1
        
        ws.cell(row=row_num, column=6, value="TOTAL:")
        ws.cell(row=row_num, column=6).font = Font(bold=True)
        cell_total = ws.cell(row=row_num, column=7, value=total_general)
        cell_total.font = Font(bold=True)
        cell_total.number_format = '#,##0.00'
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _apply_template_column_widths(self, ws):
        """Applique les largeurs de colonnes"""
        column_widths = {
            'A': 12,   # Activité
            'B': 15,   # Région
            'C': 25,   # PERIMETRE
            'D': 12,   # Famille
            'E': 18,   # N°:Cpte Analy.
            'F': 45,   # Libellés
            'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15,
            'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 15, 'R': 15,
            'S': 15, 'T': 15, 'U': 15, 'V': 15, 'W': 15, 'X': 15,
            'Y': 15, 'Z': 15, 'AA': 15, 'AB': 15,
        }
        
        for i in range(29, 53):
            col_letter = get_column_letter(i)
            column_widths[col_letter] = 12
        
        for col_letter, width in column_widths.items():
            try:
                ws.column_dimensions[col_letter].width = width
            except:
                pass
    
    def _format_number(self, value):
        if value is None:
            return 0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0